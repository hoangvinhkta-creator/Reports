"""Hợp đồng Demo V1: production thật, không bỏ dòng, không đoán giá."""

import json
import subprocess
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from app import demo
from app.modules.exporting.excel_exporter import ReportIntegrityError, export_report
from app.modules.importing.raw_reader import read_raw_rows
from tests.fixtures.synthetic_workbook import HEADER
from tests.test_105e_price_composition import write_catalog_capture, write_history_capture
from tests.test_tracking_history_reader import build_export, event


def write_sales(path, rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in (["SỔ CHI TIẾT BÁN HÀNG"], [], [], HEADER, []):
        sheet.append(row)
    for order, product, when, quantity, sell_price in rows:
        sheet.append([when, order, None, product, None, None, None, None,
                      quantity, sell_price, None, 0, "Vũ Hạnh Ly", None, None, None, None])
        sheet.cell(sheet.max_row, 4).data_type = "s"
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def inputs(tmp_path):
    history = write_history_capture(tmp_path, build_export(
        prices={"A1": 7000, "B1": 5000},
        events={"B1": {"E1": event(prev=5000, nxt=None,
                                   at=datetime(2026, 9, 2, tzinfo=timezone.utc))}},
    ))
    catalog = write_catalog_capture(tmp_path, [
        {"tracking_code": code, "name": code, "alt": [], "present_in_board": True}
        for code in ("A1", "B1")
    ])
    sales = write_sales(tmp_path / "sales.xlsx", [
        ("AUTO-1", "A1", date(2026, 9, 5), 1, 8_000_000),
        ("MIXED", "A1", date(2026, 9, 5), 2, 8_000_000),
        ("MIXED", "B1", date(2026, 9, 5), 1, 6_000_000),
        ("UNDATED", "A1", None, 1, 8_000_000),
        ("AUTO-1", "A1", date(2026, 9, 5), 3, 8_000_000),
    ])
    return dict(sales=sales, tracking_capture=history, tracking_catalog=catalog,
                output=tmp_path / "report.xlsx")


def values(sheet):
    return list(sheet.iter_rows(min_row=2, values_only=True))


def test_full_production_export_preserves_mixed_lines_and_duplicate_record_keys(inputs):
    run = demo.run_demo(**inputs)
    assert run.summary.input_orders == run.summary.accounted_orders == 3
    assert run.summary.total_lines == 5
    assert run.summary.auto_orders == 1
    assert run.summary.review_orders == 2
    assert run.summary.review_lines == 2
    assert run.summary.order_accounting_rate == 1
    assert len(run.price_records) == 5
    assert all(r.evidence.public_purchase_version_id is None for r in run.price_records)

    workbook = openpyxl.load_workbook(inputs["output"], data_only=True)
    # PHB-01 — sheet "Chưa định danh" là bản xuất cho Owner mang sang màn
    # phân loại theo tên hàng của Tracking. Nó LUÔN có mặt, kể cả khi rỗng.
    assert workbook.sheetnames == [
        "Summary", "Order Lines", "Review Queue", "Chưa định danh",
    ]
    rows = values(workbook["Order Lines"])
    assert [r[12] for r in rows] == [6, 7, 8, 9, 10]
    assert [r[9] for r in rows] == ["AUTO", "AUTO", "PENDING", "PENDING", "AUTO"]
    assert rows[1][13] == "REVIEW_QUEUE"
    assert rows[1][6:9] == (7_000_000, 2_000_000, 2_000_000)
    assert rows[2][6:9] == (None, None, None)
    assert rows[4][6:9] == (7_000_000, 3_000_000, 3_000_000)
    review = {r[9]: r for r in values(workbook["Review Queue"]) if r[9]}
    assert set(review) == {8, 9}
    record = next(r for r in run.price_records if r.raw_product_identity == "B1")
    assert record.reason.value in review[8][5]
    assert record.detail in review[8][6]
    assert review[8][10:12] == ("TRACKING", "B1")
    assert review[8][18] == record.tracking_reconstruction.reason.value
    assert review[8][15] == record.evidence.tracking_price_history_capture_id
    assert workbook["Summary"]["B12"].value == 6_000_000
    assert workbook["Order Lines"]["G3"].data_type == "n"
    assert workbook["Order Lines"].freeze_panes == "E2"
    workbook.close()


def test_capture_before_sale_is_truthful_pending(inputs):
    path = inputs["tracking_capture"]
    data = json.loads(path.read_text())
    data["captured_at"] = "2026-08-31T08:00:38+00:00"
    path.write_text(json.dumps(data))
    run = demo.run_demo(**inputs)
    assert run.summary.auto_orders == 0
    assert run.summary.review_orders == 3
    assert all(line.accounting_purchase_price is None
               for order in run.result.orders for line in order.lines)
    workbook = openpyxl.load_workbook(inputs["output"], data_only=True)
    assert workbook["Summary"]["B12"].value is None
    workbook.close()


def test_legacy_pp_is_never_loaded(inputs, monkeypatch):
    from app.modules.pricing.resolution import sources

    def forbidden(*args, **kwargs):
        pytest.fail("CLI không được đọc PP YAML hoặc default source loader")

    monkeypatch.setattr(sources, "load_public_purchase_source", forbidden)
    monkeypatch.setattr(sources, "load_price_resolution_sources", forbidden)
    assert demo.run_demo(**inputs).summary.auto_orders == 1


@pytest.mark.parametrize("damage", ["missing_line", "missing_order", "missing_record"])
def test_missing_results_cannot_claim_success(inputs, tmp_path, damage):
    run = demo.run_demo(**inputs)
    records = run.price_records
    if damage == "missing_line":
        run.result.orders[0].lines.pop()
    elif damage == "missing_order":
        run.result.orders.pop()
    else:
        records = records[:-1]
    output = tmp_path / "damaged.xlsx"
    with pytest.raises(ReportIntegrityError):
        export_report(run.result, records, read_raw_rows(inputs["sales"]),
                      sales_path=inputs["sales"], tracking_capture=inputs["tracking_capture"],
                      tracking_catalog=inputs["tracking_catalog"], output_path=output,
                      processed_at=datetime.now(timezone.utc))
    assert not output.exists()


def test_cli_from_another_directory_and_no_overwrite(inputs, tmp_path):
    command = [sys.executable, str(Path(demo.__file__).resolve())]
    for key, value in inputs.items():
        command.extend(["--" + key.replace("_", "-"), str(value)])
    result = subprocess.run(command, cwd=tmp_path, text=True, capture_output=True)
    assert result.returncode == 0, result.stderr
    assert "DEMO_COMPLETE\n" in result.stdout
    assert "ORDERS=3\nAUTO=1\nREVIEW_QUEUE=2" in result.stdout
    original = inputs["output"].read_bytes()
    result = subprocess.run(command, cwd=tmp_path, text=True, capture_output=True)
    assert result.returncode == 1
    assert "DEMO_COMPLETE" not in result.stdout
    assert inputs["output"].read_bytes() == original


@pytest.mark.parametrize("bad_input", ["missing", "malformed", "failed"])
def test_bad_capture_fails_without_payload_or_partial_report(inputs, capsys, bad_input):
    path = inputs["tracking_capture"]
    if bad_input == "missing":
        path.unlink()
    elif bad_input == "malformed":
        path.write_text('{"secret": "DO_NOT_PRINT_THIS"}')
    else:
        data = json.loads(path.read_text())
        data.update(capture_status="FAILED", failure_reason="DO_NOT_PRINT_THIS")
        path.write_text(json.dumps(data))
    args = []
    for key, value in inputs.items():
        args.extend(["--" + key.replace("_", "-"), str(value)])
    assert demo.main(args) == 1
    captured = capsys.readouterr()
    assert "DEMO_FAILED" in captured.err
    assert "DO_NOT_PRINT_THIS" not in captured.err
    assert "Traceback" not in captured.err
    assert not inputs["output"].exists()


def test_pre_cutover_golden_prices_are_not_relabelled_as_tracking(inputs):
    inputs["sales"] = Path("tests/fixtures/golden/period_2026_01.xlsx")
    run = demo.run_demo(**inputs)
    assert run.summary.input_orders == run.summary.accounted_orders
    assert run.summary.order_accounting_rate == 1
    # Các miss pre-cutover được hỏi qua composition, nhưng sale trước baseline
    # vẫn Pending; entry registry CONFIRMED không bị relabel/ghi đè.
    assert run.price_records
    assert all(record.price_vnd is None for record in run.price_records)
    workbook = openpyxl.load_workbook(inputs["output"], data_only=True)
    rows = values(workbook["Order Lines"])
    golden = next(r for r in rows if r[1] == "BH62063")
    assert golden[6] == Decimal("7000000")
    assert golden[8] == Decimal("500000")
    assert golden[11] != "TRACKING_PRICE_HISTORY"
    pending = next(r for r in rows if r[6] is None)
    assert "Missing.PurchasePrice" in pending[10]
    assert len(rows) == len(read_raw_rows(inputs["sales"]))
    workbook.close()


def test_source_text_is_not_an_excel_formula_and_inputs_stay_unchanged(inputs):
    write_sales(inputs["sales"], [
        ("TEXT", "=1+1", date(2026, 9, 5), 1, 8_000_000),
    ])
    originals = {key: inputs[key].read_bytes() for key in
                 ("sales", "tracking_capture", "tracking_catalog")}
    run = demo.run_demo(**inputs)
    assert run.summary.review_orders == 1
    for key, original in originals.items():
        assert inputs[key].read_bytes() == original
    workbook = openpyxl.load_workbook(inputs["output"])
    assert workbook["Order Lines"]["D2"].value == "=1+1"
    assert workbook["Order Lines"]["D2"].data_type == "s"
    assert workbook["Review Queue"]["D2"].data_type == "s"
    workbook.close()


def test_kpi_unavailable_is_queued_even_with_resolved_price(inputs, monkeypatch):
    from app import composition
    from app.modules.kpi.kpi_profit_engine import AUTHORITY_UNAVAILABLE

    monkeypatch.setattr(composition, "load_eligible_costs_authority",
                        lambda path: AUTHORITY_UNAVAILABLE)
    run = demo.run_demo(**inputs)
    assert run.summary.auto_orders == 0
    assert run.summary.review_orders == 3
    workbook = openpyxl.load_workbook(inputs["output"], data_only=True)
    row = values(workbook["Order Lines"])[0]
    assert row[6] == 7_000_000
    assert row[8] is None
    assert "Pending.eligible_kpi_profit" in row[10]
    workbook.close()
