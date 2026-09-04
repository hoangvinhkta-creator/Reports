"""Importer legacy (TASK-PRA-001.2): fidelity, KHÔNG tính lại, annotate lỗi.

Đây là nhóm test giữ ranh giới chấp nhận cứng của task: giá trị trong workbook
cũ đi vào bản ghi NGUYÊN TRẠNG. Công cụ được phép chỉ ra chỗ công thức tự mâu
thuẫn, nhưng KHÔNG được sửa con số — kể cả khi biết công thức cũ sai.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.legacy import LegacyImportError, parse_workbook
from app.legacy.models import SUMMARY_COLUMN_FIELDS

REPO_ROOT = Path(__file__).resolve().parent.parent


@pytest.fixture
def workbook(legacy_workbook_path):
    return parse_workbook(legacy_workbook_path)


def _row(workbook, sheet_name: str, sheet_row: int):
    return next(
        row for row in workbook.summary_rows
        if row.sheet_name == sheet_name and row.sheet_row == sheet_row
    )


# --- Fidelity -------------------------------------------------------------

def test_every_imported_summary_cell_equals_the_excel_cell(workbook, legacy_workbook_path):
    """Đối chiếu từng ô: đọc lại Excel bằng openpyxl data_only và so với bản ghi."""
    sheets = load_workbook(legacy_workbook_path, data_only=True)
    matched = mismatched = 0
    for row in workbook.summary_rows:
        sheet = sheets[row.sheet_name]
        for column, field in SUMMARY_COLUMN_FIELDS.items():
            source = sheet[f"{column}{row.sheet_row}"].value
            expected = None if not isinstance(source, (int, float)) else Decimal(str(source))
            if row.values[field] == expected:
                matched += 1
            else:
                mismatched += 1
    assert (matched, mismatched) == (matched, 0)
    assert matched > 0


def test_every_imported_daily_cell_equals_the_excel_cell(workbook, legacy_workbook_path):
    sheet = load_workbook(legacy_workbook_path, data_only=True)["DataChart 2026"]
    for item in workbook.daily_sales:
        cell = sheet.cell(row=2 + item.month, column=1 + item.day).value
        assert item.sales_vnd == Decimal(str(cell))


def test_a_non_integer_source_value_keeps_its_exact_decimal(workbook):
    """87,6 sản phẩm là một con số SAI về nghiệp vụ — và vẫn phải được giữ đúng."""
    assert _row(workbook, "Summary 2026", 5).values["products"] == Decimal("87.6")


def test_units_are_never_converted_between_sheets(workbook):
    """Summary lưu nghìn đồng, DataChart lưu VND nguyên — không quy đổi."""
    assert _row(workbook, "Summary 2026", 4).unit == "kVND"
    january = next(item for item in workbook.daily_sales if item.month == 1 and item.day == 1)
    assert january.sales_vnd == Decimal("820000000")


# --- Không tính lại -------------------------------------------------------

def test_a_value_that_contradicts_its_own_formula_is_stored_verbatim(workbook):
    """Ô F của NV-B tháng 02 mang giá trị 999 nhưng công thức là ``=G9/5.5%``.

    Importer phải lưu 999. Nếu ở đâu đó nó "sửa cho đúng công thức", con số
    lịch sử của Owner sẽ bị thay bằng con số do code nghĩ ra.
    """
    row = _row(workbook, "Summary 2026", 9)
    assert row.values["converted_revenue"] == Decimal("999")
    assert row.formula_text["F"] == "=G9/5.5%"


def test_month_total_is_stored_as_written_not_resummed(workbook):
    """Dòng tổng tháng 01 lưu đúng số Excel đã ghi, kể cả khi nó cộng thiếu."""
    total = _row(workbook, "Summary 2026", 7)
    sellers = [_row(workbook, "Summary 2026", index) for index in (4, 5, 6)]
    assert total.values["converted_revenue"] == Decimal("1152000")
    assert total.values["converted_revenue"] != sum(
        row.values["converted_revenue"] for row in sellers
    )


def _code_without_string_literals(path: Path) -> str:
    """Mã nguồn với mọi chuỗi và chú thích bị xoá — chỉ còn phần LOGIC."""
    import io
    import tokenize

    pieces = []
    with open(path, "rb") as handle:
        for token in tokenize.tokenize(handle.readline):
            if token.type in (tokenize.STRING, tokenize.COMMENT):
                continue
            pieces.append(token.string)
    return " ".join(pieces)


def test_importer_never_divides_or_multiplies_anything():
    """Một phép chia/nhân trong importer là một con số mới — điều bị cấm.

    Phép TRỪ chỉ số dòng (so độ rộng hai khoảng SUM khi phát hiện A2) được
    phép: nó so sánh cấu trúc công thức, không đụng vào giá trị nghiệp vụ.
    """
    import ast

    for path in (REPO_ROOT / "app/legacy").rglob("*.py"):
        for node in ast.walk(ast.parse(path.read_text(encoding="utf-8"))):
            if isinstance(node, ast.BinOp) and isinstance(
                node.op, (ast.Div, ast.Mult, ast.FloorDiv)
            ):
                pytest.fail(f"{path.name}:{node.lineno} — phép chia/nhân trong importer")


def test_recalculation_tokens_appear_only_inside_matching_strings():
    """CHECK-PRA001-02: `/2`, `/ 2`, `5.5%` không được nằm trong logic."""
    for path in (REPO_ROOT / "app/legacy").rglob("*.py"):
        code = _code_without_string_literals(path)
        for token in ("/2", "/ 2", "5.5%"):
            assert token not in code, f"{path.name}: {token!r} nằm trong logic"


def test_formula_text_is_preserved_for_audit(workbook):
    row = _row(workbook, "Summary 2026", 4)
    assert row.formula_text["C"] == "='01.2026 NV-A'!$B$1"
    assert row.formula_text["I"] == "=F4/1571182"


# --- Lỗi đã biết A1 / A2 / A4 / A6 ---------------------------------------

def test_a1_flags_a_non_integer_product_count_without_changing_it(workbook):
    row = _row(workbook, "Summary 2026", 5)
    assert row.known_defects["D"] == ["A1"]
    assert row.values["products"] == Decimal("87.6")


def test_a2_flags_a_month_total_column_that_sums_a_narrower_range(workbook):
    row = _row(workbook, "Summary 2026", 7)
    assert row.known_defects == {"F": ["A2"]}
    assert row.formula_text["F"] == "=SUM(F4:F5)"
    assert row.formula_text["E"] == "=SUM(E4:E6)"


def test_a4_flags_a_cell_reading_another_sellers_sheet(workbook):
    row = _row(workbook, "Summary 2026", 10)
    assert row.seller_label == "Kênh-1"
    assert "A4" in row.known_defects["D"]
    assert row.formula_text["D"] == "='02.2026 NV-A'!$E$1"


def test_a6_flags_a_hardcoded_previous_period_denominator(workbook):
    assert _row(workbook, "Summary 2026", 4).known_defects["I"] == ["A6"]


def test_a6_does_not_fire_when_the_previous_period_is_a_real_cell(workbook):
    assert "I" not in _row(workbook, "Summary 2026", 8).known_defects


def test_a_conversion_rate_divisor_is_not_mistaken_for_a_defect(workbook):
    """``=G4/5.5%`` là tỉ lệ quy đổi hợp lệ, không phải mẫu số gõ tay."""
    assert "F" not in _row(workbook, "Summary 2026", 4).known_defects


def test_clean_rows_carry_no_defect_codes(workbook):
    assert _row(workbook, "Summary 2026", 13).known_defects == {}


# --- Phân loại dòng và kỳ -------------------------------------------------

def test_row_kinds_follow_formula_structure_not_fixed_offsets(workbook):
    kinds = {
        (row.sheet_name, row.sheet_row): row.row_kind for row in workbook.summary_rows
    }
    assert kinds[("Summary 2026", 3)] == "YEAR_TOTAL"
    assert kinds[("Summary 2026", 4)] == "SELLER"
    assert kinds[("Summary 2026", 7)] == "MONTH_TOTAL"
    # Khối tháng 02 có 4 người bán, khối tháng 01 có 3 — offset cứng sẽ sai.
    assert kinds[("Summary 2026", 11)] == "SELLER"
    assert kinds[("Summary 2026", 12)] == "MONTH_TOTAL"


def test_period_comes_from_the_referenced_source_sheet(workbook):
    assert (_row(workbook, "Summary 2026", 4).year, _row(workbook, "Summary 2026", 4).month) == (2026, 1)
    assert (_row(workbook, "Summary 2026", 8).year, _row(workbook, "Summary 2026", 8).month) == (2026, 2)


def test_year_total_row_is_not_attached_to_any_month(workbook):
    row = _row(workbook, "Summary 2026", 3)
    assert (row.year, row.month) == (2026, None)


def test_seller_label_is_kept_verbatim(workbook):
    assert [row.seller_label for row in workbook.summary_rows
            if row.sheet_name == "Summary 2026" and row.row_kind == "SELLER"][:3] == [
        "NV-A", "NV-B", "Kênh-1"]


# --- DataChart ------------------------------------------------------------

def test_empty_days_are_absent_rather_than_stored_as_zero(workbook):
    """Ngày chưa có số ≠ doanh số 0 — lưu 0 sẽ là bịa một sự kiện kinh doanh."""
    assert {(item.month, item.day) for item in workbook.daily_sales} == {
        (1, 1), (1, 2), (1, 3), (2, 1), (2, 2), (3, 1)}


def test_monthly_reference_keeps_target_cells_and_leaves_unknowns_null(workbook):
    january = workbook.monthly_reference[0]
    assert january.sales_current_year_vnd == Decimal("2780000000")
    assert january.target_year == Decimal("345474000")
    assert january.average_per_day is None


# --- Hình dạng workbook ---------------------------------------------------

def test_a_workbook_missing_a_frozen_sheet_fails_with_a_clear_error(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "thieu_sheet.xlsx"
    book = Workbook()
    book.active.title = "Summary 2026"
    book.save(path)
    with pytest.raises(LegacyImportError) as exc:
        parse_workbook(path)
    assert "DataChart 2026" in str(exc.value)
    # `Summary 2025` là REFERENCE_ONLY (DEC-169) nên KHÔNG còn là sheet bắt
    # buộc: thiếu nó không phải lỗi, và không được kể tên trong lỗi này.
    assert "Summary 2025" not in str(exc.value)


def test_sheet_visibility_state_is_recorded_as_imported(workbook):
    states = {item["sheet_name"]: item["state"] for item in workbook.sheets_imported}
    assert states["Summary 2026"] == "visible"
    assert states["DataChart 2026"] == "visible"
    # `DEC-177`: sheet OPTIONAL_IMPORT ĐƯỢC ghi vào bản ghi import, kèm
    # trạng thái ẩn/hiện thật của nó. Một sheet bị Excel để "hidden" vẫn là
    # dữ liệu của chủ dự án — giấu nó khỏi bản ghi nguồn gốc là làm mất dấu
    # vết, không phải giữ phạm vi.
    assert states["Summary 2025"] == "hidden"


def test_fingerprint_is_stable_for_the_same_bytes(legacy_workbook_path, tmp_path):
    import shutil

    copy = tmp_path / "ban_sao.xlsx"
    shutil.copyfile(legacy_workbook_path, copy)
    assert parse_workbook(copy).file_fingerprint == parse_workbook(
        legacy_workbook_path).file_fingerprint


def test_importer_never_touches_the_network_or_a_database():
    text = "".join(
        path.read_text(encoding="utf-8") for path in (REPO_ROOT / "app/legacy").rglob("*.py")
    )
    for banned in ("requests", "urllib", "sqlalchemy", "socket", "http"):
        assert banned not in text
