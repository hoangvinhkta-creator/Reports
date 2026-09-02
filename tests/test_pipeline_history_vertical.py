"""TASK-PRA-002 slice A — vertical thật: pipeline → persistence → đối chiếu.

Khác `test_snapshot_repository.py` (dựng dòng nguồn bằng tay), module này chạy
ĐÚNG pipeline authoritative trên fixture golden thật rồi mới ghi lịch sử. Nó
trả lời câu hỏi mà Owner đặt ra ở mục 3.1: upload sổ nửa tháng rồi upload sổ
cả tháng — hệ thống có đếm phần chồng nhau hai lần không.

Hai snapshot được SINH RA từ golden bằng cách cắt dòng trong `tmp_path`;
`tests/fixtures/golden/**` KHÔNG bị sửa và bản cắt KHÔNG được commit.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import create_engine, func, select

import tools.db as history_db
from app import demo
from app.modules.importing.raw_reader import read_raw_rows
from app.web import history_store, history_writer
from tests.test_105e_price_composition import write_catalog_capture, write_history_capture
from tests.test_tracking_history_reader import build_export
from tools.db import schema

REPO_ROOT = Path(__file__).resolve().parents[1]
GOLDEN = REPO_ROOT / "tests/fixtures/golden/period_2026_01.xlsx"
EXPECTED = REPO_ROOT / "tests/fixtures/golden/expected/period_2026_01.json"

# Mốc cắt của snapshot A (kịch bản "upload giữa kỳ rồi upload cả kỳ").
CUT_UNTIL = date(2026, 1, 10)

DATE_COLUMN = 1  # 1-based trong openpyxl, khớp raw_reader.COLUMNS["date"] = 0
FIRST_DATA_ROW = 6


@pytest.fixture(scope="module")
def captures(tmp_path_factory):
    """Capture Tracking tối thiểu: fixture golden đã ẩn danh nên không mã nào
    khớp danh mục — mọi dòng ra PENDING. Đó KHÔNG phải vấn đề ở đây: slice này
    kiểm tầng lưu, và tầng lưu phải ghi đúng kết quả pipeline dù nó là gì."""
    tmp = tmp_path_factory.mktemp("captures")
    return (write_history_capture(tmp, build_export(prices={}, events={})),
            write_catalog_capture(tmp, []))


def cut_workbook(source: Path, target: Path, until: date) -> Path:
    """Bản cắt giữ NGUYÊN VĂN dòng 1–5 và mọi dòng có `Ngày` ≤ mốc.

    Không sửa một ô nào của dòng được giữ lại — nếu không, "phần chồng nhau
    của hai snapshot là giống hệt" sẽ là điều test tự dựng ra chứ không phải
    điều nó chứng minh.
    """
    workbook = openpyxl.load_workbook(source)
    sheet = workbook.active
    drop = []
    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        value = sheet.cell(row, DATE_COLUMN).value
        # Ô không phải ngày (dòng tiêu đề phụ, dòng "Tổng cộng") KHÔNG mang dòng
        # bán nào — giữ nguyên như bản gốc, không suy diễn.
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date) and value > until:
            drop.append(row)
    for row in reversed(drop):
        sheet.delete_rows(row)
    workbook.save(target)
    workbook.close()
    return target


def run_pipeline(sales: Path, captures, output: Path):
    history, catalog = captures
    return demo.run_demo(sales=sales, tracking_capture=history,
                         tracking_catalog=catalog, output=output)


def fresh_repository() -> history_store.SnapshotRepository:
    engine = create_engine("sqlite://")
    history_db.create_all_for_test(engine)
    return history_store.SnapshotRepository(engine)


def persist(repository, sales: Path, captures, tmp_path: Path, *, run_id: str, at: str):
    run = run_pipeline(sales, captures, tmp_path / f"{run_id}.xlsx")
    outcome = history_writer.write_run_history(
        repository, demo_run=run, run_id=run_id, workbook_path=sales,
        display_name=sales.name, created_at=at,
    )
    return run, outcome


def count(repository, table) -> int:
    with repository.engine.connect() as connection:
        return connection.execute(select(func.count()).select_from(table)).scalar()


def measured(sales: Path) -> tuple[int, int]:
    """Số dòng/số đơn ĐO TỪ CHÍNH fixture — không hard-code theo con số kế hoạch."""
    rows = read_raw_rows(sales)
    return len(rows), len({row.order_id for row in rows})


@pytest.fixture(scope="module")
def snapshot_files(tmp_path_factory):
    tmp = tmp_path_factory.mktemp("snapshots")
    return cut_workbook(GOLDEN, tmp / "snapshot_a.xlsx", CUT_UNTIL), GOLDEN


# --- CHECK-PRA002-02 ------------------------------------------------------

def test_one_real_run_persists_exactly_what_the_exporter_reported(captures, tmp_path):
    repository = fresh_repository()
    run, outcome = persist(repository, GOLDEN, captures, tmp_path,
                           run_id="run-1", at="2026-02-01T00:00:00")
    lines, orders = measured(GOLDEN)

    snapshot = repository.get_snapshot(outcome.snapshot_id)
    assert (snapshot["line_count"], snapshot["order_count"]) == (lines, orders)
    assert (lines, orders) == (351, 254), "hình dạng fixture golden 01.2026"
    assert (snapshot["sheet_data_rows"], snapshot["rows_without_order_id"]) == (352, 1)
    assert snapshot["coverage_state"] == "HEADER_CONSISTENT"
    assert outcome.counts["INSERT"] == lines and outcome.counts["SAME"] == 0

    assert count(repository, schema.order_line_source_version) == lines
    assert count(repository, schema.order_line_result_version) == lines
    assert count(repository, schema.order_line_current) == lines

    totals = repository.current_totals()
    expected_sales = Decimal(json.loads(EXPECTED.read_text())["money"]["sales_normalized"][0])
    assert totals["total_sales"] == expected_sales == Decimal("3562310000")
    assert (totals["lines"], totals["orders"]) == (lines, orders)


def test_the_persisted_status_column_matches_the_exporter_review_count(captures, tmp_path):
    repository = fresh_repository()
    run, _ = persist(repository, GOLDEN, captures, tmp_path,
                     run_id="run-1", at="2026-02-01T00:00:00")
    with repository.engine.connect() as connection:
        pending = connection.execute(
            select(func.count()).select_from(schema.order_line_result_version)
            .where(schema.order_line_result_version.c.status == "PENDING")
        ).scalar()
    assert pending == run.summary.review_lines


def test_the_snapshot_carries_the_evidence_needed_to_reopen_the_run(captures, tmp_path):
    repository = fresh_repository()
    _, outcome = persist(repository, GOLDEN, captures, tmp_path,
                         run_id="run-1", at="2026-02-01T00:00:00")
    evidence = repository.get_snapshot(outcome.snapshot_id)["evidence_json"]
    for field in ("tracking_price_history_capture_id", "tracking_catalog_capture_id",
                  "identity_store_revision", "business_timezone_label",
                  "employee_master_snapshot_id"):
        assert field in evidence
    assert evidence["employee_master_snapshot_id"]


# --- CHECK-PRA002-03 ------------------------------------------------------

def test_uploading_the_same_book_twice_never_moves_a_single_dong(captures, tmp_path):
    repository = fresh_repository()
    _, first = persist(repository, GOLDEN, captures, tmp_path,
                       run_id="run-1", at="2026-02-01T00:00:00")
    before = repository.current_totals()

    _, second = persist(repository, GOLDEN, captures, tmp_path,
                        run_id="run-2", at="2026-02-02T00:00:00")

    lines, _ = measured(GOLDEN)
    assert second.counts == {"INSERT": 0, "SAME": lines, "SOURCE_CHANGED": 0,
                             "ORDER_KEY_COLLISION": 0}
    assert second.duplicate_of_snapshot_id == first.snapshot_id
    assert count(repository, schema.order_line_source_version) == lines
    with repository.engine.connect() as connection:
        assert connection.execute(
            select(func.count()).select_from(schema.order_line_source_version)
            .where(schema.order_line_source_version.c.version_no > 1)
        ).scalar() == 0
    assert count(repository, schema.order_line_current) == lines
    assert repository.current_totals() == before
    # Run thứ hai đã chạy pipeline thật → audit kết quả tăng, nghiệp vụ không.
    assert count(repository, schema.order_line_result_version) == lines * 2
    assert repository.count_flags(kind="SOURCE_CHANGED") == 0


# --- CHECK-PRA002-04 ------------------------------------------------------

def test_a_half_month_book_then_the_full_month_equals_the_full_month_alone(
    captures, tmp_path, snapshot_files,
):
    """Đẳng thức, không phải niềm tin: state(A rồi B) == state(B một mình)."""
    snapshot_a, snapshot_b = snapshot_files
    lines_a, orders_a = measured(snapshot_a)
    lines_b, orders_b = measured(snapshot_b)
    assert 0 < lines_a < lines_b, "bản cắt phải là tập con thật sự"

    sequential = fresh_repository()
    persist(sequential, snapshot_a, captures, tmp_path, run_id="seq-1",
            at="2026-02-01T00:00:00")
    assert sequential.current_totals()["lines"] == lines_a
    _, second = persist(sequential, snapshot_b, captures, tmp_path, run_id="seq-2",
                        at="2026-02-02T00:00:00")

    assert second.counts == {"INSERT": lines_b - lines_a, "SAME": lines_a,
                             "SOURCE_CHANGED": 0, "ORDER_KEY_COLLISION": 0}
    assert count(sequential, schema.order_line_source_version) == lines_b

    alone = fresh_repository()
    persist(alone, snapshot_b, captures, tmp_path, run_id="only-1",
            at="2026-02-02T00:00:00")

    assert sequential.current_totals() == alone.current_totals()
    assert sequential.current_fingerprints() == alone.current_fingerprints()
    assert sequential.current_totals()["orders"] == orders_b


def test_the_wide_book_first_then_the_narrow_one_adds_nothing_and_changes_nothing(
    captures, tmp_path, snapshot_files,
):
    """Thứ tự upload đảo: phần chồng nhau vẫn SAME, tổng hiện hành không đổi."""
    snapshot_a, snapshot_b = snapshot_files
    lines_a, _ = measured(snapshot_a)
    repository = fresh_repository()
    persist(repository, snapshot_b, captures, tmp_path, run_id="rev-1",
            at="2026-02-01T00:00:00")
    before = repository.current_totals()

    _, second = persist(repository, snapshot_a, captures, tmp_path, run_id="rev-2",
                        at="2026-02-02T00:00:00")

    assert second.counts["SAME"] == lines_a
    assert second.counts["INSERT"] == 0
    assert second.counts["SOURCE_CHANGED"] == 0
    assert repository.current_totals() == before


# --- CHECK-PRA002-05 ------------------------------------------------------

def test_one_edited_line_produces_one_new_version_and_exactly_that_much_money(
    captures, tmp_path, snapshot_files,
):
    _, snapshot_b = snapshot_files
    edited, order_key, product, delta = edit_one_line(
        snapshot_b, tmp_path / "snapshot_b_edited.xlsx",
    )
    repository = fresh_repository()
    persist(repository, snapshot_b, captures, tmp_path, run_id="edit-1",
            at="2026-02-01T00:00:00")
    before = repository.current_totals()

    _, second = persist(repository, edited, captures, tmp_path, run_id="edit-2",
                        at="2026-02-02T00:00:00")

    assert second.counts["SOURCE_CHANGED"] == 1
    assert second.counts["INSERT"] == 0
    with repository.engine.connect() as connection:
        versions = [dict(row._mapping) for row in connection.execute(
            select(schema.order_line_source_version)
            .where(schema.order_line_source_version.c.order_key == order_key)
            .order_by(schema.order_line_source_version.c.version_no)
        )]
    assert [row["version_no"] for row in versions] == [1, 2]
    assert versions[0]["changed_fields_json"] is None, "version cũ nguyên vẹn"

    flag, = repository.list_flags(snapshot_id=second.snapshot_id)
    assert flag["kind"] == "SOURCE_CHANGED"
    assert set(flag["detail_json"]) == {"sell_price", "total_sales_raw"}

    after = repository.current_totals()
    assert after["lines"] == before["lines"], "sửa một dòng KHÔNG tạo dòng mới"
    assert after["total_sales"] - before["total_sales"] == delta


def edit_one_line(source: Path, target: Path) -> tuple[Path, str, str, Decimal]:
    """Sửa ĐÚNG một dòng: đơn giá và doanh số bán đi cùng nhau, như kế toán làm."""
    workbook = openpyxl.load_workbook(source)
    sheet = workbook.active
    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        price = sheet.cell(row, 10).value      # Đơn giá
        sales = sheet.cell(row, 11).value      # Doanh số bán
        discount = sheet.cell(row, 12).value or 0
        quantity = sheet.cell(row, 9).value
        if price and sales and quantity:
            bump = Decimal("1000000")
            new_price = Decimal(str(price)) + bump
            new_sales = new_price * Decimal(str(quantity)) - Decimal(str(discount))
            delta = new_sales - Decimal(str(sales))
            sheet.cell(row, 10).value = int(new_price)
            sheet.cell(row, 11).value = int(new_sales)
            order_key = str(sheet.cell(row, 2).value).strip()
            product = str(sheet.cell(row, 4).value).strip()
            workbook.save(target)
            workbook.close()
            return target, order_key, product, delta
    raise AssertionError("fixture golden không có dòng nào đủ giá/số lượng để sửa")


# --- ranh giới ------------------------------------------------------------

def test_the_exporter_alias_is_the_same_object_the_report_is_printed_from():
    """Không có nguồn sự thật thứ hai về AUTO/PENDING."""
    from app.modules.exporting import excel_exporter

    assert excel_exporter.present_lines is excel_exporter._present_lines
    assert excel_exporter.PresentedLine is excel_exporter._PresentedLine


def test_the_demo_run_carries_the_rows_and_lines_the_writer_needs(captures, tmp_path):
    run = run_pipeline(GOLDEN, captures, tmp_path / "boundary.xlsx")
    assert len(run.raw_rows) == len(run.presented_lines) == run.summary.total_lines
    assert {view.status for view in run.presented_lines} <= {"AUTO", "PENDING"}


# --- slice B: vắng mặt trên dữ liệu golden thật ----------------------------

def drop_one_line(source: Path, target: Path) -> tuple[Path, str, Decimal]:
    """Bỏ ĐÚNG một dòng cuối sổ — mô phỏng kế toán xoá một chứng từ.

    Chọn dòng CUỐI có Số BH để phép bỏ không làm dịch ``occurrence_index`` của
    dòng nào khác: test này nói về sự vắng mặt, không phải về hiệu ứng phụ của
    việc đánh lại chỉ số.
    """
    workbook = openpyxl.load_workbook(source)
    sheet = workbook.active
    for row in range(sheet.max_row, FIRST_DATA_ROW - 1, -1):
        order_key = sheet.cell(row, 2).value
        sales = sheet.cell(row, 11).value
        if order_key and str(order_key).strip() and sales:
            key = str(order_key).strip()
            sheet.delete_rows(row)
            workbook.save(target)
            workbook.close()
            return target, key, Decimal(str(sales))
    raise AssertionError("fixture golden không có dòng nào để bỏ")


def test_a_dropped_line_is_only_not_seen_until_someone_confirms_the_book(
    captures, tmp_path, snapshot_files,
):
    """CHECK-PRA002-07 trên dữ liệu golden: NOT_SEEN → (xác nhận) → REMOVED.

    Ở CẢ HAI trạng thái, dòng bị bỏ vẫn là dòng hiện hành và vẫn nằm trong
    tổng doanh thu — đó là toàn bộ điểm của slice B.
    """
    _, snapshot_b = snapshot_files
    dropped, order_key, _ = drop_one_line(snapshot_b, tmp_path / "snapshot_b_dropped.xlsx")
    repository = fresh_repository()
    persist(repository, snapshot_b, captures, tmp_path, run_id="drop-1",
            at="2026-02-01T00:00:00")
    before = repository.current_totals()
    facts = {table.name: count(repository, table) for table in schema.PIPELINE_TABLES}

    _, second = persist(repository, dropped, captures, tmp_path, run_id="drop-2",
                        at="2026-02-02T00:00:00")

    not_seen = [flag for flag in repository.list_flags(snapshot_id=second.snapshot_id)
                if flag["kind"] == "NOT_SEEN_IN_LATEST_SNAPSHOT"]
    assert len(not_seen) == 1 and not_seen[0]["order_key"] == order_key
    assert repository.count_flags(kind="REMOVED_IN_SOURCE_CANDIDATE") == 0
    assert repository.current_totals() == before, "chưa xác nhận: không đổi gì"

    confirmation = repository.confirm_coverage(
        second.snapshot_id, start=date(2026, 1, 1), end=date(2026, 1, 31),
        confirmed=True, confirmed_at="2026-02-03T00:00:00",
    )

    assert confirmation.removed_candidates == 1
    removed, = [flag for flag in repository.list_flags()
                if flag["kind"] == "REMOVED_IN_SOURCE_CANDIDATE"]
    assert removed["order_key"] == order_key
    assert repository.current_totals() == before, (
        "REMOVED_CANDIDATE vẫn current, vẫn tính — không xoá, không trừ tiền"
    )
    for table in schema.PIPELINE_TABLES:
        assert count(repository, table) >= facts[table.name], (
            f"{table.name}: không bảng fact nào được phép mất bản ghi"
        )


def test_confirming_the_half_month_book_never_touches_the_second_half(
    captures, tmp_path, snapshot_files,
):
    """Ranh giới phạm vi trên dữ liệu thật: xác nhận 01–10 KHÔNG đụng 11–31.

    Kịch bản ngược của mục 3.1: sổ cả tháng đã lưu, rồi Owner upload lại sổ
    nửa đầu tháng và xác nhận nó đầy đủ cho ĐÚNG 01–10/01. Các đơn ngày
    11–31/01 nằm ngoài phạm vi được xác nhận, nên chúng không phải ứng viên
    bị xoá — dù chúng "không có" trong sổ vừa xác nhận.
    """
    snapshot_a, snapshot_b = snapshot_files
    lines_a, _ = measured(snapshot_a)
    lines_b, _ = measured(snapshot_b)
    repository = fresh_repository()
    persist(repository, snapshot_b, captures, tmp_path, run_id="scope-1",
            at="2026-02-01T00:00:00")
    before = repository.current_totals()
    _, narrow = persist(repository, snapshot_a, captures, tmp_path, run_id="scope-2",
                        at="2026-02-02T00:00:00")

    assert repository.get_snapshot(narrow.snapshot_id)["n_not_seen"] == 0, (
        "sổ nửa tháng không phát biểu gì về nửa sau — không cờ vắng mặt giả"
    )

    confirmation = repository.confirm_coverage(
        narrow.snapshot_id, start=date(2026, 1, 1), end=CUT_UNTIL,
        confirmed=True, confirmed_at="2026-02-03T00:00:00",
    )

    assert confirmation.removed_candidates == 0
    assert repository.count_flags(kind="REMOVED_IN_SOURCE_CANDIDATE") == 0
    assert repository.current_totals() == before
    assert repository.current_totals()["lines"] == lines_b > lines_a


def test_confirming_the_narrow_book_for_the_whole_month_does_reach_the_rest(
    captures, tmp_path, snapshot_files,
):
    """Cùng dữ liệu, khoảng khai báo rộng hơn → thẩm quyền rộng hơn, có kiểm soát.

    Đây là mặt kia của cùng một luật: phạm vi do người xác nhận khai, và hệ
    quả bám đúng phạm vi đó. Ngay cả ở đây, hệ quả vẫn chỉ là cờ Review —
    tổng tiền không suy chuyển một đồng.
    """
    snapshot_a, snapshot_b = snapshot_files
    lines_a, _ = measured(snapshot_a)
    lines_b, _ = measured(snapshot_b)
    repository = fresh_repository()
    persist(repository, snapshot_b, captures, tmp_path, run_id="wide-1",
            at="2026-02-01T00:00:00")
    before = repository.current_totals()
    _, narrow = persist(repository, snapshot_a, captures, tmp_path, run_id="wide-2",
                        at="2026-02-02T00:00:00")

    confirmation = repository.confirm_coverage(
        narrow.snapshot_id, start=date(2026, 1, 1), end=date(2026, 1, 31),
        confirmed=True, confirmed_at="2026-02-03T00:00:00",
    )

    assert confirmation.removed_candidates == lines_b - lines_a
    assert repository.current_totals() == before, "vẫn không xoá, vẫn tính đủ"
