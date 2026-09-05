"""PHB-04 — Legacy Reference V1: kỳ lịch sử đọc được mà KHÔNG làm sai số mới.

Nhóm test này trả lời đúng mười câu của Done Gate PHB-04 mục 12, và mỗi test
chứng minh một KẾT QUẢ nghiệp vụ chứ không phải "hàm có chạy":

    A kỳ legacy đọc được · B provenance là LEGACY_REFERENCE · C không vào
    pipeline kế toán · D không đụng coverage lợi nhuận · E kỳ PHB-03 y nguyên ·
    F điều hướng phân biệt được hai nguồn · G chỉ tiêu thiếu hiện N/A chứ
    không phải 0 · H chỉ so khi hợp đồng cho phép · I chỉ tiêu tham chiếu
    không âm thầm sinh phép so · J nạp lại cùng một bản không nhân đôi số.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest
from sqlalchemy import create_engine, func, select

import tools.db as history_db
from app.legacy import LegacyImportError, parse_workbook, parse_year_workbook
from app.web import history_store, legacy_presentation, legacy_reference
from app.web import business_service, business_store
from app.web import server as web_server
from tests.fixtures.legacy.build_legacy_workbook import strip_formula_markers
from tests.fixtures.legacy.build_year_workbook import build_year_workbook
from tests.test_business_vertical import pair, persist
from tools.db import schema
from tools.tracking import live_pull

JANUARY = {"date_from": date(2026, 1, 1), "date_to": date(2026, 1, 31)}

# Giá trị cột `AH` của fixture (`tests/fixtures/legacy/build_legacy_workbook.py`):
# doanh số cùng kỳ năm trước, đơn vị VND nguyên. Chỉ ba tháng đầu có số —
# chín tháng còn lại là ô TRỐNG, và đó chính là điều test G cần.
FIXTURE_PREV_YEAR = {1: Decimal("2410000000"), 2: Decimal("1520000000"),
                     3: Decimal("870000000")}


# --- Hạ tầng dùng chung ---------------------------------------------------

@pytest.fixture
def engine():
    engine = create_engine("sqlite://")
    history_db.create_all_for_test(engine)
    return engine


@pytest.fixture
def repository(engine):
    return history_store.build(engine=engine)


@pytest.fixture
def snapshots(engine):
    return history_store.SnapshotRepository(engine)


@pytest.fixture
def service(engine):
    return business_service.BusinessReportService(
        engine=engine, store=business_store.BusinessDecisionStore(engine))


@pytest.fixture
def app(monkeypatch, tmp_path, repository, snapshots):
    monkeypatch.setattr(web_server, "UPLOAD_DIR", tmp_path / "uploads")
    monkeypatch.setattr(web_server, "ARTIFACT_DIR", (tmp_path / "outputs").resolve())
    monkeypatch.setattr(web_server, "select_latest_valid_captures", lambda: None)
    monkeypatch.setattr(live_pull, "is_configured", lambda env=None: False)
    application = web_server.create_app(db_path=tmp_path / "runs.db",
                                        history=repository, snapshots=snapshots)
    application.testing = True
    return application


@pytest.fixture
def client(app):
    return app.test_client()


def upload(client, path):
    """Nạp một workbook vào history store DÙNG CHUNG với `client` (qua
    `HISTORY_STORE` của Flask app) — GIỮ NGUYÊN tên file thật làm
    `source_file_name`, đúng hành vi upload cũ (`_safe_display_name` chỉ cắt
    basename, không đổi tên).

    `POST /du-lieu/legacy` đã khóa vĩnh viễn (`DEC-181`, repair R2-B01) —
    route không còn tạo import nào, kể cả trong test, nên seed đi thẳng qua
    tầng repository. Dùng đúng bộ nhận dạng hình dạng workbook
    (`_looks_like_year_workbook`) mà route từng dùng để hai hình dạng
    (workbook năm độc lập / workbook nhiều tháng) vẫn phân biệt đúng như cũ.
    """
    repository = client.application.config["HISTORY_STORE"]
    workbook = (
        web_server.parse_year_workbook(path)
        if web_server._looks_like_year_workbook(path)
        else web_server.parse_workbook(path)
    )
    return repository.create_import(workbook)


@pytest.fixture
def loaded(client, legacy_workbook_path):
    upload(client, legacy_workbook_path)
    return client


def page(client, path="/lich-su") -> str:
    response = client.get(path)
    assert response.status_code == 200, f"{path} → {response.status_code}"
    return response.get_data(as_text=True)


def table_counts(engine) -> dict:
    """Số dòng của MỌI bảng trong history store — dấu vân tay của trạng thái."""
    with engine.connect() as connection:
        return {
            table.name: connection.execute(
                select(func.count()).select_from(table)).scalar()
            for table in schema.METADATA.sorted_tables
        }


# --- A. Kỳ legacy đọc được ------------------------------------------------

class TestLegacyPeriodIsReadable:
    def test_reference_year_periods_come_from_the_prev_year_column(self, loaded, repository):
        """Kỳ 2025 sinh ra từ cột `AH` của DataChart 2026, không từ sheet nào khác."""
        periods = legacy_reference.reference_periods(
            repository.query_monthly_reference())
        assert legacy_reference.reference_years(periods) == [2025]
        by_month = {item.month: item.value for item in periods}
        for month, expected in FIXTURE_PREV_YEAR.items():
            assert by_month[month] == expected, f"tháng {month}"

    def test_the_page_shows_the_reference_period_values(self, loaded):
        html = page(loaded)
        assert "Tháng 01/2025" in html
        # 2.410.000.000 — định dạng vi-VN của ô AH3.
        assert "2.410.000.000" in html

    def test_a_workbook_year_period_is_still_readable(self, loaded, repository):
        """Kỳ báo cáo tay của năm workbook không bị PHB-04 làm hụt đi."""
        assert (2026, 1) in repository.available_periods()

    def test_the_page_says_so_when_nothing_has_been_imported(self, client):
        assert "Chưa nhập bản báo cáo cũ nào" in page(client)


# --- B. Provenance --------------------------------------------------------

class TestProvenanceIsExplicit:
    def test_every_projected_period_carries_legacy_reference(self, loaded, repository):
        periods = legacy_reference.reference_periods(
            repository.query_monthly_reference())
        assert periods
        assert {item.provenance for item in periods} == {"LEGACY_REFERENCE"}

    def test_the_page_names_the_provenance_and_the_source_cell(self, loaded):
        html = page(loaded)
        assert "LEGACY_REFERENCE" in html
        assert legacy_reference.PROVENANCE_LABEL in html
        assert "DataChart 2026!AH" in html

    def test_no_reference_value_appears_without_a_legacy_badge(self, loaded, repository):
        """Mọi ô đi qua `cell()` nên luôn có nhãn nguồn và đơn vị."""
        rows = legacy_presentation.reference_rows(
            legacy_reference.reference_periods(repository.query_monthly_reference()))
        assert rows
        for row in rows:
            assert row["cell"]["unit"] == "đồng (số cũ)"

    def test_the_page_never_claims_the_current_engine_produced_these_numbers(self, loaded):
        html = page(loaded)
        assert "KHÔNG do công cụ hiện tại tính lại" in html
        for forbidden in ("AUTO_CALCULATED", "ERP_RECONCILED", "CURRENT_ENGINE"):
            assert forbidden not in html


# --- C. Không vào pipeline kế toán hiện hành ------------------------------

class TestLegacyStaysOutOfTheAccountingPipeline:
    def test_legacy_import_writes_nothing_into_pipeline_tables(
        self, client, legacy_workbook_path, engine
    ):
        pipeline_tables = ("order_line_source_version", "order_line_result_version",
                           "order_line_current", "source_snapshot", "snapshot_line",
                           "reconciliation_flag")
        upload(client, legacy_workbook_path)
        counts = table_counts(engine)
        assert counts["legacy_monthly_reference"] > 0, "chưa nhập được gì thì test vô nghĩa"
        for table in pipeline_tables:
            assert counts[table] == 0, table

    def test_reading_the_legacy_page_writes_nothing_at_all(
        self, loaded, engine
    ):
        before = table_counts(engine)
        page(loaded)
        page(loaded)
        assert table_counts(engine) == before

    def test_every_legacy_row_is_constrained_to_the_legacy_origin(self, loaded, engine):
        """Không phải quy ước đặt tên: đây là CHECK constraint ở tầng schema."""
        with engine.connect() as connection:
            for table in (schema.legacy_import, schema.legacy_summary_row,
                          schema.legacy_daily_sales, schema.legacy_monthly_reference):
                origins = connection.execute(
                    select(table.c.origin).distinct()).scalars().all()
                assert set(origins) <= {"LEGACY_REFERENCE"}, table.name


# --- D. Không đụng coverage lợi nhuận của số mới --------------------------

class TestLegacyDoesNotTouchCurrentProfitCoverage:
    def test_coverage_is_identical_before_and_after_a_legacy_import(
        self, client, legacy_workbook_path, snapshots, service
    ):
        """Nạp một bản legacy KHÔNG được làm nhúc nhích coverage giá nhập."""
        persist(snapshots, [pair("BH1", kpi_purchase=None, kpi_profit=None)])
        before = service.period(**JANUARY).totals

        upload(client, legacy_workbook_path)
        after = service.period(**JANUARY).totals

        assert after.coverage == before.coverage
        assert after.coverage.covered_lines == 0
        assert after.coverage.missing_price_lines == 1
        assert after.kpi_profit is None          # NULL, không phải 0
        assert after.official_kpi_profit is None

    def test_legacy_money_never_reaches_the_current_period_totals(
        self, client, legacy_workbook_path, snapshots, service
    ):
        persist(snapshots, [pair("BH1", sell="8000000", kpi_purchase="5000000",
                                 kpi_profit="3000000")])
        upload(client, legacy_workbook_path)
        totals = service.period(**JANUARY).totals
        # Doanh thu vẫn đúng một dòng bán; không cộng thêm đồng nào của số cũ.
        assert totals.sales_revenue == Decimal("8000000")
        assert totals.lines == 1


# --- E. Kỳ PHB-03 y nguyên ------------------------------------------------

class TestCurrentBusinessPagesAreUnchanged:
    def test_the_business_summary_still_renders_after_a_legacy_import(
        self, loaded, snapshots
    ):
        persist(snapshots, [pair("BH1", kpi_purchase="5000000", kpi_profit="3000000")])
        html = page(loaded, "/kinh-doanh?ky=2026-01")
        assert "8.000.000" in html

    def test_the_legacy_pages_of_pra_001_still_render(self, loaded):
        assert "SỐ CŨ" in page(loaded, "/nhan-vien")
        page(loaded, "/doanh-so-ngay")


# --- F. Điều hướng phân biệt được hai nguồn -------------------------------

class TestNavigationDistinguishesOrigins:
    def test_a_period_with_both_origins_is_labelled_with_both(self):
        rows = legacy_reference.period_navigation(
            legacy_summary_periods=[(2026, 1)],
            legacy_reference_periods=[],
            pipeline_periods=[(2026, 1)],
        )
        assert len(rows) == 1
        assert rows[0].both
        assert rows[0].origin_labels == ["SỐ MỚI", "SỐ CŨ"]

    def test_a_reference_year_period_is_legacy_only(self, loaded, repository):
        periods = legacy_reference.reference_periods(
            repository.query_monthly_reference())
        rows = legacy_reference.period_navigation(
            legacy_summary_periods=[], legacy_reference_periods=periods,
            pipeline_periods=[],
        )
        assert rows
        assert all(row.year == 2025 and row.has_legacy and not row.has_pipeline
                   for row in rows)

    def test_an_empty_reference_cell_never_becomes_a_navigable_period(
        self, loaded, repository
    ):
        """Tháng 04–12/2025 có ô AH TRỐNG ⟹ không phải một kỳ có dữ liệu."""
        periods = legacy_reference.reference_periods(
            repository.query_monthly_reference())
        rows = legacy_reference.period_navigation(
            legacy_summary_periods=[], legacy_reference_periods=periods,
            pipeline_periods=[],
        )
        assert {row.month for row in rows} == set(FIXTURE_PREV_YEAR)

    def test_the_page_links_to_both_the_legacy_and_the_current_view(
        self, loaded, snapshots
    ):
        persist(snapshots, [pair("BH1", kpi_purchase="5000000", kpi_profit="3000000")])
        html = page(loaded)
        assert "/nhan-vien?ky=2026-01" in html
        assert "/kinh-doanh?ky=2026-01" in html

    def test_the_page_says_when_the_current_period_list_is_unreadable(
        self, monkeypatch, tmp_path, repository, legacy_workbook_path
    ):
        """Thiếu snapshot store ⟹ danh mục kỳ KHÔNG được im lặng chỉ có legacy.

        Hôm nay `_build_snapshots()` dựng snapshot repo trên chính engine của
        history store, nên một app CÓ history luôn CÓ cả snapshot — trạng thái
        dưới đây phải dựng bằng monkeypatch. Nhánh này vẫn được giữ và được
        kiểm: nó là cùng một kỷ luật với `not_configured()` của PRA-001 — một
        danh sách thiếu nguồn không bao giờ được trông giống danh sách đủ.
        """
        monkeypatch.setattr(web_server, "UPLOAD_DIR", tmp_path / "uploads")
        monkeypatch.setattr(web_server, "ARTIFACT_DIR", (tmp_path / "outputs").resolve())
        monkeypatch.setattr(web_server, "select_latest_valid_captures", lambda: None)
        monkeypatch.setattr(live_pull, "is_configured", lambda env=None: False)
        monkeypatch.setattr(web_server, "_build_snapshots", lambda legacy: None)
        application = web_server.create_app(db_path=tmp_path / "runs.db",
                                            history=repository)
        application.testing = True
        client = application.test_client()
        upload(client, legacy_workbook_path)
        assert "không phải là toàn bộ kỳ đang có" in page(client)


# --- G. Chỉ tiêu thiếu hiện N/A, KHÔNG phải 0 -----------------------------

class TestMissingMetricsAreNeverFabricated:
    def test_an_empty_reference_month_renders_a_dash_not_a_zero(self, loaded, repository):
        rows = legacy_presentation.reference_rows(
            legacy_reference.reference_periods(repository.query_monthly_reference()))
        empty = [row for row in rows if not row["available"]]
        assert empty, "fixture phải có tháng trống, nếu không test này vô nghĩa"
        for row in empty:
            assert row["cell"]["text"] == "—"
            assert row["cell"]["empty"] is True

    def test_unavailable_metrics_are_declared_not_invented(self):
        keys = {rule.key for rule
                in legacy_reference.unavailable_metrics("REFERENCE_YEAR")}
        # Năm trước KHÔNG có bằng chứng cho bất cứ thứ gì ngoài doanh số tháng.
        assert keys == {"orders", "products", "converted_revenue", "profit",
                        "target", "by_employee", "daily_sales"}

    def test_the_only_supported_reference_year_metric_is_monthly_sales(self):
        """`DEC-180` — vẫn ĐÚNG MỘT chỉ tiêu, nhưng nay nó SO ĐƯỢC.

        Chủ dự án đã bác lý do chặn cũ ("báo cáo tay trừ chiết khấu khác
        cách công cụ trừ") bằng chính cách sổ tay ghi chiết khấu. Số lượng
        chỉ tiêu có bằng chứng KHÔNG đổi — chỉ lớp của nó đổi.
        """
        supported = legacy_reference.supported_metrics("REFERENCE_YEAR")
        assert [rule.key for rule in supported] == ["sales_vnd"]
        assert supported[0].metric_class == legacy_reference.COMPARABLE

    def test_the_page_states_why_a_metric_is_unavailable(self, loaded):
        html = page(loaded)
        assert "Không có bằng chứng" in html


# --- H/I. Cổng so sánh ----------------------------------------------------

class TestComparisonGate:
    def test_exactly_the_total_sales_pairs_are_comparable_and_no_others(self):
        """`DEC-180` mở ĐÚNG hai cặp Tổng bán, không mở lây một cặp nào khác.

        Đây là khẳng định canh "scope creep" của chính bản sửa này: chủ dự án
        chứng minh Tổng bán cùng nghĩa, KHÔNG chứng minh lợi nhuận, DS quy
        đổi, số đơn hay số SP cùng nghĩa. Lý do chặn của bốn cặp đó chưa ai
        bác, nên chúng phải còn nguyên `False`.
        """
        allowed = {(rule.legacy_key, rule.current_key)
                   for rule in legacy_reference.CROSS_ORIGIN_CONTRACT
                   if rule.allowed}
        assert allowed == {("sales", "sales_revenue"),
                           ("sales_vnd", "sales_revenue")}
        blocked = {(rule.legacy_key, rule.current_key)
                   for rule in legacy_reference.CROSS_ORIGIN_CONTRACT
                   if not rule.allowed}
        assert blocked == {("profit", "kpi_profit"),
                           ("converted_revenue", "converted_sales"),
                           ("orders", "orders"),
                           ("products", "qualifying_quantity")}

    def test_the_metric_class_agrees_with_the_gate_for_total_sales(self):
        """Hai cách ghi cùng một phán quyết KHÔNG được phép mâu thuẫn nhau.

        `MetricRule.metric_class` và `CROSS_ORIGIN_CONTRACT` là hai biểu diễn
        của cùng một câu hỏi. Mở cổng mà quên đổi lớp chỉ tiêu là để màn hình
        nói "không so được" ngay cạnh một tỉ lệ vừa được tính ra.
        """
        assert legacy_reference.has_comparable_metric() is True
        comparable = {rule.key for period_class in legacy_reference.CONTRACTS
                      for rule in legacy_reference.rules(period_class)
                      if rule.metric_class == legacy_reference.COMPARABLE}
        assert comparable == {"sales", "sales_vnd"}

    def test_a_reference_only_metric_produces_no_percentage(self):
        """Cặp CHƯA được chứng minh vẫn không sinh ra một con số nào."""
        result = legacy_reference.compare(
            Decimal("1000"), Decimal("1500"),
            legacy_key="profit", current_key="kpi_profit")
        assert result.allowed is False
        assert result.percent is None
        assert "Không so được" in result.note

    def test_the_proven_total_sales_pair_now_produces_a_percentage(self):
        """`DEC-180` — cặp đã được chứng minh thì cổng phải TÍNH, không chặn."""
        result = legacy_reference.compare(
            Decimal("1000000"), Decimal("1500000"),
            legacy_key="sales", current_key="sales_revenue")
        assert result.allowed is True
        assert result.percent == Decimal("50")

    def test_a_pair_outside_the_contract_is_refused_rather_than_guessed(self):
        result = legacy_reference.compare(
            Decimal("1000"), Decimal("1500"),
            legacy_key="bonus", current_key="kpi_profit")
        assert result.allowed is False
        assert result.percent is None
        assert "chưa xét cặp chỉ tiêu này" in result.note

    def test_the_gate_reads_the_contract_instead_of_hardcoding_a_refusal(self):
        """Cổng phải là cổng THẬT: hợp đồng cho phép ⟹ nó tính."""
        permissive = (legacy_reference.CrossOriginRule(
            "sales_vnd", "sales_revenue", True, "giả định của test"),)
        result = legacy_reference.compare(
            Decimal("1000"), Decimal("1500"),
            legacy_key="sales_vnd", current_key="sales_revenue",
            contract=permissive)
        assert result.allowed is True
        assert result.percent == Decimal("50")

    def test_a_permitted_pair_with_a_missing_side_still_refuses_to_invent(self):
        permissive = (legacy_reference.CrossOriginRule(
            "sales_vnd", "sales_revenue", True, "giả định của test"),)
        for legacy_value, current_value in ((None, Decimal("10")),
                                            (Decimal("10"), None),
                                            (Decimal("10"), Decimal("0"))):
            result = legacy_reference.compare(
                legacy_value, current_value, legacy_key="sales_vnd",
                current_key="sales_revenue", contract=permissive)
            assert result.percent is None
            assert "không tính được" in result.note

    def test_the_page_states_which_metric_is_comparable_and_which_are_not(
        self, loaded, snapshots
    ):
        """Trang lịch sử phải NÓI RA phán quyết mới, không giữ câu chữ cũ.

        Một trang còn viết "chưa chỉ tiêu nào đạt điều đó" trong khi cổng đã
        tính được một tỉ lệ là một trang nói dối về chính hệ thống của nó.
        """
        persist(snapshots, [pair("BH1", kpi_purchase="5000000", kpi_profit="3000000")])
        html = page(loaded)
        assert "đúng một chỉ tiêu đạt điều đó" in html
        assert "chưa chỉ tiêu nào đạt điều đó" not in html
        # Bốn cặp còn lại vẫn phải hiện nguyên phán quyết chặn của chúng.
        assert "Không so được" in html

    def test_every_contract_pair_records_its_reason_whichever_way_it_went(self):
        """Cho phép hay từ chối, mỗi dòng hợp đồng đều phải nói VÌ SAO."""
        rows = legacy_reference.comparison_summary()
        for row in rows:
            assert row["reason"].strip(), row["legacy_key"]
        assert [row["allowed"] for row in rows].count(True) == 2


# --- J. Nạp lại cùng một bản không nhân đôi số ----------------------------

class TestReloadingIsIdempotent:
    def test_uploading_the_same_workbook_twice_creates_no_second_version(
        self, client, legacy_workbook_path, engine, repository
    ):
        upload(client, legacy_workbook_path)
        after_first = table_counts(engine)
        first_import = repository.current_import()["import_id"]

        upload(client, legacy_workbook_path)

        assert table_counts(engine) == after_first
        assert repository.current_import()["import_id"] == first_import
        assert repository.count_imports() == 1

    def test_the_projected_reference_periods_do_not_duplicate(
        self, client, legacy_workbook_path, repository
    ):
        upload(client, legacy_workbook_path)
        upload(client, legacy_workbook_path)
        periods = legacy_reference.reference_periods(
            repository.query_monthly_reference())
        keys = [(item.year, item.month) for item in periods]
        assert len(keys) == len(set(keys))


# --- `DEC-177` — Summary 2025 là OPTIONAL_IMPORT, không phải sheet bị cấm --

class TestSummary2025IsInScope:
    """Chủ dự án đính chính (`DEC-177`): 2025 CÓ Summary và CÓ chi tiết nhân
    viên. `DEC-169` nói *"Owner KHÔNG yêu cầu"* — một tuyên bố PHẠM VI, không
    phải lệnh cấm sản phẩm. Nhóm test này khoá đúng ngữ nghĩa mới, và khoá
    luôn điều KHÔNG được đổi: guard DEC-168 trên sheet REQUIRED_IMPORT.
    """

    def test_a_classifiable_summary_2025_sheet_is_imported(self, loaded, engine):
        with engine.connect() as connection:
            sheets = connection.execute(
                select(schema.legacy_summary_row.c.sheet_name).distinct()
            ).scalars().all()
        assert "Summary 2025" in sheets

    def test_imported_summary_2025_rows_keep_the_legacy_origin(self, loaded, engine):
        with engine.connect() as connection:
            origins = connection.execute(
                select(schema.legacy_summary_row.c.origin)
                .where(schema.legacy_summary_row.c.sheet_name == "Summary 2025")
                .distinct()
            ).scalars().all()
        assert origins == ["LEGACY_REFERENCE"]

    def test_the_summary_contract_applies_to_every_year_not_just_the_workbook_year(self):
        """Summary 2025 và Summary 2026 dùng CÙNG 16 cột ⟹ cùng hợp đồng."""
        assert legacy_reference.WORKBOOK_YEAR_CONTRACT is \
            legacy_reference.SUMMARY_SHEET_CONTRACT
        assert len(legacy_reference.SUMMARY_SHEET_CONTRACT) == 16

    def test_a_value_only_summary_2025_never_breaks_the_2026_import(
        self, client, legacy_workbook_path, repository
    ):
        """Hình dạng workbook THẬT (0 công thức trong Summary 2025).

        Đây là bất biến của `DEC-169` phải sống sót qua `DEC-177`: mở phạm vi
        KHÔNG được biến hình dạng value-only thành một lần nhập trượt.
        """
        stripped = strip_formula_markers(legacy_workbook_path,
                                         sheet_name="Summary 2025")
        upload(client, stripped)
        rows = repository.query_summary(2026)
        assert rows, "phần 2026 phải nhập được bình thường"
        assert repository.query_summary(2025) == []

    def test_unreadable_rows_are_counted_and_surfaced_never_swallowed(
        self, client, legacy_workbook_path, repository
    ):
        stripped = strip_formula_markers(legacy_workbook_path,
                                         sheet_name="Summary 2025")
        upload(client, stripped)
        unread = legacy_reference.unread_sheets(
            repository.current_import()["sheets_imported"])
        assert [item.sheet_name for item in unread] == ["Summary 2025"]
        assert unread[0].unclassified_rows == 3
        assert unread[0].imported_rows == 0

    def test_the_page_tells_the_owner_what_is_still_unread(
        self, client, legacy_workbook_path
    ):
        stripped = strip_formula_markers(legacy_workbook_path,
                                         sheet_name="Summary 2025")
        upload(client, stripped)
        html = page(client)
        assert "CHƯA đọc được" in html
        assert "Summary 2025" in html

    def test_the_dec_168_guard_still_fires_on_a_required_sheet(
        self, legacy_workbook_path
    ):
        """Điều KHÔNG được đổi: sheet REQUIRED value-only vẫn phải FAIL TO."""
        stripped = strip_formula_markers(legacy_workbook_path,
                                         sheet_name="Summary 2026")
        with pytest.raises(LegacyImportError) as exc:
            parse_workbook(stripped)
        assert "Summary 2026" in str(exc.value)


# --- Câu hỏi thật của chủ dự án về một kỳ 2025 (chỉ thị §11) --------------

class TestOwnerQuestionsAboutALegacyYear:
    """A–I của chỉ thị đính chính, hỏi bằng đúng ngôn ngữ của chủ dự án."""

    def test_a_total_historical_sales_for_a_2025_month(self, loaded, repository):
        """A. Tháng 01/2025 tổng bán bao nhiêu? — lấy từ dòng TỔNG của kỳ."""
        rows = repository.query_summary(2025, 1)
        totals = [r for r in rows if r["row_kind"] == "MONTH_TOTAL"]
        assert len(totals) == 1
        assert totals[0]["sales"] == Decimal("2030000")

    def test_b_the_other_accepted_summary_metrics_of_that_month(self, loaded, repository):
        total = [r for r in repository.query_summary(2025, 1)
                 if r["row_kind"] == "MONTH_TOTAL"][0]
        assert total["converted_revenue"] == Decimal("1050000")
        assert total["profit"] == Decimal("57750")

    def test_c_which_employees_had_data_that_month(self, loaded, repository):
        years = legacy_reference.summary_years(repository.query_all_summary())
        year_2025 = next(y for y in years if y.year == 2025)
        assert year_2025.has_employee_detail
        assert set(year_2025.sellers) == {"NV-A", "NV-B"}
        # Dòng tổng tháng KHÔNG được lẫn vào danh sách nhân viên.
        assert "Tổng T01" not in year_2025.sellers

    def test_d_the_accepted_metrics_for_one_employee(self, loaded, repository):
        """Tháng 01/2025, NV-A: bán bao nhiêu, DS quy đổi, lợi nhuận."""
        row = next(r for r in repository.query_summary(2025, 1)
                   if r["seller_label"] == "NV-A")
        assert row["sales"] == Decimal("1120000")
        assert row["converted_revenue"] == Decimal("580000")
        assert row["profit"] == Decimal("31900")

    def test_e_every_displayed_2025_value_stays_legacy_reference(self, loaded, repository):
        rows = repository.query_summary(2025)
        assert rows
        assert {r["origin"] for r in rows} == {"LEGACY_REFERENCE"}
        matrix = legacy_presentation.matrix(rows)
        for row in matrix:
            for cell in row["cells"]:
                assert cell["unit"], "mọi ô legacy phải mang nhãn đơn vị"

    def test_f_none_of_it_contaminates_current_engine_coverage(
        self, client, legacy_workbook_path, snapshots, service
    ):
        persist(snapshots, [pair("BH1", kpi_purchase=None, kpi_profit=None)])
        before = service.period(**JANUARY).totals
        upload(client, legacy_workbook_path)
        after = service.period(**JANUARY).totals
        assert after.coverage == before.coverage
        assert after.sales_revenue == before.sales_revenue
        assert after.kpi_profit is None

    def test_g_the_owner_can_navigate_from_2025_summary_to_employee_detail(self, loaded):
        html = page(loaded)
        assert "Năm 2025" in html
        assert "/nhan-vien?ky=2025-01" in html
        detail = page(loaded, "/nhan-vien?ky=2025-01")
        assert "NV-A" in detail and "NV-B" in detail

    def test_h_an_unavailable_2025_metric_shows_a_dash_not_a_zero(
        self, loaded, repository
    ):
        """Fixture 2025 không có cột lương/thưởng ⟹ phải là `—`, không phải 0."""
        rows = repository.query_summary(2025, 1)
        availability = {
            item.rule.key: item
            for item in legacy_reference.summary_year_availability(rows)
        }
        assert availability["total_salary"].availability == \
            legacy_reference.NOT_AVAILABLE
        assert availability["total_salary"].filled_rows == 0
        cell = legacy_presentation.cell(rows[0], "total_salary", "kvnd")
        assert cell["text"] == "—" and cell["empty"] is True

    def test_h2_an_available_2025_metric_is_measured_from_the_rows(
        self, loaded, repository
    ):
        """Tính sẵn có được ĐO trên dòng thật, không phải hằng số viết tay."""
        rows = repository.query_summary(2025, 1)
        availability = {
            item.rule.key: item
            for item in legacy_reference.summary_year_availability(rows)
        }
        assert availability["sales"].availability == \
            legacy_reference.AVAILABLE_WITH_ACCEPTED_EVIDENCE
        assert availability["sales"].filled_rows == 3   # 2 người bán + 1 dòng tổng

    def test_i_only_total_sales_crosses_the_engine_boundary_for_2025(
        self, loaded, snapshots
    ):
        """`DEC-180` — Tổng bán so được; các chỉ tiêu khác của 2025 thì không."""
        persist(snapshots, [pair("BH1", kpi_purchase="5000000", kpi_profit="3000000")])
        html = page(loaded)
        assert "Không so được" in html, "bốn cặp còn lại vẫn bị chặn"
        allowed = legacy_reference.compare(
            Decimal("1120000"), Decimal("8000000"),
            legacy_key="sales", current_key="sales_revenue")
        assert allowed.allowed is True and allowed.percent is not None
        blocked = legacy_reference.compare(
            Decimal("1120000"), Decimal("8000000"),
            legacy_key="converted_revenue", current_key="converted_sales")
        assert blocked.allowed is False and blocked.percent is None


# ==========================================================================
# `DEC-178` — THẨM QUYỀN NGUỒN LỊCH SỬ 2025.
#
# Chủ dự án đã freeze: workbook lịch sử MỘT NĂM độc lập là nguồn CHUẨN của
# năm đó; bản sao `Summary <năm>` nhúng trong workbook năm hiện hành chỉ là
# bằng chứng THỨ CẤP. Lệch nhau ⟹ bản độc lập thắng.
#
# Hai fixture cố ý mang SỐ KHÁC NHAU cho cùng kỳ 01.2025 / NV-A:
#   build_year_workbook()   (nguồn chuẩn)  Tổng bán = 1.180.000
#   build_legacy_workbook() (bản thứ cấp)  Tổng bán = 1.120.000
# nên mọi bài test dưới đây đo được một quy tắc thật, không phải một trùng hợp.
# ==========================================================================

AUTHORITATIVE_SALES = Decimal("1180000")   # SOURCE_A — nguồn chuẩn
SNAPSHOT_SALES = Decimal("1120000")        # SOURCE_B — bản sao thứ cấp


@pytest.fixture
def year_workbook_path(tmp_path):
    return build_year_workbook(tmp_path / "bao_cao_2025_doc_lap.xlsx")


def upload_both(client, snapshot_path, year_path):
    """Nạp bản THỨ CẤP trước, rồi bản CHUẨN — thứ tự bất lợi nhất cho quy tắc.

    Nếu hệ thống chỉ chạy theo "ai ghi sau thì thắng", thứ tự này vẫn cho ra
    kết quả đúng một cách tình cờ; nên các bài test dưới đây kiểm cả thứ tự
    ngược lại.
    """
    upload(client, snapshot_path)
    upload(client, year_path)


class TestSourceAuthority2025:
    def test_a_the_standalone_year_workbook_is_authoritative(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        upload_both(client, legacy_workbook_path, year_workbook_path)
        authority = repository.authority_import_for_year(2025)
        assert authority is not None
        assert authority["source_authority"] == "AUTHORITATIVE_YEAR"
        assert "doc_lap" in authority["source_file_name"]

    def test_b_the_snapshot_cannot_overwrite_the_authoritative_source(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        """Nạp bản thứ cấp SAU bản chuẩn vẫn không đổi được số của 2025."""
        upload(client, year_workbook_path)
        before = repository.query_summary(2025, 1)
        upload(client, legacy_workbook_path)
        after = repository.query_summary(2025, 1)
        assert [r["sheet_name"] for r in after] == [r["sheet_name"] for r in before]
        assert all(row["sheet_name"] == "Summary" for row in after)

    def test_c_a_conflicting_value_resolves_to_the_authoritative_source(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        """Ô lệch nhau THẬT: 1.180.000 (chuẩn) và 1.120.000 (thứ cấp)."""
        upload_both(client, legacy_workbook_path, year_workbook_path)
        row = next(r for r in repository.query_summary(2025, 1)
                   if r["seller_label"] == "NV-A")
        assert row["sales"] == AUTHORITATIVE_SALES
        assert row["sales"] != SNAPSHOT_SALES

    def test_c2_the_same_holds_when_the_snapshot_is_uploaded_last(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        """Không phải "ai ghi sau thì thắng" — đảo thứ tự, kết quả không đổi."""
        upload(client, year_workbook_path)
        upload(client, legacy_workbook_path)
        row = next(r for r in repository.query_summary(2025, 1)
                   if r["seller_label"] == "NV-A")
        assert row["sales"] == AUTHORITATIVE_SALES

    def test_c3_no_value_is_ever_averaged_or_merged(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        """Không trộn: giá trị hiện ra phải là MỘT trong hai số nguồn."""
        upload_both(client, legacy_workbook_path, year_workbook_path)
        row = next(r for r in repository.query_summary(2025, 1)
                   if r["seller_label"] == "NV-A")
        midpoint = (AUTHORITATIVE_SALES + SNAPSHOT_SALES) / 2
        assert row["sales"] != midpoint
        assert row["sales"] in (AUTHORITATIVE_SALES, SNAPSHOT_SALES)

    def test_the_snapshot_is_still_readable_as_secondary_evidence(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        """Thứ cấp KHÔNG bị xoá — vẫn tra được khi chỉ đích danh bản nhập đó."""
        upload_both(client, legacy_workbook_path, year_workbook_path)
        snapshot = next(item for item in repository.list_imports()
                        if item["source_authority"] != "AUTHORITATIVE_YEAR")
        rows = repository.query_summary(
            2025, 1, import_id=snapshot["import_id"])
        row = next(r for r in rows if r["seller_label"] == "NV-A")
        assert row["sales"] == SNAPSHOT_SALES

    def test_the_year_workbook_does_not_steal_the_current_pointer(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        """Nạp workbook 2025 KHÔNG được làm biến mất các kỳ 2026."""
        upload_both(client, legacy_workbook_path, year_workbook_path)
        periods = repository.available_periods()
        assert (2026, 1) in periods, "kỳ 2026 phải còn nguyên"
        assert (2025, 1) in periods, "kỳ 2025 phải xuất hiện"
        assert "bao_cao" in repository.current_import()["source_file_name"]

    def test_the_year_workbook_alone_still_becomes_readable(
        self, client, year_workbook_path, repository
    ):
        """Chỉ có workbook 2025 và không có gì khác — vẫn phải đọc được."""
        upload(client, year_workbook_path)
        assert repository.query_summary(2025, 1)
        assert (2025, 1) in repository.available_periods()


class TestYearWorkbookStructure:
    def test_d_every_month_present_in_the_source_is_discovered(
        self, client, year_workbook_path, repository
    ):
        upload(client, year_workbook_path)
        months = sorted({m for y, m in repository.available_periods() if y == 2025})
        assert months == [1, 2]

    def test_e_detail_sheets_are_recorded_but_never_ingested(
        self, client, year_workbook_path, repository
    ):
        """`LEGACY_LINE_DETAIL_2025 = DEFERRED` — ghi tên, không đọc ô nào."""
        upload(client, year_workbook_path)
        current = repository.query_summary(2025)
        assert all(row["sheet_name"] == "Summary" for row in current)
        entries = repository.list_imports()[0]["sheets_imported"]
        deferred = [e for e in entries if e["scope"] == "DETAIL_NOT_INGESTED"]
        assert len(deferred) == 5
        assert all(e["imported_rows"] == "0" for e in deferred)

    def test_f_summary_employee_rows_are_recognised(self, client, year_workbook_path,
                                                    repository):
        upload(client, year_workbook_path)
        sellers = [r["seller_label"] for r in repository.query_summary(2025, 1)
                   if r["row_kind"] == "SELLER"]
        assert sellers == ["NV-A", "NV-B"]

    def test_g_month_total_rows_are_a_separate_row_kind(
        self, client, year_workbook_path, repository
    ):
        upload(client, year_workbook_path)
        kinds = [r["row_kind"] for r in repository.query_summary(2025, 1)]
        assert kinds.count("MONTH_TOTAL") == 1

    def test_h_a_total_row_is_never_treated_as_an_employee(
        self, client, year_workbook_path, repository
    ):
        upload(client, year_workbook_path)
        years = legacy_reference.summary_years(repository.query_all_summary())
        year = next(y for y in years if y.year == 2025)
        assert set(year.sellers) == {"NV-A", "NV-B", "NV-C"}
        for forbidden in ("Tổng", "Total", "Summary", None):
            assert forbidden not in year.sellers

    def test_i_supported_metrics_preserve_the_source_values(
        self, client, year_workbook_path, repository
    ):
        """Giá trị đi thẳng từ ô Excel ra, KHÔNG tính lại bằng công thức mới."""
        upload(client, year_workbook_path)
        row = next(r for r in repository.query_summary(2025, 1)
                   if r["seller_label"] == "NV-A")
        assert row["orders"] == Decimal("41")
        assert row["products"] == Decimal("111")
        assert row["sales"] == AUTHORITATIVE_SALES
        assert row["converted_revenue"] == Decimal("601000")
        assert row["profit"] == Decimal("33055")

    def test_j_a_blank_metric_stays_unavailable_not_zero(
        self, client, year_workbook_path, repository
    ):
        upload(client, year_workbook_path)
        row = next(r for r in repository.query_summary(2025, 1)
                   if r["seller_label"] == "NV-A")
        assert row["total_salary"] is None
        cell = legacy_presentation.cell(row, "total_salary", "kvnd")
        assert cell["text"] == "—" and cell["empty"] is True

    def test_the_kpi_recap_block_is_excluded_with_its_row_count_recorded(
        self, client, year_workbook_path, repository
    ):
        """Khối tổng kết KPI: cột mang nghĩa khác ⟹ loại trừ TƯỜNG MINH."""
        upload(client, year_workbook_path)
        entry = next(e for e in repository.list_imports()[0]["sheets_imported"]
                     if e["sheet_name"] == "Summary")
        assert entry["recap_rows_excluded"] == "3"
        # Và không dòng nào của khối đó lọt vào dữ liệu người bán.
        rows = repository.query_summary(2025)
        assert all(row["orders"] is None or row["orders"] > 1 or
                   row["row_kind"] != "SELLER" for row in rows)

    def test_a_progress_ratio_row_never_becomes_a_period(
        self, client, year_workbook_path, repository
    ):
        """Dòng tiến độ (`C = B/A`) có `month = NULL` ⟹ ngoài mọi khung nhìn kỳ."""
        upload(client, year_workbook_path)
        for _year, month in repository.available_periods():
            assert month is not None
        assert all(r["row_kind"] != "PROGRESS"
                   for r in repository.query_summary(2025, 1))

    def test_a_multi_year_workbook_fails_loudly_instead_of_being_guessed(self, tmp_path):
        from openpyxl import Workbook
        path = tmp_path / "hai_nam.xlsx"
        book = Workbook()
        book.active.title = "Summary"
        book.create_sheet("01.2024 NV-A")
        book.create_sheet("01.2025 NV-A")
        book.save(path)
        with pytest.raises(LegacyImportError) as exc:
            parse_year_workbook(path)
        assert "NHIỀU năm" in str(exc.value)

    def test_the_dec_168_guard_applies_to_the_authoritative_summary(self, tmp_path):
        """Nguồn CHUẨN là REQUIRED: một dòng không phân loại được vẫn FAIL TO."""
        stripped = strip_formula_markers(
            build_year_workbook(tmp_path / "mat_formula.xlsx"), sheet_name="Summary")
        with pytest.raises(LegacyImportError) as exc:
            parse_year_workbook(stripped)
        assert "Summary" in str(exc.value)


class TestOwnerNavigationForALegacyYear:
    def test_k_the_owner_can_reach_employee_metrics_for_a_2025_month(
        self, client, year_workbook_path
    ):
        upload(client, year_workbook_path)
        html = page(client)
        assert "Năm 2025" in html
        assert "/nhan-vien?ky=2025-01" in html
        detail = page(client, "/nhan-vien?ky=2025-01")
        assert "NV-A" in detail and "NV-B" in detail

    def test_l_legacy_values_stay_labelled_as_legacy_reference(
        self, client, year_workbook_path, repository
    ):
        upload(client, year_workbook_path)
        rows = repository.query_summary(2025)
        assert {r["origin"] for r in rows} == {"LEGACY_REFERENCE"}
        assert "SỐ CŨ" in page(client, "/nhan-vien?ky=2025-01")

    def test_the_page_names_the_source_of_each_year(self, client, year_workbook_path):
        upload(client, year_workbook_path)
        html = page(client)
        assert "Nguồn chuẩn của năm" in html
        assert "doc_lap" in html

    def test_the_page_says_detail_sheets_are_deliberately_deferred(
        self, client, year_workbook_path
    ):
        upload(client, year_workbook_path)
        html = page(client)
        assert "cố ý chưa nhập" in html
        assert "địa chỉ khách hàng" in html


class TestIdempotencyAndIsolationForTheYearWorkbook:
    def test_m_repeated_import_of_the_authoritative_source_is_idempotent(
        self, client, year_workbook_path, engine, repository
    ):
        upload(client, year_workbook_path)
        after_first = table_counts(engine)
        upload(client, year_workbook_path)
        assert table_counts(engine) == after_first
        assert repository.count_imports() == 1
        # Khoá DUY NHẤT thật của bảng là (import_id, sheet_name, sheet_row):
        # hai dòng "tiến độ" khác nhau đều có (year=2025, month=NULL), nên bộ
        # bốn trường nghiệp vụ KHÔNG phải khoá và không dùng để đo trùng lặp.
        keys = [(r["sheet_name"], r["sheet_row"])
                for r in repository.query_summary(2025)]
        assert len(keys) == len(set(keys))

    def test_n_repeated_snapshot_import_cannot_alter_authoritative_values(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        upload(client, year_workbook_path)
        for _ in range(3):
            upload(client, legacy_workbook_path)
        row = next(r for r in repository.query_summary(2025, 1)
                   if r["seller_label"] == "NV-A")
        assert row["sales"] == AUTHORITATIVE_SALES

    def test_o_p_q_current_engine_numbers_are_unchanged_by_the_2025_import(
        self, client, year_workbook_path, snapshots, service
    ):
        """O + P + Q: coverage, doanh thu, và giá nhập Owner nhập tay."""
        persist(snapshots, [pair("BH1", kpi_purchase=None, kpi_profit=None)])
        data = service.period(**JANUARY)
        first = data.details[0]
        service.store.set_purchase_price(
            order_key=first["order_key"], product_key=first["product_key"],
            occurrence_index=first["occurrence_index"],
            price=Decimal("6000000"), auto_price=None)
        before = service.period(**JANUARY).totals

        upload(client, year_workbook_path)
        after = service.period(**JANUARY).totals

        assert after.coverage == before.coverage
        assert after.sales_revenue == before.sales_revenue
        assert after.kpi_profit == before.kpi_profit
        assert after.kpi_profit == Decimal("2000000"), "giá Owner nhập vẫn còn"

    def test_r_no_product_identity_or_tracking_path_is_reachable_from_legacy(self):
        """Ranh giới bằng CẤU TRÚC: `app/legacy` không import các module đó."""
        import ast
        import pathlib

        forbidden = ("product", "pricing", "tracking", "profit", "kpi")
        for path in pathlib.Path("app/legacy").glob("*.py"):
            tree = ast.parse(path.read_text(encoding="utf-8"))
            for node in ast.walk(tree):
                names = []
                if isinstance(node, ast.Import):
                    names = [alias.name for alias in node.names]
                elif isinstance(node, ast.ImportFrom):
                    names = [node.module or ""]
                for name in names:
                    assert not any(f"modules.{token}" in name for token in forbidden), \
                        f"{path.name} import {name}"

    def test_t_the_2026_workbook_behaviour_is_unchanged(
        self, client, legacy_workbook_path, year_workbook_path, repository
    ):
        upload_both(client, legacy_workbook_path, year_workbook_path)
        rows = repository.query_summary(2026, 1)
        assert [r["seller_label"] for r in rows if r["row_kind"] == "SELLER"] == [
            "NV-A", "NV-B", "Kênh-1"]
        assert repository.query_daily(2026, 1), "DataChart 2026 vẫn đọc được"


class TestComparisonStaysBlockedForTheAuthoritativeSource:
    def test_s_being_the_authoritative_source_does_not_open_other_metrics(
        self, client, year_workbook_path
    ):
        """"Nguồn chuẩn lịch sử" KHÔNG đồng nghĩa "cùng nghĩa với số mới".

        `DEC-178` nói bản nhập nào THẮNG khi hai nguồn cùng nói về một năm.
        Nó không nói gì về việc chỉ tiêu nào so được với số mới — câu đó
        thuộc về `CROSS_ORIGIN_CONTRACT`, và `DEC-180` chỉ mở đúng Tổng bán.
        """
        upload(client, year_workbook_path)
        blocked = legacy_reference.compare(
            AUTHORITATIVE_SALES, Decimal("8000000"),
            legacy_key="profit", current_key="kpi_profit")
        assert blocked.allowed is False
        assert blocked.percent is None
        assert "Không so được" in page(client)
