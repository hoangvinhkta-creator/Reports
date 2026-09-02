"""Dựng workbook legacy tổng hợp (anonymized) cho test PRA-001.

Vì sao là fixture chứ không phải file thật: workbook "Báo cáo Kinh doanh
2026.xlsx" chứa dữ liệu cá nhân khách hàng và KHÔNG được commit
(``.gitignore``: ``data/samples/``). Fixture này tái tạo ĐÚNG hình dạng cấu
trúc mà importer phải xử lý — khối tháng có số dòng người bán khác nhau,
dòng tổng tháng, dòng tổng năm chia đôi — và cố ý cài sẵn bốn lỗi công thức
đã biết A1/A2/A4/A6 của ``docs/analysis/05_EXCEPTIONS.md``.

Giá trị ô và công thức được ghi ĐỘC LẬP với nhau: ô mang giá trị mà Excel đã
lưu lần tính cuối, công thức chỉ là văn bản. Nhờ vậy test chứng minh được
importer đọc GIÁ TRỊ NGUỒN chứ không tính lại từ công thức — xem ô F9 của
tháng 02 (giá trị 999 nhưng công thức là ``=G16/5.5%``).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

# (nhãn, tháng, sheet nguồn được tham chiếu, giá trị theo cột)
SELLERS_MONTH_01 = [
    ("NV-A", "01.2026 NV-A", {"C": 42, "D": 118, "E": 1_240_500, "F": 640_000,
                              "G": 35_200, "H": 0.0284, "I": 0.9, "J": 0.61,
                              "K": 33_100, "M": 1_300_000, "N": 0.49, "O": 3_200,
                              "P": 28, "Q": 4_846, "R": 780, "S": 8_826}),
    # A1 — số SP không nguyên (E1 sheet nguồn trừ đi một tỉ lệ phần trăm).
    ("NV-B", "01.2026 NV-B", {"C": 31, "D": 87.6, "E": 980_000, "F": 512_000,
                              "G": 28_160, "H": 0.0287, "I": 0.88, "J": 0.55,
                              "K": 26_900, "M": 1_300_000, "N": 0.39, "O": 2_560,
                              "P": 26, "Q": 4_500, "R": 780, "S": 7_840}),
    ("Kênh-1", "01.2026 Kênh-1", {"C": 0, "D": 1_517, "E": 14_452_000,
                                  "F": 14_452_000, "G": 289_040, "H": 0.02,
                                  "I": 1.08, "J": 0.72, "K": 275_000,
                                  "M": 12_000_000, "N": 1.2, "O": 72_260,
                                  "P": 28, "Q": 4_846, "R": 780, "S": 77_886}),
]

SELLERS_MONTH_02 = [
    ("NV-A", "02.2026 NV-A", {"C": 45, "D": 126, "E": 1_310_000, "F": 700_000,
                              "G": 38_500, "H": 0.0294, "I": 1.09, "J": 0.63,
                              "K": 36_400, "M": 1_300_000, "N": 0.54, "O": 3_500,
                              "P": 27, "Q": 4_673, "R": 780, "S": 8_953}),
    ("NV-B", "02.2026 NV-B", {"C": 34, "D": 94, "E": 1_020_000, "F": 999,
                              "G": 30_100, "H": 0.0295, "I": 1.04, "J": 0.57,
                              "K": 28_700, "M": 1_300_000, "N": 0.00077, "O": 5,
                              "P": 26, "Q": 4_500, "R": 780, "S": 5_285}),
    # A4 — nhãn dòng là "Kênh-1" nhưng số SP lấy từ sheet của NV-A.
    ("Kênh-1", "02.2026 NV-A", {"C": 0, "D": 1_602, "E": 15_100_000,
                                "F": 15_100_000, "G": 302_000, "H": 0.02,
                                "I": 1.04, "J": 0.74, "K": 288_000,
                                "M": 12_000_000, "N": 1.26, "O": 75_500,
                                "P": 28, "Q": 4_846, "R": 780, "S": 81_126}),
    ("NV-C", "02.2026 NV-C", {"C": 12, "D": 33, "E": 402_000, "F": 210_000,
                              "G": 11_550, "H": 0.0287, "I": 1.0, "J": 0.5,
                              "K": 10_900, "M": 900_000, "N": 0.23, "O": 1_050,
                              "P": 22, "Q": 3_808, "R": 660, "S": 5_518}),
]

SELLERS_MONTH_03 = [
    ("NV-A", "03.2026 NV-A", {"C": 48, "D": 133, "E": 1_402_000, "F": 745_000,
                              "G": 40_975, "H": 0.0292, "I": 1.06, "J": 0.64,
                              "K": 38_900, "M": 1_300_000, "N": 0.57, "O": 3_725,
                              "P": 28, "Q": 4_846, "R": 780, "S": 9_351}),
    ("NV-B", "03.2026 NV-B", {"C": 36, "D": 99, "E": 1_090_000, "F": 560_000,
                              "G": 30_800, "H": 0.0283, "I": 1.03, "J": 0.58,
                              "K": 29_400, "M": 1_300_000, "N": 0.43, "O": 2_800,
                              "P": 27, "Q": 4_673, "R": 780, "S": 8_253}),
]

# Ánh xạ cột → ô nguồn trên sheet người bán (docs/analysis/02_FORMULA_MAPPING.md §3).
_SOURCE_CELL = {"C": "B1", "D": "E1", "E": "H1", "G": "I1", "J": "C1", "K": "M1"}


def _write_seller_row(sheet, row: int, label: str, source_sheet: str,
                      values: dict, *, hardcoded_prev: int | None,
                      prev_row: int | None = None) -> None:
    sheet[f"B{row}"] = label
    for column, value in values.items():
        sheet[f"{column}{row}"] = value
    for column, source_cell in _SOURCE_CELL.items():
        sheet[f"{column}{row}"] = values[column]
    # Công thức ghi đè phần "văn bản" của ô ở bước ghi thứ hai bên dưới; ở
    # đây chỉ lưu lại để hàm gọi dựng sheet công thức song song.
    sheet._legacy_formulas = getattr(sheet, "_legacy_formulas", {})
    for column, source_cell in _SOURCE_CELL.items():
        sheet._legacy_formulas[f"{column}{row}"] = f"='{source_sheet}'!${source_cell[0]}${source_cell[1:]}"
    sheet._legacy_formulas[f"F{row}"] = f"=G{row}/5.5%"
    sheet._legacy_formulas[f"H{row}"] = f"=G{row}/E{row}"
    if hardcoded_prev is not None:
        # A6 — so kỳ trước bằng hằng số gõ tay.
        sheet._legacy_formulas[f"I{row}"] = f"=F{row}/{hardcoded_prev}"
    elif prev_row is not None:
        sheet._legacy_formulas[f"I{row}"] = f"=F{row}/F{prev_row}"
    sheet._legacy_formulas[f"N{row}"] = f"=IFERROR(F{row}/M{row},\"\")"
    sheet._legacy_formulas[f"S{row}"] = f"=IF(P{row}>0,SUM(O{row}+Q{row}+R{row}),\"\")"


def _write_month_total(sheet, row: int, first: int, last: int, values: dict,
                       *, narrow_last: int) -> None:
    sheet[f"B{row}"] = f"Tổng T{values['month']:02d}"
    for column, value in values["cells"].items():
        sheet[f"{column}{row}"] = value
    sheet._legacy_formulas = getattr(sheet, "_legacy_formulas", {})
    for column in ("E", "G", "K"):
        sheet._legacy_formulas[f"{column}{row}"] = f"=SUM({column}{first}:{column}{last})"
    # A2 — cột F cộng trên khoảng HẸP HƠN các cột khác của cùng dòng tổng.
    sheet._legacy_formulas[f"F{row}"] = f"=SUM(F{first}:F{narrow_last})"


def build_legacy_workbook(path: Path) -> Path:
    """Ghi workbook fixture ra ``path`` và trả về chính đường dẫn đó."""
    workbook = Workbook()
    summary_2026 = workbook.active
    summary_2026.title = "Summary 2026"
    summary_2025 = workbook.create_sheet("Summary 2025")
    summary_2025.sheet_state = "hidden"
    datachart = workbook.create_sheet("DataChart 2026")

    formulas: dict[str, dict[str, str]] = {}

    # --- Summary 2026 -------------------------------------------------
    # Dòng 3: tổng năm (SUM chia đôi vì vùng chứa cả dòng tổng tháng).
    summary_2026["A3"] = 365
    summary_2026["B3"] = 90
    summary_2026["C3"] = 0.2466
    summary_2026["E3"] = 21_998_250
    summary_2026["M3"] = 345_474_000
    summary_2026._legacy_formulas = {
        "C3": "=B3/A3",
        "E3": "=SUM(E4:E30)/2",
        "F3": "=E3/M3",
    }

    row = 4
    blocks = [
        (1, SELLERS_MONTH_01, 1_571_182),
        (2, SELLERS_MONTH_02, None),
        (3, SELLERS_MONTH_03, None),
    ]
    previous_row_by_label: dict[str, int] = {}
    for month, sellers, hardcoded in blocks:
        first = row
        for label, source_sheet, values in sellers:
            _write_seller_row(summary_2026, row, label, source_sheet, values,
                              hardcoded_prev=hardcoded,
                              prev_row=previous_row_by_label.get(label))
            previous_row_by_label[label] = row
            row += 1
        last = row - 1
        totals = {
            column: sum(values[column] for _, _, values in sellers)
            for column in ("E", "G", "K")
        }
        # Giá trị cột F của dòng tổng phản ánh đúng khoảng hẹp (thiếu dòng cuối).
        totals["F"] = sum(values["F"] for _, _, values in sellers[:-1])
        _write_month_total(summary_2026, row, first, last,
                           {"month": month, "cells": totals}, narrow_last=last - 1)
        row += 1

    # --- Summary 2025 (kỳ lịch sử, ít cột hơn) -------------------------
    summary_2025["B4"] = "NV-A"
    summary_2025["C4"] = 38
    summary_2025["D4"] = 104
    summary_2025["E4"] = 1_120_000
    summary_2025["F4"] = 580_000
    summary_2025["G4"] = 31_900
    summary_2025["B5"] = "NV-B"
    summary_2025["C5"] = 29
    summary_2025["D5"] = 80
    summary_2025["E5"] = 910_000
    summary_2025["F5"] = 470_000
    summary_2025["G5"] = 25_850
    summary_2025["B6"] = "Tổng T01"
    summary_2025["E6"] = 2_030_000
    summary_2025["F6"] = 1_050_000
    summary_2025["G6"] = 57_750
    summary_2025._legacy_formulas = {
        "C4": "='01.2025 NV-A'!$B$1", "D4": "='01.2025 NV-A'!$E$1",
        "E4": "='01.2025 NV-A'!$H$1", "G4": "='01.2025 NV-A'!$I$1",
        "C5": "='01.2025 NV-B'!$B$1", "D5": "='01.2025 NV-B'!$E$1",
        "E5": "='01.2025 NV-B'!$H$1", "G5": "='01.2025 NV-B'!$I$1",
        "E6": "=SUM(E4:E5)", "F6": "=SUM(F4:F5)", "G6": "=SUM(G4:G5)",
    }

    # --- DataChart 2026 ------------------------------------------------
    datachart["J15"] = 345_474_000
    datachart["P15"] = 987_068.57
    datachart["AJ2"] = 28_789_481_081
    datachart._legacy_formulas = {"P15": "=J15/350"}
    daily_by_month = {1: [820_000_000, 910_000_000, 1_050_000_000],
                      2: [760_000_000, 880_000_000],
                      3: [990_000_000]}
    prev_year = {1: 2_410_000_000, 2: 1_520_000_000, 3: 870_000_000}
    for month in range(1, 13):
        sheet_row = 2 + month
        amounts = daily_by_month.get(month, [])
        for day, amount in enumerate(amounts, start=1):
            datachart.cell(row=sheet_row, column=1 + day, value=amount)
        if amounts:
            datachart[f"AG{sheet_row}"] = sum(amounts)
            datachart[f"AH{sheet_row}"] = prev_year[month]
            datachart[f"AI{sheet_row}"] = round(sum(amounts) / prev_year[month], 4)
            datachart[f"AJ{sheet_row}"] = round(sum(amounts) / 28_789_481_081, 4)
            datachart._legacy_formulas[f"AG{sheet_row}"] = f"=SUM(B{sheet_row}:AF{sheet_row})"
            datachart._legacy_formulas[f"AI{sheet_row}"] = f"=AG{sheet_row}/AH{sheet_row}"
            datachart._legacy_formulas[f"AJ{sheet_row}"] = f"=AG{sheet_row}/$AJ$2"

    for sheet in (summary_2026, summary_2025, datachart):
        formulas[sheet.title] = dict(getattr(sheet, "_legacy_formulas", {}))

    path = Path(path)
    workbook.save(path)
    workbook.close()
    _overlay_formulas(path, formulas)
    return path


def _overlay_formulas(path: Path, formulas: dict[str, dict[str, str]]) -> None:
    """Ghi công thức vào bản "formula" và giá trị vào bản "cached value".

    openpyxl không giữ được cả hai mặt của một ô trong cùng một lần ghi. Cách
    duy nhất để có một file mà ``data_only=True`` trả GIÁ TRỊ còn
    ``data_only=False`` trả CÔNG THỨC là chèn cached value vào XML sau khi
    lưu — đúng hình dạng file Excel thật, và là điều kiện để test chứng minh
    importer không tính lại từ công thức.
    """
    import re
    import shutil
    import zipfile
    from openpyxl import load_workbook

    values_snapshot: dict[str, dict[str, object]] = {}
    workbook = load_workbook(path)
    for title, cells in formulas.items():
        sheet = workbook[title]
        values_snapshot[title] = {ref: sheet[ref].value for ref in cells}
        for ref, formula in cells.items():
            sheet[ref] = formula
    workbook.save(path)
    workbook.close()

    sheet_files = _sheet_xml_map(path)
    temporary = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            title = sheet_files.get(item.filename)
            if title is not None:
                data = _inject_cached_values(data, values_snapshot.get(title, {}))
            target.writestr(item, data)
    shutil.move(temporary, path)


def _sheet_xml_map(path: Path) -> dict[str, str]:
    import re
    import zipfile

    with zipfile.ZipFile(path) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
        rels = archive.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    target_by_id: dict[str, str] = {}
    for element in re.findall(r"<Relationship\b[^>]*/?>", rels):
        rel_id = re.search(r'Id="([^"]+)"', element)
        target = re.search(r'Target="([^"]+)"', element)
        if rel_id and target:
            # Target có thể tuyệt đối ("/xl/worksheets/sheet1.xml") hoặc
            # tương đối với thư mục xl/ ("worksheets/sheet1.xml").
            path_in_zip = target.group(1).lstrip("/")
            if not path_in_zip.startswith("xl/"):
                path_in_zip = f"xl/{path_in_zip}"
            target_by_id[rel_id.group(1)] = path_in_zip
    mapping: dict[str, str] = {}
    for name, rel_id in re.findall(r'<sheet name="([^"]+)"[^>]*?r:id="([^"]+)"', workbook_xml):
        if rel_id in target_by_id:
            mapping[target_by_id[rel_id]] = name
    return mapping


def _inject_cached_values(data: bytes, values: dict[str, object]) -> bytes:
    import re

    text = data.decode("utf-8")
    for ref, value in values.items():
        if value is None:
            continue
        # openpyxl ghi ô công thức thành `<c r="C4"><f>...</f><v /></c>` —
        # thẻ <v> rỗng tự đóng. Thay đúng thẻ đó bằng cached value.
        pattern = re.compile(
            r'(<c r="%s"[^>]*>\s*<f>[^<]*</f>\s*)(?:<v\s*/>|<v>[^<]*</v>)?'
            % re.escape(ref)
        )

        def _replace(match: "re.Match[str]") -> str:
            return f"{match.group(1)}<v>{value}</v>"

        text, count = pattern.subn(_replace, text, count=1)
        if count != 1:
            raise AssertionError(f"Không chèn được cached value cho ô {ref}")
    return text.encode("utf-8")


def strip_formula_markers(path: Path, sheet_name: str = "Summary 2025") -> Path:
    """Giữ nguyên GIÁ TRỊ nghiệp vụ, xoá sạch công thức khỏi ``sheet_name``.

    Tái tạo đúng case của Independent Review (FIND-PRA001-R01): một sheet
    Summary vẫn đầy đủ số của Owner nhưng không còn dấu hiệu công thức nào
    để parser bám vào phân loại dòng. Trước repair, cả sheet biến mất khỏi
    bản nhập mà verifier vẫn báo `mismatched=0`.

    Tham số ``sheet_name`` cho phép chĩa đúng case này vào một sheet
    REQUIRED_IMPORT (`Summary 2026`) — nơi guard DEC-168 PHẢI còn hiệu lực —
    thay vì chỉ vào sheet REFERENCE_ONLY (`Summary 2025`), nơi sau DEC-169
    hình dạng value-only là hợp lệ và không được làm import trượt.
    """
    from openpyxl import load_workbook

    # Đọc GIÁ TRỊ đã cache của riêng sheet đích trước...
    values_book = load_workbook(path, data_only=True)
    values = {
        cell.coordinate: cell.value
        for row in values_book[sheet_name].iter_rows()
        for cell in row
        if cell.value is not None
    }
    values_book.close()

    # ...rồi ghi đè vào bản CÓ công thức, chỉ trên sheet đó. Các sheet khác
    # giữ nguyên công thức để bài test cô lập đúng một sheet hỏng.
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None
    for coordinate, value in values.items():
        sheet[coordinate] = value
    workbook.save(path)
    workbook.close()
    return path
