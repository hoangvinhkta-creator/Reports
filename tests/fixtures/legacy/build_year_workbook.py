"""Workbook lịch sử MỘT NĂM độc lập (anonymized) — hình dạng của SOURCE_A.

Vì sao là fixture chứ không phải file thật: `Báo cáo Kinh doanh 2025.xlsx`
chứa **tên, số điện thoại và địa chỉ khách hàng** trên cả 74 sheet chi tiết,
nên nó KHÔNG được commit. Fixture này tái tạo đúng những đặc điểm cấu trúc mà
`parse_year_workbook()` phải xử lý, đo trên chính file thật ở phiên S119:

* sheet tổng hợp tên đúng là ``Summary`` — **không mang năm trong tên**, nên
  năm phải suy từ tên các sheet chi tiết;
* dòng người bán liên kết công thức chéo sheet tới ``MM.YYYY <Nhãn>``;
* dòng tổng tháng dùng ``SUM`` trên khối tháng ngay phía trên;
* dòng "tiến độ" ``C = B/A`` (số ngày đã qua ÷ số ngày trong tháng);
* khối tổng kết KPI cuối năm mở đầu bằng ô ``C = "Tổng KPI"``, nơi các cột
  mang ý nghĩa KHÁC HẲN bảng chính và vì vậy phải bị loại trừ;
* sheet chi tiết có cột dữ liệu cá nhân — ở đây để rỗng, chỉ giữ tiêu đề, vì
  fixture chỉ cần chứng minh sheet TỒN TẠI và KHÔNG được nhập.

Giá trị nghiệp vụ được chọn KHÁC với `build_legacy_workbook()` để hai nguồn
nói về cùng một kỳ 2025 mà lệch nhau — điều kiện để kiểm quy tắc thẩm quyền
nguồn của `DEC-178`.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from tests.fixtures.legacy.build_legacy_workbook import _overlay_formulas

YEAR = 2025

# (tháng, [(nhãn người bán, {cột: giá trị})]) — cố ý khác số của
# `build_legacy_workbook()` cho CÙNG kỳ 01.2025 / NV-A, NV-B.
MONTHS: dict[int, list[tuple[str, dict]]] = {
    1: [
        ("NV-A", {"C": 41, "D": 111, "E": 1_180_000, "F": 601_000, "G": 33_055,
                  "H": 0.028, "J": 0.6, "K": 31_400, "M": 1_270_000, "P": 24}),
        ("NV-B", {"C": 30, "D": 84, "E": 947_000, "F": 486_000, "G": 26_730,
                  "H": 0.0282, "J": 0.55, "K": 25_100, "M": 1_270_000, "P": 22}),
    ],
    2: [
        ("NV-A", {"C": 44, "D": 119, "E": 1_262_000, "F": 655_000, "G": 36_025,
                  "H": 0.0285, "J": 0.62, "K": 34_300, "M": 1_270_000, "P": 23}),
        ("NV-B", {"C": 33, "D": 91, "E": 1_004_000, "F": 519_000, "G": 28_545,
                  "H": 0.0284, "J": 0.57, "K": 27_100, "M": 1_270_000, "P": 21}),
        ("NV-C", {"C": 11, "D": 29, "E": 361_000, "F": 188_000, "G": 10_340,
                  "H": 0.0286, "J": 0.5, "K": 9_900, "M": 900_000, "P": 18}),
    ],
}

DETAIL_HEADER = [
    "Date", "Trans", "Mã Sản phẩm", "Số lượng", "Giá nhập TT", "Giá bán",
    "Tổng bán", "Lợi nhuận", "Giao hàng", "Chi phí giao", "Giá thực nhập",
    "Lợi nhuận gộp", "Tên khách hàng", "Số điện thoại", "Địa chỉ",
]

# Ô nguồn trên sheet chi tiết, giống ánh xạ của workbook thật.
_SOURCE_CELL = {"C": "B1", "D": "E1", "E": "H1", "G": "I1", "J": "C1", "K": "M1"}

RECAP_HEADER_ROW_LABEL = "Tổng KPI"


def build_year_workbook(path: Path, *, year: int = YEAR) -> Path:
    """Ghi workbook một-năm ra ``path`` và trả về chính đường dẫn đó."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    formulas: dict[str, dict[str, str]] = {}
    summary_formulas: dict[str, str] = {}

    summary["B1"] = "Seller"
    summary["C2"] = "Tổng đơn"
    summary["D2"] = "Tổng số SP"
    summary["E2"] = "Tổng bán"

    row = 4
    seller_rows: dict[str, list[int]] = {}
    for month in sorted(MONTHS):
        first = row
        for label, values in MONTHS[month]:
            sheet_name = f"{month:02d}.{year} {label}"
            summary[f"B{row}"] = label
            for column, value in values.items():
                summary[f"{column}{row}"] = value
            for column, source_cell in _SOURCE_CELL.items():
                summary_formulas[f"{column}{row}"] = (
                    f"='{sheet_name}'!${source_cell[0]}${source_cell[1:]}")
            summary_formulas[f"F{row}"] = f"=G{row}/5.5%"
            summary_formulas[f"H{row}"] = f"=G{row}/E{row}"
            summary_formulas[f"N{row}"] = f'=IFERROR(F{row}/M{row},"")'
            seller_rows.setdefault(label, []).append(row)
            row += 1
        last = row - 1

        # Dòng tổng tháng.
        for column in ("E", "F", "G", "K"):
            summary[f"{column}{row}"] = sum(
                values.get(column, 0) for _label, values in MONTHS[month])
            summary_formulas[f"{column}{row}"] = f"=SUM({column}{first}:{column}{last})"
        row += 1

        # Dòng tiến độ: C = B/A trên vùng NHÃN, không phải "Tổng đơn".
        summary[f"A{row}"] = 31
        summary[f"B{row}"] = 30
        summary[f"C{row}"] = 30 / 31
        summary_formulas[f"C{row}"] = f"=B{row}/A{row}"
        row += 2

    # Khối tổng kết KPI cuối năm — cột mang ý nghĩa khác hẳn bảng chính.
    recap_header = row
    summary[f"C{recap_header}"] = RECAP_HEADER_ROW_LABEL
    summary[f"D{recap_header}"] = "KPI trung bình"
    row += 1
    for label, rows_of_label in seller_rows.items():
        summary[f"B{row}"] = label
        summary[f"C{row}"] = 1.5
        summary[f"D{row}"] = 0.75
        summary_formulas[f"C{row}"] = "+".join(
            f"=N{r}" if i == 0 else f"N{r}"
            for i, r in enumerate(rows_of_label))
        summary_formulas[f"D{row}"] = f"=C{row}/{len(rows_of_label)}"
        row += 1
    formulas["Summary"] = summary_formulas

    # Sheet chi tiết: CHỈ tiêu đề. Fixture không cần dòng nào — điều đang được
    # chứng minh là sheet tồn tại, được ghi tên, và KHÔNG được nhập.
    for month in sorted(MONTHS):
        for label, _values in MONTHS[month]:
            sheet = workbook.create_sheet(f"{month:02d}.{year} {label}")
            for index, header in enumerate(DETAIL_HEADER, start=1):
                sheet.cell(row=2, column=index, value=header)

    path = Path(path)
    workbook.save(path)
    workbook.close()
    _overlay_formulas(path, formulas)
    return path
