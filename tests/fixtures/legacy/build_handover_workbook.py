"""Workbook provenance 2026 hình dạng THẬT, phủ tới kỳ bàn giao 08/2026.

Vì sao cần thêm một fixture nữa bên cạnh ``build_legacy_workbook()``: bản kia
dừng ở 03/2026, nên nó không chạm được ca bàn giao mà `DEC-181` §14 bắt phải
chứng minh — 08/2026 là kỳ lịch sử CUỐI, 09/2026 là kỳ số mới ĐẦU. Fixture
này giữ nguyên mọi đặc điểm cấu trúc mà ``parse_workbook()`` phải xử lý
(dòng người bán liên kết chéo sheet, dòng tổng tháng dùng ``SUM``, DataChart
theo ngày + tổng tháng, bản sao ``Summary 2025`` nhúng) và chỉ mở rộng SỐ KỲ.

Bản sao ``Summary 2025`` nhúng ở đây cố ý MANG SỐ KHÁC với workbook 2025 độc
lập (``build_year_workbook``) cho cùng kỳ 01/2025 — điều kiện để đo được
rằng bản nhúng KHÔNG BAO GIỜ ghi đè nguồn chuẩn (`DEC-178`, `DEC-181` §1).

File thật của chủ dự án chứa dữ liệu cá nhân khách hàng nên KHÔNG được
commit; đây là bản anonymized tái tạo hình dạng, không phải bản sao dữ liệu.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from tests.fixtures.legacy.build_legacy_workbook import (
    _overlay_formulas, _write_month_total, _write_seller_row,
)

YEAR = 2026

# Số của bản sao `Summary 2025` NHÚNG — cố ý lệch bản chuẩn 01/2025 của
# `build_year_workbook()` (NV-A: 1.180.000) để quy tắc thẩm quyền đo được.
EMBEDDED_2025_NVA_SALES = 1_120_000
EMBEDDED_2025_MONTH_TOTAL = 2_030_000

# Tổng bán (kVND) của dòng TỔNG THÁNG từng kỳ trên `Summary 2026`.
# 08/2026 = kỳ bàn giao: đây là con số mà MoM của 09/2026 phải so vào.
MONTH_TOTAL_SALES_KVND = {
    1: 2_220_500, 2: 2_330_000, 3: 2_492_000, 4: 2_510_000,
    5: 2_604_000, 6: 2_688_000, 7: 2_742_000, 8: 1_000,
}

# Doanh số theo ngày của DataChart (VND nguyên), và tổng tháng suy từ chính
# các ngày đó — KHÔNG gõ một tổng rời rạc, để fixture không tự mâu thuẫn.
DATACHART_DAILY_VND = {
    month: [820_000_000 + month * 1_000_000, 910_000_000 + month * 1_000_000]
    for month in range(1, 9)
}
DATACHART_PREV_YEAR_VND = {month: 1_500_000_000 + month * 10_000_000
                           for month in range(1, 9)}


def _seller_values(month: int, index: int) -> dict:
    """Giá trị một dòng người bán — khác nhau theo kỳ để không có kỳ nào
    trùng khít kỳ khác (một phép so nhầm kỳ sẽ lộ ra ngay)."""
    base = 1_000_000 + month * 10_000 + index * 200_000
    return {
        "C": 40 + month + index, "D": 110 + month + index, "E": base,
        "F": base // 2, "G": base // 40, "H": 0.028, "I": 1.0,
        "J": 0.6, "K": base // 42, "M": 1_300_000, "N": 0.5,
        "O": 3_200, "P": 28, "Q": 4_846, "R": 780, "S": 8_826,
    }


def build_handover_workbook(
    path: Path, *, summary_months=tuple(range(1, 9)),
    datachart_months=tuple(range(1, 9)),
    embed_2025: bool = True,
) -> Path:
    """Ghi workbook 2026 ra ``path`` và trả về chính đường dẫn đó.

    ``summary_months`` / ``datachart_months`` tách nhau để dựng được ca
    fallback của `DEC-180` §13: một kỳ chỉ có bằng chứng DataChart, không có
    dòng tổng tháng trên Summary.
    """
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary 2026"
    datachart = workbook.create_sheet("DataChart 2026")
    embedded = workbook.create_sheet("Summary 2025") if embed_2025 else None
    formulas: dict[str, dict[str, str]] = {}

    # --- Summary 2026 -------------------------------------------------
    row = 4
    previous_row_by_label: dict[str, int] = {}
    for month in summary_months:
        first = row
        for index, label in enumerate(("NV-A", "NV-B")):
            _write_seller_row(
                summary, row, label, f"{month:02d}.{YEAR} {label}",
                _seller_values(month, index), hardcoded_prev=None,
                prev_row=previous_row_by_label.get(label),
            )
            previous_row_by_label[label] = row
            row += 1
        last = row - 1
        totals = {
            "E": MONTH_TOTAL_SALES_KVND[month],
            "F": MONTH_TOTAL_SALES_KVND[month] // 2,
            "G": MONTH_TOTAL_SALES_KVND[month] // 40,
            "K": MONTH_TOTAL_SALES_KVND[month] // 42,
        }
        _write_month_total(summary, row, first, last,
                           {"month": month, "cells": totals}, narrow_last=last)
        row += 1

    # --- Summary 2025 nhúng (bản sao THỨ CẤP) --------------------------
    if embedded is not None:
        embedded.sheet_state = "hidden"
        embedded["B4"] = "NV-A"
        embedded["C4"] = 38
        embedded["D4"] = 104
        embedded["E4"] = EMBEDDED_2025_NVA_SALES
        embedded["B5"] = "Tổng T01"
        embedded["E5"] = EMBEDDED_2025_MONTH_TOTAL
        embedded._legacy_formulas = {
            "C4": "='01.2025 NV-A'!$B$1", "D4": "='01.2025 NV-A'!$E$1",
            "E4": "='01.2025 NV-A'!$H$1",
            "E5": "=SUM(E4:E4)",
        }

    # --- DataChart 2026 ------------------------------------------------
    datachart["J15"] = 345_474_000
    datachart["P15"] = 987_068.57
    datachart["AJ2"] = 28_789_481_081
    datachart._legacy_formulas = {"P15": "=J15/350"}
    for month in datachart_months:
        sheet_row = 2 + month
        amounts = DATACHART_DAILY_VND[month]
        for day, amount in enumerate(amounts, start=1):
            datachart.cell(row=sheet_row, column=1 + day, value=amount)
        prev_year = DATACHART_PREV_YEAR_VND[month]
        datachart[f"AG{sheet_row}"] = sum(amounts)
        datachart[f"AH{sheet_row}"] = prev_year
        datachart[f"AI{sheet_row}"] = round(sum(amounts) / prev_year, 4)
        datachart[f"AJ{sheet_row}"] = round(sum(amounts) / 28_789_481_081, 4)
        datachart._legacy_formulas[f"AG{sheet_row}"] = \
            f"=SUM(B{sheet_row}:AF{sheet_row})"
        datachart._legacy_formulas[f"AI{sheet_row}"] = \
            f"=AG{sheet_row}/AH{sheet_row}"
        datachart._legacy_formulas[f"AJ{sheet_row}"] = \
            f"=AG{sheet_row}/$AJ$2"

    for sheet in (summary, datachart) + ((embedded,) if embedded else ()):
        formulas[sheet.title] = dict(getattr(sheet, "_legacy_formulas", {}))

    path = Path(path)
    workbook.save(path)
    workbook.close()
    _overlay_formulas(path, formulas)
    return path
