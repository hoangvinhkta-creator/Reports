"""GB-3 — biến workbook production thật thành Golden fixture đã ẩn danh.

**Chạy TAY, đúng MỘT lần, ngoài CI.** Script này nhận đường dẫn tới file thô
production do Owner cung cấp. File thô KHÔNG BAO GIỜ nằm trong repository
(DEC-108, `.gitignore`, `governance/product/17_DATA_GOVERNANCE_PRIVACY.md`);
chỉ **kết quả** — fixture đã ẩn danh — được commit.

    python3 -m tests.fixtures.golden.anonymize <raw.xlsx> <period-label> <out.xlsx>

Owner Decision `OD-GB-1 = A + A1` (xem PLAN §Phần B).

## MINIMIZE trước, ANONYMIZE sau

Nguyên tắc: một trường chỉ được giữ nếu **business logic thật sự đọc nó**.
Việc "đọc" được xác định bằng đường code, không bằng cảm tính:

    note_raw     -> LeadSourceClassifier._note_matches_ads()   ĐỌC
    product_raw  -> rules.is_non_product_line() qua matches_any() ĐỌC
                    + ProductGroupProvider.classify()
    employee_raw -> EmployeeMapper.resolve() (khớp prefix)      ĐỌC
    Ngày/SL/Đơn giá/Doanh số/Chiết khấu/Lương chuyến/Lợi nhuận  ĐỌC
    customer, customer_code, address, phone, shipper_raw, imei
                 -> chỉ được `normalizer` chép qua `WorkingLine`,
                    KHÔNG rule nào đọc                          KHÔNG ĐỌC

Vì vậy `address`, `phone`, `shipper_raw`, `imei` bị **XOÁ HẲN**, không thay
bằng surrogate: giữ surrogate cho một trường không ai đọc chỉ làm phình
fixture. `customer`/`customer_code` giữ surrogate vì hai lý do cụ thể — bảo
toàn *lực lượng* (cardinality) và quan hệ **không** 1-1 giữa mã và tên (đo
được: 06.2026 có 135 tên nhưng 133 mã), và giữ fixture ở hình dạng
production-realistic theo `V4_1_POLICY_FREEZE.md` §5.

Việc xoá bốn trường trên **không** làm dịch chuyển nghiệp vụ. Chứng minh
không bằng lập luận mà bằng phép đo, và phép đo đó là một test:
`test_golden_anonymization_preserves_business_output` chạy pipeline trên bản
gốc và trên fixture rồi so từng trường nghiệp vụ. Test đó chỉ chạy được khi
Owner cung cấp lại file thô (`GOLDEN_RAW_01`/`GOLDEN_RAW_06`), và tự SKIP khi
không có — nó là bằng chứng E1 của phiên tạo fixture, không phải cổng CI.

## `Diễn giải` — A1

`Diễn giải` là văn xuôi tự do người nhập gõ, và trên dữ liệu thật nó **có**
chứa tên và số điện thoại khách. Business logic hỏi nó đúng một câu: *chuỗi
viết hoa có chứa từ khoá ADS không?* (`config/lead_source.yaml`). Nên fixture
giữ đúng câu trả lời đó và không giữ gì khác:

    normalize rỗng      -> ""          (giữ nguyên trạng thái "trống")
    có chứa từ khoá ADS -> "ADS"
    còn lại             -> "BAN_HANG"

Nhãn được tính bằng **chính** `LeadSourceClassifier` production, không bằng
một bản `in` viết lại — nên fixture không thể lệch khỏi ngữ nghĩa production.

Trên hai kỳ này, số dòng chứa "ADS" đo được là **0/351** và **0/180**, khớp
`docs/analysis/_evidence/evidence.json` → `ads_keyword_cell_hits` = 0 và
DEC-109. Hệ quả trung thực: Golden **không** phủ nhánh ADS-qua-từ-khoá, vì
dữ liệu thật không có dòng nào đi qua nhánh đó. Nhánh đó do
`tests/test_pipeline.py::test_order_with_ads_line_propagates_to_all_lines`
phủ trên fixture tổng hợp. KHÔNG bịa thêm một dòng ADS vào fixture — làm vậy
là chế tạo dữ liệu nghiệp vụ.

## Giữ nguyên văn, có chủ đích

- `Số BH` — số chứng từ nội bộ, không phải dữ liệu cá nhân. Giữ để bảo toàn
  danh tính đơn (`order_graph`, đếm OrderID duy nhất) và để Owner truy được
  một dòng Golden ngược về đơn thật khi regression nổ.
- `NVBH` — master data nhân viên, đã có sẵn trong `config/employees.yaml` và
  `tests/fixtures/baseline/`. `EmployeeMapper` khớp theo prefix trên chuỗi
  **thô**, nên đổi chuỗi này là phá `I-11`.
- `Tên hàng trên chứng từ` — tên model sản phẩm, đầu vào của rule dòng phụ.
  Đã quét: 0 chuỗi chứa pattern số điện thoại, 0 chuỗi chứa tên khách, 0
  chuỗi chứa địa chỉ, ở cả hai kỳ.
- Dòng `Tổng cộng` cuối sheet — do chính ERP ghi, KHÔNG do engine tính. Nó là
  oracle độc lập duy nhất có (`I-02`), nên bắt buộc giữ.

Surrogate là **deterministic**: đánh số theo thứ tự xuất hiện đầu tiên trong
file. Cùng một file vào ⇒ cùng một fixture ra. Bảng ánh xạ ngược (surrogate →
giá trị thật) **không bao giờ** được ghi ra đĩa.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[3]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.modules.importing.raw_reader import COLUMNS, FIRST_DATA_ROW  # noqa: E402
from app.modules.lead_source.classifier import LeadSourceClassifier  # noqa: E402
from app.modules.validation.text import normalize_text  # noqa: E402

#: Tăng khi quy tắc biến đổi đổi. Ghi vào provenance của expected output, nên
#: một fixture sinh bằng luật cũ không thể im lặng đi cùng expected output mới.
ANONYMIZATION_VERSION = "1.0.0"

CONFIG_DIR = REPO_ROOT / "config"

#: Cột bị xoá hẳn — không rule nào đọc chúng (xem docstring).
DROPPED_COLUMNS = ("address", "phone", "shipper", "imei")

#: Cột được thay bằng surrogate deterministic, giữ lực lượng và quan hệ.
SURROGATE_COLUMNS = {"customer": "CUSTOMER", "customer_code": "CUSTCODE"}

#: Cột giữ nguyên văn.
VERBATIM_COLUMNS = (
    "date", "order_id", "product", "qty", "unit_price",
    "sales", "discount", "employee", "trip_pay", "profit",
)

_TOTAL_ROW_LABEL = "Tổng cộng"


class _Surrogates:
    """Ánh xạ giá trị thật -> nhãn thay thế, đánh số theo thứ tự xuất hiện.

    Bảng chỉ sống trong bộ nhớ của lần chạy này và không bao giờ được ghi ra
    đĩa: một reverse map đã commit thì tương đương commit chính dữ liệu gốc.
    """

    def __init__(self, prefix: str) -> None:
        self._prefix = prefix
        self._seen: dict[str, str] = {}

    def __call__(self, value: object) -> Optional[str]:
        text = normalize_text(value)
        if not text:
            return None
        if text not in self._seen:
            self._seen[text] = f"{self._prefix}_{len(self._seen) + 1:04d}"
        return self._seen[text]

    def __len__(self) -> int:
        return len(self._seen)


def _note_label(classifier: LeadSourceClassifier, value: object) -> str:
    """Nhãn A1 cho `Diễn giải`, quyết định bằng CHÍNH classifier production."""
    if not normalize_text(value):
        return ""
    # `classify_auto` trả "Auto:ADS Rule" khi và chỉ khi từ khoá khớp. Dùng nó
    # thay vì viết lại phép so — một bản viết lại sẽ trôi khỏi production.
    verdict = classifier.classify_auto([value], None, None)
    return "ADS" if verdict.source_of_value == "Auto:ADS Rule" else "BAN_HANG"


def anonymize_workbook(raw_path: Path, period_label: str, out_path: Path) -> dict:
    """Đọc workbook thật, ghi fixture đã ẩn danh, trả thống kê để ghi provenance.

    Không đọc bất kỳ giá trị nào ra ngoài hàm này ngoài các con số đếm.
    """
    classifier = LeadSourceClassifier.from_yaml(CONFIG_DIR / "lead_source.yaml")
    surrogates = {
        name: _Surrogates(prefix) for name, prefix in SURROGATE_COLUMNS.items()
    }

    source = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    try:
        sheet = source.active
        rows = list(sheet.iter_rows(values_only=True))
        sheet_title = sheet.title
    finally:
        source.close()

    out = openpyxl.Workbook()
    target = out.active
    target.title = sheet_title

    # Dòng 1..FIRST_DATA_ROW-1: khối tiêu đề. Giữ nguyên — nó mang bố cục
    # (header ở dòng 4, dòng 5 là header phụ) mà `raw_reader` phụ thuộc.
    for values in rows[: FIRST_DATA_ROW - 1]:
        target.append(list(values))

    data_rows = 0
    total_rows = 0
    ads_labels = 0
    for values in rows[FIRST_DATA_ROW - 1:]:
        if all(v is None for v in values):
            continue
        row = list(values)

        first = normalize_text(row[0]) if row else ""
        if first == _TOTAL_ROW_LABEL:
            # Dòng "Tổng cộng" của chính ERP — oracle độc lập (I-02). Nó chỉ
            # chứa số tổng, không chứa dữ liệu cá nhân, nên đi qua nguyên vẹn.
            target.append(row)
            total_rows += 1
            continue

        for name in DROPPED_COLUMNS:
            row[COLUMNS[name]] = None
        for name, surrogate in surrogates.items():
            row[COLUMNS[name]] = surrogate(row[COLUMNS[name]])
        label = _note_label(classifier, row[COLUMNS["note"]])
        row[COLUMNS["note"]] = label
        if label == "ADS":
            ads_labels += 1

        target.append(row)
        data_rows += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out.save(out_path)

    return {
        "period_label": period_label,
        "anonymization_version": ANONYMIZATION_VERSION,
        "data_rows": data_rows,
        "total_rows_kept": total_rows,
        "distinct_customer_surrogates": len(surrogates["customer"]),
        "distinct_customer_code_surrogates": len(surrogates["customer_code"]),
        "rows_labelled_ads": ads_labels,
        "dropped_columns": list(DROPPED_COLUMNS),
        "surrogate_columns": sorted(SURROGATE_COLUMNS),
        "verbatim_columns": list(VERBATIM_COLUMNS),
    }


def main(argv: list[str]) -> int:
    if len(argv) != 4:
        print(__doc__)
        print("Usage: python3 -m tests.fixtures.golden.anonymize "
              "<raw.xlsx> <period-label> <out.xlsx>")
        return 2
    stats = anonymize_workbook(Path(argv[1]), argv[2], Path(argv[3]))
    for key, value in stats.items():
        print(f"{key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
