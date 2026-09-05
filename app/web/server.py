"""Reports Web Shared Online Beta (S071) — thin server-side adapter.

Kiến trúc bắt buộc: Browser → tầng này → ``app.owner_usability``/``app.demo``
(đường production đã accepted) → artifact. Module này KHÔNG tính lại business
rule, KHÔNG tự phân loại AUTO/Review/Product Identity/PP — chỉ gọi đúng
adapter đã có (``run_owner_report``, cùng đường ``owner_launcher.py`` Tkinter
dùng) và trình bày lại ``ReportSummary`` đã authoritative.

S071 thay đổi so với S070 (Web Beta V1, local-only):

1. Run registry đổi từ ``dict`` process-local sang SQLite persistent
   (``app.web.run_registry``) — sống qua restart, chia sẻ được giữa nhiều
   viewer/nhiều worker process cùng phục vụ ``reports.tinphatcrm.com``.
2. Khi ``TRACKING_REPORT_SOURCE_URL``/``TRACKING_REPORT_API_KEY`` được cấu
   hình (triển khai cloud), mỗi lần chạy PULL LIVE từ Tracking
   (``tools.tracking.live_pull``) thay vì đọc capture cục bộ trên máy Owner —
   Owner Mac không còn nằm trên critical path. Khi CHƯA cấu hình (môi trường
   phát triển/local Owner), hành vi local S068–S070 giữ nguyên không đổi.
3. Thêm trang lịch sử run (``/history``) — sếp mở web thấy cùng run mới nhất
   và lịch sử mà Owner thấy, không phụ thuộc session trình duyệt của ai.

S071B thay tiếp SQLite + đĩa persistent bằng Cloudflare R2 khi đã cấu hình
(``app.web.storage_backend``, ``tools.storage.r2_store``) để runtime STATELESS
— xem ``docs/deployment/S071_DEPLOYMENT.md``. Module này KHÔNG biết đang chạy
trên backend nào; toàn bộ đọc/ghi đi qua ``store`` (``storage_backend.RunStore``).

Trust boundary: browser chỉ gửi workbook + lựa chọn feedback; browser không
bao giờ nhận secret, raw Tracking payload, hay đường dẫn filesystem tuyệt đối.
Download chỉ được resolve từ ``run_id`` qua registry do chính server tạo.
"""

from __future__ import annotations

import os
import time
import uuid
from dataclasses import replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from flask import Flask, abort, redirect, render_template, request, url_for
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge

from app import beta_feedback, beta_telemetry
from app.beta_presentation import REASON_DISPLAY_LABELS
from app.legacy import (
    LegacyImportError, is_standalone_year_workbook, parse_workbook,
    parse_year_workbook,
)
from app.legacy.models import SOURCE_AUTHORITY_YEAR
from app.owner_usability import (
    OwnerUsabilityError, run_owner_report, select_latest_valid_captures,
)
from app.owner_usability import SelectedCaptures
from app.history import coverage as history_coverage
from app.history import models as history_models
from app.modules.reporting import business_metrics
from app.modules.reporting.rate_routing import gia_dung_workflow_applies
from app.web import (
    analytics_presentation, analytics_queries, business_presentation,
    business_service, business_store, history_store, history_writer,
    legacy_presentation, legacy_reference, run_registry, sales_presentation,
    sales_queries, storage_backend,
)
import tools.db as history_db
from tools.db import HistoryConfigurationError
from tools.storage.errors import CorruptRunRecordError, StorageUnavailableError
from tools.tracking import live_pull

REPO_ROOT = Path(__file__).resolve().parents[2]

# S071 Deployment Gate: một số nhà cung cấp hosting managed (vd Render Web
# Service) chỉ cho gắn ĐÚNG MỘT persistent disk trên mỗi service — registry
# SQLite và artifact/upload/tracking-tạm phải cùng nằm dưới một gốc mount
# duy nhất để cả hai sống qua restart/redeploy. `REPORTS_DATA_ROOT` cho phép
# production trỏ toàn bộ state runtime vào gốc disk đó; mặc định (không đặt
# biến này) giữ NGUYÊN các đường dẫn tương đối REPO_ROOT đã dùng từ S070 —
# không đổi hành vi local/test nào.
DATA_ROOT = Path(os.environ.get("REPORTS_DATA_ROOT") or REPO_ROOT)
UPLOAD_DIR = DATA_ROOT / "data" / "uploads"
ARTIFACT_DIR = (DATA_ROOT / "outputs" / "reports").resolve()
TRACKING_TEMP_DIR = DATA_ROOT / "data" / "tracking_live_tmp"

# Beta technical safety limit — không phải quyết định nghiệp vụ của Owner.
# Workbook kế toán thật (mẫu đã audit) nằm dưới vài MB; 25MB để dư biên độ.
MAX_UPLOAD_BYTES = 25 * 1024 * 1024

HISTORY_PAGE_LIMIT = 50

LEGACY_IMPORT_PAGE_LIMIT = 50

# `R3` — hai provenance có nghĩa "chính Owner đã quyết con số này". Đọc thẳng
# từ `business_metrics` để bộ lọc không bao giờ lệch với ngữ nghĩa đã freeze.
_OWNER_EDITED_PROVENANCE = (
    business_metrics.PROVENANCE_MANUAL,
    business_metrics.PROVENANCE_MANUAL_OVERRIDE,
)


def _guarded(fn, *args, **kwargs):
    """Gọi ``fn`` (một lời gọi store) và biến lỗi storage (R2 unavailable/
    JSON hỏng) thành HTTP 503 rõ ràng — KHÔNG âm thầm coi như "không tìm
    thấy"/"lịch sử rỗng" (S071B: phải phân biệt lỗi storage với dữ liệu
    thật sự rỗng)."""
    try:
        return fn(*args, **kwargs)
    except (StorageUnavailableError, CorruptRunRecordError,
            history_store.HistoryUnavailableError):
        abort(503)


def _readiness_text() -> str:
    """Đúng nguyên semantics/wording đã review truthful ở owner_launcher.py.

    Chế độ cloud (live pull cấu hình xong) luôn "sẵn sàng" theo nghĩa cục bộ
    — sẵn sàng thật hay không quyết định LÚC CHẠY (mỗi lần fetch Tracking),
    không phải bằng cách quét capture cũ trên đĩa máy chủ.
    """
    if live_pull.is_configured():
        return "Sẵn sàng — dữ liệu Tracking lấy trực tiếp (live) mỗi lần chạy"
    try:
        select_latest_valid_captures()
    except Exception:
        return "Chưa sẵn sàng"
    return "Có capture hợp lệ trên máy"


def _review_reason_lines(counts: dict[str, int]) -> list[tuple[str, int]]:
    return [
        (REASON_DISPLAY_LABELS.get(reason, reason), count)
        for reason, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    ]


def _build_view(summary, *, dropped_lines: int = 0) -> dict:
    return {
        "input_orders": summary.input_orders,
        "auto_orders": summary.auto_orders,
        "review_orders": summary.review_orders,
        "error_count": summary.error_count,
        "dropped_lines": dropped_lines,
        "accounting_rate": f"{summary.order_accounting_rate:.0%}",
        "review_reason_lines": _review_reason_lines(summary.review_reason_counts),
    }


def _record_telemetry(run_id: str, summary, duration_ms: int) -> None:
    # Byproduct vận hành, không phải core path — không chặn Owner nếu ghi lỗi.
    try:
        record = beta_telemetry.build_run_record(
            run_id=run_id, summary=summary, processing_duration_ms=duration_ms,
        )
        beta_telemetry.record_run(record)
    except Exception:
        pass


def _safe_display_name(filename: str) -> str:
    """Chỉ giữ basename để hiển thị — không bao giờ đường dẫn client gửi lên
    (chống lộ cấu trúc thư mục máy khách, không liên quan tên file lưu trên
    đĩa server — tên đó luôn do server sinh, xem ``/run``)."""
    name = Path(filename).name or "workbook.xlsx"
    return name[:120]


def _selected_period(periods: list[tuple[int, Optional[int]]]) -> Optional[tuple[int, Optional[int]]]:
    """Kỳ đang xem: lấy từ query string nếu hợp lệ, nếu không thì kỳ mới nhất.

    Chỉ chấp nhận kỳ THỰC SỰ có trong dữ liệu đã nhập — một kỳ do người dùng
    gõ tay mà không có dữ liệu sẽ rơi về ``None`` để trang hiện trạng thái
    rỗng trung thực, thay vì một bảng toàn số 0.
    """
    raw = request.args.get("ky") or ""
    if not raw:
        return periods[0] if periods else None
    year_text, _, month_text = raw.partition("-")
    try:
        chosen = (int(year_text), int(month_text) if month_text else None)
    except ValueError:
        return None
    return chosen if chosen in periods else None


def _select_captures_for_run() -> tuple[Optional[SelectedCaptures], Optional[dict], Optional[live_pull.LiveSelectedCaptures]]:
    """Trả về ``(captures, tracking_evidence, live_handle)``.

    ``live_handle`` khác ``None`` khi captures đến từ live pull — caller phải
    gọi ``live_handle.cleanup()`` sau khi dùng xong (finally), bất kể thành
    công hay lỗi, để không giữ authority thô của Tracking lâu hơn một lần
    chạy trên đĩa máy chủ (S071 §10).
    """
    if not live_pull.is_configured():
        return None, None, None
    live = live_pull.pull_live_captures(out_dir=TRACKING_TEMP_DIR)
    captures = SelectedCaptures(
        tracking_capture=live.tracking_capture,
        tracking_catalog=live.tracking_catalog,
        tracking_inv_map=live.tracking_inv_map,
    )
    return captures, live.evidence, live


def _build_history(env=None) -> Optional[history_store.LegacyRepository]:
    """Dựng repository history nếu môi trường đã sẵn sàng.

    ``REPORTS_REQUIRE_HISTORY_DB=1`` (production) → lỗi cấu hình được ném
    tiếp ra ngoài và app KHÔNG khởi động: thà chết lúc khởi động còn hơn
    chạy lên rồi hiển thị lịch sử rỗng trong khi thật ra không có database.
    Ngoài chế độ đó (máy dev chưa `alembic upgrade head`), trả ``None`` và
    các trang legacy nói thẳng là history store CHƯA cấu hình — vẫn không
    bao giờ giả vờ "chưa có dữ liệu".
    """
    values = os.environ if env is None else env
    required = (values.get("REPORTS_REQUIRE_HISTORY_DB") or "").strip() == "1"
    if not required and not (values.get("HISTORY_DATABASE_URL") or "").strip():
        # Chưa cấu hình gì: chỉ dùng SQLite mặc định nếu database ĐÃ được
        # tạo bằng `alembic upgrade head`. Không tự tạo file/thư mục database
        # khi khởi động — một database rỗng do app tự sinh ra sẽ khiến trang
        # legacy trông như "chưa nhập gì" trong khi thật ra chưa ai migrate.
        if not history_db.default_sqlite_path(values).exists():
            return None
    try:
        return history_store.build(values)
    except HistoryConfigurationError:
        if required:
            raise
        return None


SNAPSHOT_PAGE_LIMIT = 50

FLAG_PAGE_LIMIT = 200


def _build_snapshots(
    legacy: Optional[history_store.LegacyRepository],
) -> Optional[history_store.SnapshotRepository]:
    """Repository PRA-002 trên ĐÚNG engine mà LegacyRepository đang dùng.

    Dùng chung một ``Engine`` là điều kiện để lịch sử snapshot và bản nhập
    legacy nằm trong cùng một database — hai origin tách bảng, không tách nơi
    lưu. Không có history store thì cũng không có snapshot: ``None``, và tầng
    trên nói thẳng là run không được lưu lịch sử.
    """
    if legacy is None:
        return None
    return history_store.build_snapshots(engine=legacy.engine, verify_schema=False)


def _looks_like_year_workbook(path: Path) -> bool:
    """Workbook có phải bản lịch sử MỘT NĂM độc lập không.

    Chỉ đọc DANH SÁCH TÊN SHEET (``read_only``), không phân tích ô nào — nên
    phép thử này rẻ và không thể làm hỏng đường nhập hiện có: file nào không
    khớp hình dạng đó vẫn đi đúng nhánh cũ.
    """
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True)
    try:
        return is_standalone_year_workbook(book.sheetnames)
    finally:
        book.close()


def _workbook_year(workbook) -> str:
    for entry in workbook.sheets_imported:
        if entry.get("year"):
            return entry["year"]
    return "?"


def create_app(
    *,
    db_path: Path = run_registry.DEFAULT_DB_PATH,
    store: Optional[storage_backend.RunStore] = None,
    history: Optional[history_store.LegacyRepository] = None,
    snapshots: Optional[history_store.SnapshotRepository] = None,
) -> Flask:
    app = Flask(__name__)
    app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES
    if store is None:
        store = storage_backend.build(db_path=db_path, artifact_dir=ARTIFACT_DIR)
    app.config["RUN_REGISTRY"] = store
    # Tên cục bộ khác tên endpoint: một view function tên ``history`` sẽ che
    # mất biến này trong closure của create_app.
    history_repo = history if history is not None else _build_history()
    app.config["HISTORY_STORE"] = history_repo
    snapshot_repo = snapshots if snapshots is not None else _build_snapshots(history_repo)
    app.config["SNAPSHOT_STORE"] = snapshot_repo
    # PHB-03 — Summary/Employee nghiệp vụ. Dùng CHUNG engine với snapshot repo:
    # quyết định của Owner về một dòng hàng chỉ có nghĩa cạnh chính dòng đó,
    # nên chúng phải sống trong cùng một database, cùng một transaction domain.
    business = (
        None if snapshot_repo is None
        else business_service.BusinessReportService(
            engine=snapshot_repo.engine,
            store=business_store.BusinessDecisionStore(snapshot_repo.engine),
        )
    )
    app.config["BUSINESS_SERVICE"] = business
    app.jinja_env.globals["LEGACY_BADGE"] = legacy_presentation.ORIGIN_BADGE
    app.jinja_env.globals["LEGACY_BADGE_TITLE"] = legacy_presentation.ORIGIN_TITLE
    app.jinja_env.globals["LEGACY_PROVENANCE"] = legacy_reference.PROVENANCE
    app.jinja_env.globals["LEGACY_PROVENANCE_LABEL"] = legacy_reference.PROVENANCE_LABEL
    app.jinja_env.globals["LEGACY_PROVENANCE_NOTE"] = legacy_reference.PROVENANCE_NOTE
    app.jinja_env.globals["PIPELINE_BADGE"] = analytics_presentation.ORIGIN_BADGE
    app.jinja_env.globals["PIPELINE_BADGE_TITLE"] = analytics_presentation.ORIGIN_TITLE
    app.jinja_env.globals["QUANTITY_NOTE"] = analytics_presentation.QUANTITY_NOTE
    app.jinja_env.globals["ORDER_COLUMN_NOTE"] = analytics_presentation.ORDER_COLUMN_NOTE
    app.jinja_env.globals["BOTH_SOURCES_NOTE"] = analytics_presentation.BOTH_SOURCES_NOTE
    for _name in ("MULTI_DATE_NOTE", "MULTI_EMPLOYEE_NOTE", "NO_ORDERS_NOTE",
                  "PRODUCT_GROUPING_NOTE", "PRODUCT_ITEM_COUNT_LABEL",
                  "PRODUCT_ORDER_COUNT_NOTE", "REASON_LABEL"):
        app.jinja_env.globals[_name] = getattr(sales_presentation, _name)
    for _name in ("CONVERTED_SALES_NOTE", "DERIVED_COLUMNS_NOTE",
                  "DISCOUNT_ROW_NOTE",
                  "INCOMPLETE_NOTE", "NET_SALES_NOTE", "OFFICIAL_NOTE",
                  "QUALIFYING_QUANTITY_NOTE", "UNKNOWN_EMPLOYEE",
                  "UNRESOLVED_EMPLOYEE_NOTE"):
        app.jinja_env.globals[_name] = getattr(business_presentation, _name)
    app.jinja_env.globals["BUSINESS_ORDER_COLUMN_NOTE"] = \
        business_presentation.ORDER_COLUMN_NOTE

    def _require_history() -> history_store.LegacyRepository:
        if history_repo is None:
            # Không phải "chưa có dữ liệu" — là chưa có nơi lưu dữ liệu.
            abort(503)
        return history_repo

    def _legacy_page(template: str, **context):
        return render_template(
            template,
            history_configured=history_repo is not None,
            legacy_import=_guarded(history_repo.current_import) if history_repo else None,
            **context,
        )

    def _page(*, error: Optional[str] = None, run_id: Optional[str] = None,
             not_found: bool = False, feedback_ok: bool = False, status: int = 200):
        record = _guarded(store.get_run, run_id) if run_id else None
        return render_template(
            "index.html",
            readiness=_readiness_text(),
            error=error,
            result=record.view if record and record.view else None,
            run_id=run_id if record else None,
            run_created_at=record.created_at if record else None,
            workbook_display_name=record.workbook_display_name if record else None,
            not_found=not_found,
            feedback_ok=feedback_ok,
            history_configured=snapshot_repo is not None,
            feedback_categories=beta_feedback.FEEDBACK_CATEGORIES,
        ), status

    @app.get("/")
    def landing():
        """R1 (`GỠ TRÙNG UX`) — `/` mở BÁO CÁO, không còn màn hình upload.

        Chỉ đọc kỳ mới nhất từ Current Engine (`analytics_queries`, cùng
        nguồn `/kinh-doanh` đã dùng) — KHÔNG dò kỳ mới nhất bắc qua Legacy:
        hợp nhất "mới nhất" giữa hai nguồn là việc của R2, chưa làm ở đây.
        Không có snapshot store hoặc chưa có kỳ nào ⟹ về thẳng `/kinh-doanh`
        không kèm `ky`, để trang đó tự nói tình trạng của nó (503/rỗng).
        """
        ky = None
        if snapshot_repo is not None:
            periods = _guarded(analytics_queries.available_periods, snapshot_repo.engine)
            if periods:
                ky = business_presentation.period_value(periods[0])
        return redirect(url_for("business_summary", **({"ky": ky} if ky else {})))

    @app.get("/du-lieu/chay-bao-cao")
    def index():
        """Chạy báo cáo (upload workbook, xem kết quả) — dời từ `/` sang đây
        dưới DỮ LIỆU (R1 §3): đây là công cụ vận hành, không phải màn hình
        Owner/Director đọc báo cáo. Tên hàm ``index`` giữ nguyên để mọi
        ``url_for("index", ...)`` hiện có (trang Dữ liệu, POST /run, feedback)
        tự trỏ đúng đường mới, không phải sửa từng nơi gọi.
        """
        run_id = request.args.get("run_id") or None
        found = run_id is not None and _guarded(store.get_run, run_id) is not None
        return _page(
            run_id=run_id if found else None,
            not_found=run_id is not None and not found,
            feedback_ok=request.args.get("feedback") == "ok",
        )

    @app.get("/du-lieu")
    def data_tab():
        """Tab "Dữ liệu": các lần chạy pipeline VÀ các bản nhập legacy.

        Hai origin nằm trong hai bảng tách biệt trên trang, mỗi bảng ghi rõ
        nguồn — không bao giờ trộn số pipeline với số cũ vào một danh sách.
        """
        runs = _guarded(store.list_runs, limit=HISTORY_PAGE_LIMIT)
        imports = (
            _guarded(history_repo.list_imports, limit=LEGACY_IMPORT_PAGE_LIMIT)
            if history_repo else []
        )
        snapshots, runs_with_snapshot = [], set()
        if snapshot_repo is not None:
            snapshots = _guarded(snapshot_repo.list_snapshots, limit=SNAPSHOT_PAGE_LIMIT)
            # Một run có trên store mà KHÔNG có snapshot nghĩa là lần ghi lịch
            # sử đó đã hỏng. Trang phải nói ra, không im lặng bỏ qua.
            runs_with_snapshot = _guarded(
                snapshot_repo.run_ids_with_snapshot, [run.run_id for run in runs],
            )
        return _legacy_page(
            "du_lieu.html", runs=runs, imports=imports, snapshots=snapshots,
            snapshots_configured=snapshot_repo is not None,
            runs_with_snapshot=runs_with_snapshot,
            imported=request.args.get("imported") or None,
            error=request.args.get("loi") or None,
        )

    def _snapshot_page(snapshot_id: str, *, message=None, error=None, status=200):
        """Trang chỉ-đọc của MỘT snapshot: coverage, số đếm reconcile, cờ.

        Không hiển thị PII: bảng cờ chỉ mang khoá đơn/dòng, loại cờ và các
        trường nghiệp vụ đã đổi — tên/SĐT/địa chỉ khách không có mặt trong bất
        kỳ bảng nào của PRA-002 nên cũng không có đường nào ra tới đây.
        """
        if snapshot_repo is None:
            abort(503)
        snapshot = _guarded(snapshot_repo.get_snapshot, snapshot_id)
        if snapshot is None:
            abort(404)
        flags = _guarded(snapshot_repo.list_flags, snapshot_id=snapshot_id,
                         limit=FLAG_PAGE_LIMIT)
        totals = _guarded(
            snapshot_repo.current_totals,
            date_from=snapshot["detected_date_min"], date_to=snapshot["detected_date_max"],
        )
        return _legacy_page(
            "snapshot.html", snapshot=snapshot, flags=flags, totals=totals,
            coverage_label=history_coverage.coverage_label(snapshot["coverage_state"]),
            can_confirm=snapshot["coverage_state"] != history_models.CONFIRMED_COMPLETE,
            confirm_message=message, confirm_error=error,
        ), status

    @app.get("/du-lieu/snapshot/<snapshot_id>")
    def snapshot_detail(snapshot_id: str):
        return _snapshot_page(
            snapshot_id, message=request.args.get("xac_nhan") or None,
        )

    @app.post("/du-lieu/snapshot/<snapshot_id>/xac-nhan-du")
    def confirm_snapshot(snapshot_id: str):
        """Xác nhận TƯỜNG MINH rằng sổ này đầy đủ cho một khoảng ngày.

        Đây là hành động DUY NHẤT nâng coverage lên ``CONFIRMED_COMPLETE``, và
        nó luôn là một hành động riêng của con người sau khi đã nhìn thấy phạm
        vi hệ thống đo được (mục 7.3, DEC-171 #4). Không có suy diễn nào ở đây:
        không tick ô → 400; khoảng ngày không bao trọn dữ liệu → 400; xác nhận
        lần hai → 409. Mọi nhánh từ chối đều KHÔNG ghi gì.
        """
        if snapshot_repo is None:
            abort(503)
        try:
            confirmation = _guarded(
                snapshot_repo.confirm_coverage, snapshot_id,
                start=history_coverage.parse_iso_date(request.form.get("tu_ngay")),
                end=history_coverage.parse_iso_date(request.form.get("den_ngay")),
                confirmed=request.form.get("xac_nhan") == "1",
                confirmed_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
            )
        except KeyError:
            abort(404)
        except history_store.CoverageAlreadyConfirmedError as exc:
            return _snapshot_page(snapshot_id, error=str(exc), status=409)
        except history_store.CoverageRangeError as exc:
            return _snapshot_page(snapshot_id, error=str(exc), status=400)
        return redirect(url_for(
            "snapshot_detail", snapshot_id=snapshot_id,
            xac_nhan=(
                f"Đã ghi nhận xác nhận đầy đủ cho {confirmation.confirmed_range_start} "
                f"→ {confirmation.confirmed_range_end}. "
                f"{confirmation.removed_candidates} dòng hiện hành trong khoảng này không "
                "có trong sổ vừa xác nhận — đã đưa vào Review, KHÔNG xoá và VẪN tính."
            ),
        ))

    @app.get("/history")
    def history_redirect():
        # Đường cũ của S071 — giữ để link/bookmark đã phát ra không gãy.
        return redirect(url_for("data_tab"), code=302)

    @app.post("/du-lieu/legacy")
    def import_legacy():
        repository = _require_history()
        upload = request.files.get("workbook")
        if upload is None or not upload.filename:
            return redirect(url_for("data_tab", loi="Hãy chọn workbook legacy .xlsx."))
        if not upload.filename.lower().endswith(".xlsx"):
            return redirect(url_for("data_tab", loi="Chỉ chấp nhận file .xlsx."))

        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        # Tên file lưu trên đĩa LUÔN do server sinh — tên client gửi lên chỉ
        # dùng để hiển thị, nên không có chuỗi đường dẫn nào của client chạm
        # tới filesystem (chống path traversal).
        temp_path = UPLOAD_DIR / f"{uuid.uuid4().hex}.xlsx"
        upload.save(temp_path)
        try:
            # `DEC-178` — hai HÌNH DẠNG workbook, một ô tải lên. Nhận dạng
            # bằng CẤU TRÚC SHEET của chính file (có `Summary` + các sheet
            # `MM.YYYY <Nhãn>` ⟹ workbook một năm độc lập), không bằng tên
            # file: tên file do người dùng đặt và đổi được, cấu trúc thì không.
            workbook = (
                parse_year_workbook(temp_path)
                if _looks_like_year_workbook(temp_path)
                else parse_workbook(temp_path)
            )
            workbook = replace(
                workbook, source_file_name=_safe_display_name(upload.filename),
            )
            result = _guarded(
                repository.create_import, workbook,
                version_label=(request.form.get("version_label") or "").strip()[:120],
            )
        except LegacyImportError as exc:
            return redirect(url_for("data_tab", loi=str(exc)))
        except HTTPException:
            # `_guarded` biến lỗi history store thành abort(503) — mà abort
            # ném HTTPException. Không cho `except Exception` bên dưới nuốt
            # nó, nếu không sự cố DB sẽ hiện ra thành "không đọc được
            # workbook": đổ lỗi cho file của Owner vì một lỗi hạ tầng, và
            # phá đúng CHECK-PRA001-06 (FIND-PRA001-R02).
            raise
        except Exception:
            return redirect(url_for(
                "data_tab",
                loi="Không đọc được workbook legacy. Kiểm tra file và thử lại.",
            ))
        finally:
            # Workbook cũ chứa dữ liệu kinh doanh: không giữ lại trên đĩa
            # máy chủ quá một lần import, kể cả khi import lỗi.
            temp_path.unlink(missing_ok=True)
        if not result.created:
            message = (f"File này đã được nhập trước đó ({result.import_id})"
                       " — không tạo bản mới.")
        elif workbook.source_authority == SOURCE_AUTHORITY_YEAR:
            year = _workbook_year(workbook)
            message = (f"Đã nhập bản legacy {result.import_id} — nguồn CHUẨN "
                       f"của năm {year}.")
        else:
            message = f"Đã nhập bản legacy {result.import_id}."
        return redirect(url_for("data_tab", imported=message))

    @app.post("/du-lieu/legacy/<import_id>/chon")
    def choose_legacy(import_id: str):
        repository = _require_history()
        try:
            _guarded(repository.set_current, import_id)
        except KeyError:
            abort(404)
        return redirect(url_for("data_tab", imported=f"Đang xem bản {import_id}."))

    def _pipeline_period(periods: list[tuple[int, int]]) -> Optional[tuple[int, int]]:
        """Kỳ SỐ MỚI đang xem. Mặc định "Toàn bộ dữ liệu" (``None``).

        Mặc định đó là lựa chọn có chủ đích: nó là kỳ DUY NHẤT không bao giờ
        giấu bớt dòng nào và không cần một kỳ so sánh, nên trang mở ra lần đầu
        không thể nói sai. Một ``ky`` không có trong dữ liệu cũng rơi về đây —
        trang hiện tổng thật, thay vì một bảng toàn số 0 cho một tháng bịa.
        """
        raw = request.args.get("ky") or ""
        year_text, _, month_text = raw.partition("-")
        try:
            chosen = (int(year_text), int(month_text))
        except ValueError:
            return None
        return chosen if chosen in periods else None

    def _pipeline_view(engine) -> dict:
        """Kỳ + tổng kỳ + tổng kỳ trước, dùng chung cho Tổng quan và Nhân viên."""
        periods = _guarded(analytics_queries.available_periods, engine)
        period = _pipeline_period(periods)
        bounds = analytics_queries.month_bounds(*period) if period else (None, None)
        previous = None
        if period is not None:
            # Kỳ so sánh LUÔN được truy vấn khi đang xem một tháng: chính kết
            # quả "kỳ trước không có dòng nào" là thứ trang phải nói ra, nên
            # không được bỏ qua truy vấn đó.
            previous = _guarded(analytics_queries.period_totals, engine,
                                **dict(zip(("date_from", "date_to"),
                                           analytics_queries.month_bounds(
                                               *analytics_presentation.previous_period(period)))))
        return {
            "periods": analytics_presentation.period_options(periods),
            "period": period, "bounds": bounds,
            "selected_period": analytics_presentation.period_value(period),
            "totals": _guarded(analytics_queries.period_totals, engine,
                               date_from=bounds[0], date_to=bounds[1]),
            "previous": previous,
        }

    @app.get("/tong-quan")
    def overview():
        """Tổng quan SỐ MỚI — 10 ô đã qua Minimum-Value Filter, không hơn."""
        if snapshot_repo is None:
            abort(503)
        view = _pipeline_view(snapshot_repo.engine)
        return render_template(
            "tong_quan.html", periods=view["periods"],
            selected_period=view["selected_period"],
            overview=analytics_presentation.overview(
                view["totals"], view["previous"], period=view["period"],
                undated=_guarded(analytics_queries.undated_lines, snapshot_repo.engine),
            ),
        )

    @app.get("/ban-hang")
    def sales():
        """Danh sách đơn của kỳ — bậc đầu tiên của đường truy vết.

        Cùng bộ chọn kỳ, cùng ngữ nghĩa kỳ với Tổng quan: Owner chọn "Tháng
        09/2026" ở cả hai trang và phải thấy CÙNG một tập đơn. Không có kho dữ
        liệu ⟹ 503 giống hệt ``/tong-quan`` — lỗi database không bao giờ được
        hiện thành "chưa có dữ liệu".
        """
        if snapshot_repo is None:
            abort(503)
        view = _pipeline_view(snapshot_repo.engine)
        return render_template(
            "ban_hang.html", periods=view["periods"],
            selected_period=view["selected_period"],
            period_label=analytics_presentation.period_label(view["period"]),
            columns=sales_presentation.ORDER_COLUMNS,
            orders=sales_presentation.order_rows(
                _guarded(sales_queries.order_list, snapshot_repo.engine,
                         date_from=view["bounds"][0], date_to=view["bounds"][1])),
        )

    @app.get("/ban-hang/<order_key>")
    def sales_order(order_key: str):
        """Chi tiết MỘT đơn: khối tổng hợp, các dòng hiện hành, lý do kiểm tra.

        Mã đơn không có dòng hiện hành nào trong kỳ ⟹ 404. KHÔNG dựng một
        trang rỗng trông như "đơn này không có dòng nào": hai tình huống đó
        khác nhau, và chỉ một trong hai là sự thật.
        """
        if snapshot_repo is None:
            abort(503)
        view = _pipeline_view(snapshot_repo.engine)
        detail = _guarded(sales_queries.order_detail, snapshot_repo.engine, order_key,
                          date_from=view["bounds"][0], date_to=view["bounds"][1])
        if detail is None:
            abort(404)
        return render_template(
            "ban_hang_chi_tiet.html", periods=view["periods"],
            selected_period=view["selected_period"],
            period_label=analytics_presentation.period_label(view["period"]),
            line_columns=sales_presentation.LINE_COLUMNS,
            order=sales_presentation.order_detail(detail),
        )

    @app.get("/san-pham")
    def products():
        """SẢN PHẨM — mặt hàng trên chứng từ (TASK-PRA-005), mặc định sắp
        Doanh thu giảm dần. Gộp theo mô tả thô đã chuẩn hoá trên chứng từ
        (OD-PRA005-01, DEC-173) — KHÔNG phải canonical Product Identity.
        Bao gồm TẤT CẢ dòng chứng từ, kể cả dịch vụ/phí (OD-PRA005-02) —
        KHÔNG lọc bằng ``is_non_product_line()``. Không có kho dữ liệu ⟹ 503
        giống hệt ``/tong-quan``/``/ban-hang``.
        """
        if snapshot_repo is None:
            abort(503)
        view = _pipeline_view(snapshot_repo.engine)
        rows = _guarded(sales_queries.product_totals, snapshot_repo.engine,
                        date_from=view["bounds"][0], date_to=view["bounds"][1])
        return render_template(
            "san_pham.html", periods=view["periods"],
            selected_period=view["selected_period"],
            period_label=analytics_presentation.period_label(view["period"]),
            columns=sales_presentation.PRODUCT_COLUMNS,
            summary=sales_presentation.product_summary(rows, view["totals"]),
            products=sales_presentation.product_rows(rows),
        )

    @app.get("/nhan-vien")
    def sellers():
        """Ma trận tháng × người bán từ Summary cũ (đơn vị nghìn đồng).

        ``?nguon=moi`` chuyển sang bảng SỐ MỚI. MỌI giá trị khác — kể cả không
        có tham số và các giá trị lạ — giữ NGUYÊN VẸN đường legacy: bảo toàn
        bằng chứng non-regression của TASK-PRA-001 quan trọng hơn sự đối xứng
        của route, và một tham số gõ sai không được phép thành HTTP 500.
        """
        if request.args.get("nguon") == "moi" and snapshot_repo is not None:
            view = _pipeline_view(snapshot_repo.engine)
            return render_template(
                "nhan_vien.html", pipeline_view=True,
                periods=view["periods"], selected_period=view["selected_period"],
                columns=analytics_presentation.EMPLOYEE_COLUMNS,
                period_label=analytics_presentation.period_label(view["period"]),
                rows=analytics_presentation.employee_rows(
                    _guarded(analytics_queries.employee_totals, snapshot_repo.engine,
                             date_from=view["bounds"][0], date_to=view["bounds"][1]),
                    view["totals"],
                ),
            )
        if history_repo is None:
            return _legacy_page("nhan_vien.html", periods=[], selected=None,
                                rows=[], columns=legacy_presentation.MATRIX_COLUMNS,
                                pipeline_view=False), 503
        periods = _guarded(history_repo.available_periods)
        selected = _selected_period(periods)
        rows = (
            _guarded(history_repo.query_summary, selected[0], selected[1])
            if selected else []
        )
        return _legacy_page(
            "nhan_vien.html", periods=periods, selected=selected,
            rows=legacy_presentation.matrix(rows),
            columns=legacy_presentation.MATRIX_COLUMNS,
            pipeline_view=False,
        )

    @app.get("/doanh-so-ngay")
    def daily_sales():
        """Doanh số theo ngày từ DataChart cũ (đơn vị VND nguyên)."""
        if history_repo is None:
            return _legacy_page("doanh_so_ngay.html", periods=[], selected=None,
                                days=[], monthly=None, monthly_cells={}), 503
        periods = _guarded(history_repo.available_periods)
        selected = _selected_period(periods)
        days, monthly = [], None
        if selected and selected[1]:
            days = legacy_presentation.daily_grid(
                _guarded(history_repo.query_daily, selected[0], selected[1])
            )
            monthly = next(
                (row for row in _guarded(history_repo.query_monthly_reference, selected[0])
                 if row["month"] == selected[1]),
                None,
            )
        return _legacy_page(
            "doanh_so_ngay.html", periods=periods, selected=selected, days=days,
            monthly=monthly, monthly_cells=legacy_presentation.monthly_cells(monthly),
        )

    @app.get("/lich-su")
    def legacy_reference_page():
        """PHB-04 — Tham chiếu lịch sử: kỳ legacy, hợp đồng, điều hướng.

        Trang này KHÔNG ghi gì và KHÔNG gọi pipeline. Nó đọc đúng ba thứ đã
        có: kỳ Summary cũ, dòng tham chiếu tháng của DataChart, và danh mục kỳ
        của số mới — rồi xếp cạnh nhau với nhãn origin. Không kỳ nào bị hợp
        nhất thành một con số duy nhất: `DEC-166 E` cấm cộng chung số cũ với
        số mới, và PHB-04 không xin ngoại lệ nào cho điều đó.

        Danh mục kỳ số mới chỉ có khi snapshot store đã cấu hình. Thiếu nó,
        trang vẫn hiện đầy đủ phần legacy và nói rõ phần số mới chưa đọc được
        — KHÔNG im lặng hiện một danh sách chỉ có legacy như thể đó là tất cả.
        """
        if history_repo is None:
            return _legacy_page(
                "lich_su.html", reference_rows=[], reference_years=[],
                reference_has_value=False, navigation=[], pipeline_configured=False,
                summary_periods=[], summary_years=[], unread=[],
                comparison=legacy_reference.comparison_summary(),
                reference_contract=legacy_presentation.contract_rows(
                    legacy_reference.REFERENCE_YEAR_CONTRACT),
                workbook_contract=legacy_presentation.contract_rows(
                    legacy_reference.SUMMARY_SHEET_CONTRACT),
            ), 503

        monthly = _guarded(history_repo.query_monthly_reference)
        periods = legacy_reference.reference_periods(monthly)
        summary_periods = _guarded(history_repo.available_periods)
        pipeline_periods = (
            _guarded(analytics_queries.available_periods, snapshot_repo.engine)
            if snapshot_repo is not None else []
        )
        # Mọi dòng Summary đã nhập, MỌI năm. `DEC-177`: chi tiết theo nhân
        # viên của một năm lịch sử sống ở đây, nên trang phải đọc cả năm
        # lịch sử chứ không chỉ năm workbook.
        all_summary = _guarded(history_repo.query_all_summary)
        years = legacy_reference.summary_years(all_summary)
        current = _guarded(history_repo.current_import)
        # `DEC-178` — mỗi năm ghi rõ nó đang đọc từ nguồn nào. Đây là chỗ
        # quy tắc "nguồn chuẩn thắng" trở thành thứ chủ dự án NHÌN THẤY,
        # không chỉ là một dòng trong tài liệu.
        authority_by_year = {
            year.year: _guarded(history_repo.authority_import_for_year, year.year)
            for year in years
        }
        return _legacy_page(
            "lich_su.html",
            reference_rows=legacy_presentation.reference_rows(periods),
            reference_years=legacy_reference.reference_years(periods),
            reference_has_value=legacy_reference.has_any_value(periods),
            summary_periods=summary_periods,
            summary_years=legacy_presentation.summary_year_rows(
                years, all_summary, authority_by_year, current),
            unread=legacy_reference.unread_sheets(
                (current or {}).get("sheets_imported")),
            navigation=legacy_reference.period_navigation(
                legacy_summary_periods=summary_periods,
                legacy_reference_periods=periods,
                pipeline_periods=pipeline_periods,
            ),
            pipeline_configured=snapshot_repo is not None,
            comparison=legacy_reference.comparison_summary(),
            reference_contract=legacy_presentation.contract_rows(
                legacy_reference.REFERENCE_YEAR_CONTRACT),
            workbook_contract=legacy_presentation.contract_rows(
                legacy_reference.SUMMARY_SHEET_CONTRACT),
        )

    @app.get("/giai-thich")
    def giai_thich():
        """R1 §7 — một nơi DUY NHẤT cho định nghĩa KPI lặp lại trên nhiều
        trang báo cáo (Tổng số SP, DS quy đổi, tiền hiển thị nghìn đồng,
        huy hiệu SỐ MỚI). Trang này KHÔNG đọc dữ liệu, không có kỳ, không
        có nhân viên — chỉ prose tĩnh, nên không cần snapshot/history store.
        """
        return render_template("giai_thich.html")

    # ------------------------------------------------------------------
    # PHB-03 — Summary + Employee Business Parity V1.
    #
    # Bốn trang, MỘT vertical: Tổng hợp (kỳ) · Nhân viên (nhân viên + kỳ) ·
    # Hoàn thiện giá nhập · Phân loại Gia dụng. Cố ý KHÔNG dựng một tab cho mỗi
    # nhân viên (`R-E1`, `P1`) — 56 sheet tay trở thành MỘT trang có bộ chọn.
    # ------------------------------------------------------------------

    def _require_business():
        if business is None:
            # Không phải "chưa có dữ liệu" — là chưa có nơi lưu dữ liệu.
            abort(503)
        return business

    def _business_period_choice(periods: list[tuple[int, int]]):
        """Kỳ đang xem, đọc từ `request.values` chứ không riêng query string.

        Các trang này có FORM POST (ghi giá nhập, tick Gia dụng) mang `ky`
        trong BODY. Đọc riêng query string ở đó sẽ âm thầm rơi về "Toàn bộ dữ
        liệu" và chuyển hướng người dùng sang một kỳ họ không chọn. Ngoài
        nguồn đọc, ngữ nghĩa giữ đúng `_pipeline_period`: giá trị lạ hoặc kỳ
        không có trong dữ liệu đều rơi về `None` ("Toàn bộ dữ liệu"), vì đó là
        kỳ DUY NHẤT không giấu bớt dòng nào.
        """
        raw = request.values.get("ky") or ""
        year_text, _, month_text = raw.partition("-")
        try:
            chosen = (int(year_text), int(month_text))
        except ValueError:
            return None
        return chosen if chosen in periods else None

    def _legacy_previous_month(period, previous) -> Optional[dict]:
        """Mốc "So tháng trước" lấy từ SỐ CŨ — `DEC-180` §9.

        Đây là chỗ DUY NHẤT của vertical nghiệp vụ chạm tới hai origin, và nó
        cố ý nằm ở tầng RÁP chứ không trong `business_*`: ba module kia phải
        tiếp tục không biết gì về số cũ, nếu không một ngày nào đó cổng so
        sánh liên-origin sẽ chặn nhầm một phép so cùng-engine.

        Bốn cửa, mỗi cửa đóng một cách sai khác nhau:

        1. Chỉ chạy khi đang xem MỘT tháng. "Toàn bộ dữ liệu" không có tháng
           liền trước nào để so.
        2. Chỉ chạy khi tháng trước KHÔNG có dòng số mới nào. Một tháng đã có
           số mới không bao giờ bị số cũ thay chỗ.
        3. Đi qua `authoritative_period_sales`: MỘT kỳ ⟹ MỘT nguồn ⟹ MỘT giá
           trị. Không cộng hai nguồn, không trộn dòng thô.
        4. Giá trị trả về đã chuẩn hoá về VND bằng `to_vnd()` trong chính hàm
           đó — `Summary` là kVND, số mới là VND, và quên hệ số 1.000 ở đây
           cho ra một tỉ lệ trông như thật.

        Trang NHÂN VIÊN cố ý KHÔNG dùng đường này: số cũ của một tháng là
        tổng của CẢ CÔNG TY, nên đem nó làm mẫu số cho doanh thu của một người
        là một phép so sai. Ghép tên người bán trong sổ cũ với nhân viên hiện
        hành là một bài toán ánh xạ riêng, chưa có quyết định nào cho phép.
        """
        if period is None or history_repo is None:
            return None
        if previous is not None and previous.totals.lines > 0:
            return None
        year, month = analytics_presentation.previous_period(period)
        summary_rows = _guarded(history_repo.query_summary, year, month)
        monthly_rows = _guarded(history_repo.query_monthly_reference, year)
        resolved = legacy_reference.authoritative_period_sales(
            year=year, month=month,
            summary_rows=summary_rows, monthly_rows=monthly_rows)
        if resolved is None:
            return None
        return {
            "sales_revenue": resolved.sales_vnd,
            "origin_label": resolved.origin_label,
            "source_label": resolved.source_label,
        }

    def _business_period() -> dict:
        """Kỳ đang xem + số của kỳ đó + số của kỳ liền trước.

        Kỳ so sánh LUÔN được truy vấn khi đang xem một tháng: chính kết quả
        "tháng trước không có dòng nào" là thứ `DEC-PHB02-07` bắt phải nói ra,
        nên không được bỏ qua truy vấn đó.
        """
        service = _require_business()
        periods = _guarded(analytics_queries.available_periods, snapshot_repo.engine)
        period = _business_period_choice(periods)
        bounds = analytics_queries.month_bounds(*period) if period else (None, None)
        data = _guarded(service.period, date_from=bounds[0], date_to=bounds[1])
        previous = None
        if period is not None:
            previous_bounds = analytics_queries.month_bounds(
                *analytics_presentation.previous_period(period))
            previous = _guarded(service.period, date_from=previous_bounds[0],
                                date_to=previous_bounds[1])
        return {
            "service": service, "period": period, "bounds": bounds, "data": data,
            "previous": previous,
            "previous_fallback": _legacy_previous_month(period, previous),
            "periods": business_presentation.period_options(periods),
            "selected_period": business_presentation.period_value(period),
        }

    def _period_employees(view: dict):
        """Bộ chọn nhân viên của kỳ, ĐÃ tính cả những lần Owner gán lại.

        Truyền `data` vào là điều kiện để `OD-5` khép kín: ngay sau khi Owner
        gán một dòng cho "Vinh", tên đó phải chọn được — nếu bộ chọn vẫn là
        danh sách thô của pipeline thì trang nhân viên trả 404 đúng lúc thao
        tác vừa thành công.
        """
        return _guarded(view["service"].employees,
                        date_from=view["bounds"][0], date_to=view["bounds"][1],
                        data=view["data"])

    def _selected_employee(view: dict) -> tuple[Optional[str], Optional[str], bool]:
        """`(tên, nhóm, đã chọn hợp lệ)` từ tham số `nhan-vien`.

        Một tên không có trong kỳ KHÔNG được dựng thành một trang toàn số 0 —
        đó là một nhân viên bịa. Trang rơi về "chưa chọn ai" và nói rõ.
        """
        employees = _period_employees(view)
        # `request.values`, không phải `request.args`: form POST của trang tick
        # Gia dụng mang `nhan-vien` trong BODY, và đọc thiếu nó ở đó sẽ biến
        # một thao tác hợp lệ thành 404.
        raw = request.values.get("nhan-vien")
        if raw is None:
            return None, None, False
        for name, group in employees:
            if (name or "") == raw:
                return name, group, True
        return None, None, False

    def _employee_context(view: dict) -> dict:
        employees = _period_employees(view)
        name, group, chosen = _selected_employee(view)
        return {
            "employees": business_presentation.employee_options(employees),
            "selected_employee": request.values.get("nhan-vien") or "",
            "employee": name, "employee_group": group, "chosen": chosen,
        }

    @app.get("/kinh-doanh")
    def business_summary():
        """SUMMARY V1 — `R-S1`…`R-S8`. Một kỳ, sáu chỉ tiêu đã freeze.

        `R1` cộng thêm ĐÚNG một cảnh báo: sổ nạp gần nhất không thấy lại một
        số dòng đang được tính vào các con số trên trang này. Cảnh báo đó
        KHÔNG sửa, không gộp và không loại bất kỳ dòng nào — nó chỉ dẫn Owner
        sang trang snapshot để tự soi.
        """
        view = _business_period()
        totals = view["data"].totals
        absence = (None if snapshot_repo is None
                   else _guarded(snapshot_repo.latest_snapshot_absence))
        return render_template(
            "kinh_doanh.html", periods=view["periods"],
            selected_period=view["selected_period"],
            not_seen=business_presentation.not_seen_warning(absence),
            summary=business_presentation.summary(
                totals, period=view["period"],
                previous_totals=None if view["previous"] is None
                                else view["previous"].totals,
                undated=_guarded(view["service"].undated_lines),
                previous_fallback=view["previous_fallback"],
            ),
            columns=business_presentation.EMPLOYEE_COLUMNS,
            rows=business_presentation.employee_rows(
                business_metrics.group_by_employee(view["data"].lines), totals),
        )

    @app.get("/kinh-doanh/nhan-vien")
    def business_employee():
        """EMPLOYEE V1 — `R-E1`…`R-E8`. Owner chọn NHÂN VIÊN + KỲ."""
        view = _business_period()
        context = _employee_context(view)
        detail = None
        if context["chosen"]:
            scoped = view["data"].for_employee(context["employee"])
            previous = (None if view["previous"] is None
                        else view["previous"].for_employee(context["employee"]))
            detail = business_presentation.employee_detail(
                context["employee"], context["employee_group"], scoped.totals,
                period=view["period"],
                previous_totals=None if previous is None else previous.totals,
                gia_dung=gia_dung_workflow_applies(context["employee_group"]),
            )
        return render_template(
            "kinh_doanh_nhan_vien.html", periods=view["periods"],
            selected_period=view["selected_period"], detail=detail, **context)

    # Bốn chế độ lọc của bảng kê. Chúng KHÔNG chồng lên nhau và mỗi cái trả
    # lời một câu hỏi khác của Owner — gộp lại thành một danh sách "còn thiếu"
    # chung chính là cái đã khiến `B03` xảy ra.
    #
    # `S120`: MẶC ĐỊNH là `tat-ca`, không còn là `thieu-gia`. Bảng kê chi tiết
    # là KHUNG NHÌN BÁO CÁO của một nhân viên/kỳ, không phải hàng đợi việc
    # tồn: mặc định lọc bỏ mọi dòng đã đủ giá khiến Owner mở trang ra và thấy
    # một tập con mà trang không nói là tập con — và khi coverage đã 100 %,
    # cùng đường dẫn đó cho ra một bảng RỖNG. Ba bộ lọc thu hẹp vẫn còn
    # nguyên, chỉ khác là Owner phải chọn chúng một cách tường minh.
    _DETAIL_FILTERS = {
        "tat-ca": lambda line: True,
        # Việc Owner gõ được ngay bây giờ.
        "thieu-gia": lambda line: line.purchase_price is None,
        # Dòng đã có lãi nhưng chưa biết của ai (`OD-5`).
        "chua-ro-nv": lambda line: (line.contributes_profit
                                    and not line.employee_resolved),
        # `R3` — "dòng tôi đã sửa": CHỈ đọc lại provenance đã lưu (giá nhập
        # Owner nhập/sửa, hoặc nhân viên Owner gán lại). Không trạng thái mới,
        # không workflow mới, không ghi gì.
        "owner-sua": lambda line: (
            line.purchase_provenance in _OWNER_EDITED_PROVENANCE
            or line.employee_provenance == "MANUAL"),
    }
    _DEFAULT_DETAIL_FILTER = "tat-ca"

    @app.get("/kinh-doanh/gia-nhap")
    def business_purchase_price():
        """BẢNG KÊ CHI TIẾT — một dòng hàng một dòng bảng, sửa ngay tại chỗ.

        Trang này vừa là nơi hoàn thiện giá nhập (`R-P1`…`R-P4`) vừa là "trang
        tính" mà chỉ thị `ORDER DETAIL TABLE` mô tả: giá nhập và nhân viên sửa
        được ngay trên dòng, còn doanh thu · lợi nhuận KPI · DS quy đổi là ba
        ô SUY RA tự tính lại sau mỗi lần lưu. Không có bước "tính" riêng.

        Bộ lọc mặc định là TẤT CẢ DÒNG của kỳ/nhân viên đang xem (`S120`):
        một khung nhìn báo cáo không được âm thầm giấu bớt dòng. Ba chế độ thu
        hẹp — còn thiếu giá, chưa rõ nhân viên, dòng Owner đã sửa — nằm ngay
        trên bảng và Owner chọn tường minh.
        """
        view = _business_period()
        context = _employee_context(view)
        data = (view["data"].for_employee(context["employee"])
                if context["chosen"] else view["data"])
        # `tat-ca=1` là cách viết cũ còn nằm trong bookmark/redirect — giữ nó
        # tương đương chế độ "tất cả" thay vì âm thầm rơi về danh sách khác.
        mode = request.args.get("loc") or _DEFAULT_DETAIL_FILTER
        if mode not in _DETAIL_FILTERS:
            mode = _DEFAULT_DETAIL_FILTER
        keep = _DETAIL_FILTERS[mode]
        details = [d for d in data.details if keep(d["line"])]
        return render_template(
            "kinh_doanh_gia_nhap.html", periods=view["periods"],
            selected_period=view["selected_period"],
            period_label=business_presentation.period_label(view["period"]),
            columns=business_presentation.DETAIL_COLUMNS,
            rows=business_presentation.detail_rows(details),
            coverage=business_presentation.coverage_cell(data.totals.coverage),
            assignable=business_presentation.assignable_employee_options(
                view["service"].assignable_employees()),
            mode=mode, show_all=(mode == "tat-ca"),
            message=request.args.get("da-luu") or None,
            error=request.args.get("loi") or None, **context)

    @app.post("/kinh-doanh/gia-nhap")
    def business_save_purchase_price():
        """Ghi MỘT giá nhập. Provenance do server quyết, không do form khai."""
        view = _business_period()
        service = view["service"]
        try:
            occurrence = int(request.form.get("occurrence_index") or "")
        except ValueError:
            abort(400)
        keys = {
            "order_key": request.form.get("order_key") or "",
            "product_key": request.form.get("product_key") or "",
            "occurrence_index": occurrence,
        }
        exists, auto_price = service.auto_price_of(data=view["data"], **keys)
        if not exists:
            abort(404)
        redirect_args = {
            "ky": request.values.get("ky") or "tat-ca",
            "nhan-vien": request.form.get("nhan-vien") or None,
            "loc": request.form.get("loc") or None,
        }
        if request.form.get("hanh-dong") == "go":
            _guarded(service.store.clear_purchase_price, **keys)
            return redirect(url_for(
                "business_purchase_price",
                **{k: v for k, v in redirect_args.items() if v},
                **{"da-luu": "Đã gỡ giá nhập do Owner nhập. Dòng trở lại giá "
                             "tự động của hệ thống."}))
        try:
            price = business_store.parse_purchase_price(request.form.get("gia_nhap"))
        except business_store.InvalidPurchasePriceError as exc:
            return redirect(url_for(
                "business_purchase_price",
                **{k: v for k, v in redirect_args.items() if v}, loi=str(exc)))
        provenance = _guarded(service.store.set_purchase_price,
                              price=price, auto_price=auto_price, **keys)
        return redirect(url_for(
            "business_purchase_price",
            **{k: v for k, v in redirect_args.items() if v},
            **{"da-luu": (
                "Đã ghi giá nhập Owner sửa (thay giá tự động)."
                if provenance == "MANUAL_OVERRIDE"
                else "Đã ghi giá nhập Owner nhập.")}))

    @app.post("/kinh-doanh/nhan-vien-dong")
    def business_save_employee():
        """`OD-5` — gán MỘT dòng hàng cho một nhân viên, hoặc gỡ việc gán đó.

        Ranh giới cố ý hẹp: đây KHÔNG phải một trình sửa đơn hàng. Nó ghi đúng
        một trường, trên đúng một dòng đã tồn tại trong kỳ đang xem, và tên
        nhân viên phải nằm trong master `config/employees.yaml` — gõ tự do một
        cái tên vào KPI là mở lại đúng lớp lỗi mà `HD-110-06` đã đóng.

        Bằng chứng gốc không bị đụng: `order_line_result_version` vẫn giữ
        nguyên tên mà sổ ghi, và bảng override lưu lại tên đó ở cột
        `source_employee_at_entry`.
        """
        view = _business_period()
        service = view["service"]
        try:
            occurrence = int(request.form.get("occurrence_index") or "")
        except ValueError:
            abort(400)
        keys = {
            "order_key": request.form.get("order_key") or "",
            "product_key": request.form.get("product_key") or "",
            "occurrence_index": occurrence,
        }
        detail = service.detail_of(data=view["data"], **keys)
        if detail is None:
            abort(404)
        redirect_args = {
            "ky": request.values.get("ky") or "tat-ca",
            "nhan-vien": request.form.get("nhan-vien") or None,
            "loc": request.form.get("loc") or None,
        }

        def _back(**extra):
            return redirect(url_for(
                "business_purchase_price",
                **{k: v for k, v in redirect_args.items() if v}, **extra))

        if request.form.get("hanh-dong") == "go":
            _guarded(service.store.clear_employee, **keys)
            return _back(**{"da-luu": (
                "Đã gỡ việc gán nhân viên. Dòng trở lại đúng tên mà sổ ghi.")})

        chosen = (request.form.get("nhan_vien_moi") or "").strip()
        groups = dict(service.assignable_employees())
        if chosen not in groups:
            # Không đoán hộ, và không im lặng: một tên lạ nghĩa là master chưa
            # có người đó, và đó là việc sửa master chứ không phải việc của
            # trang này.
            return _back(loi=(
                f"{chosen!r} không có trong danh sách nhân viên. Hãy chọn một "
                "tên trong danh sách."))
        _guarded(service.store.set_employee,
                 employee=chosen, employee_group=groups[chosen],
                 source_employee=detail["line"].source_employee, **keys)
        return _back(**{"da-luu": (
            f"Đã gán dòng này cho {chosen}. Lợi nhuận của dòng đã chuyển sang "
            f"bảng của {chosen}; tổng của cả kỳ không đổi.")})

    @app.get("/kinh-doanh/gia-dung")
    def business_gia_dung():
        """Tick Gia dụng — CHỈ nhóm Nội thành (`DEC-PHB02-05`).

        Nhân viên bán lẻ thường không được thấy luồng này: chỉ thị PHB-03 §2E
        nói rõ "Do not expose the Gia dụng workflow to ordinary retail
        employees". Chọn sai nhân viên ⟹ 404, không phải một trang tick không
        có tác dụng gì.
        """
        view = _business_period()
        context = _employee_context(view)
        if not context["chosen"]:
            abort(404)
        if not gia_dung_workflow_applies(context["employee_group"]):
            abort(404)
        data = view["data"].for_employee(context["employee"])
        return render_template(
            "kinh_doanh_gia_dung.html", periods=view["periods"],
            selected_period=view["selected_period"],
            period_label=business_presentation.period_label(view["period"]),
            columns=business_presentation.GIA_DUNG_COLUMNS,
            rows=business_presentation.gia_dung_rows(
                view["service"].products(data)),
            message=request.args.get("da-luu") or None, **context)

    @app.post("/kinh-doanh/gia-dung")
    def business_save_gia_dung():
        """Ghi/gỡ phân loại Gia dụng của MỘT mặt hàng.

        Quyền tick được kiểm lại ở đây, không chỉ ở trang GET: một POST dựng
        tay vẫn phải đi qua đúng ranh giới `DEC-PHB02-05`.
        """
        view = _business_period()
        context = _employee_context(view)
        if not context["chosen"] or not gia_dung_workflow_applies(
                context["employee_group"]):
            abort(404)
        product_key = request.form.get("product_key") or ""
        products = {item["product_key"]: item for item in view["service"].products(
            view["data"].for_employee(context["employee"]))}
        product = products.get(product_key)
        if product is None:
            abort(404)
        if request.form.get("gia_dung") == "1":
            _guarded(view["service"].store.set_product_group,
                     product_key=product_key, product_group="GIA_DUNG",
                     product_label=product["product_label"])
            saved = "Đã đánh dấu mặt hàng này là Gia dụng (tỉ lệ quy đổi 8%)."
        else:
            _guarded(view["service"].store.clear_product_group,
                     product_key=product_key)
            saved = "Đã bỏ đánh dấu Gia dụng. Mặt hàng trở lại tỉ lệ mặc định."
        return redirect(url_for(
            "business_gia_dung", ky=request.values.get("ky") or "tat-ca",
            **{"nhan-vien": context["selected_employee"], "da-luu": saved}))

    @app.post("/run")
    def run_report():
        upload = request.files.get("workbook")
        if upload is None or not upload.filename:
            return _page(error="Hãy chọn một workbook .xlsx trước khi chạy.", status=400)
        if not upload.filename.lower().endswith(".xlsx"):
            return _page(error="Chỉ chấp nhận file .xlsx.", status=400)
        display_name = _safe_display_name(upload.filename)

        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        temp_path = UPLOAD_DIR / f"{uuid.uuid4().hex}.xlsx"
        upload.save(temp_path)
        started = time.monotonic()
        live_handle = None
        try:
            try:
                captures, tracking_evidence, live_handle = _select_captures_for_run()
            except live_pull.TrackingUnavailableError as exc:
                return _page(
                    error=(
                        "Không lấy được dữ liệu Tracking trực tiếp (nguồn: "
                        f"{exc.node}). Đây KHÔNG phải lỗi của workbook — vui "
                        "lòng thử lại sau."
                    ),
                    status=503,
                )
            try:
                owner_run = run_owner_report(sales=temp_path, captures=captures)
            except OwnerUsabilityError as exc:
                return _page(error=str(exc), status=400)
            except Exception:
                return _page(
                    error="Không thể tạo báo cáo. Kiểm tra workbook và thử lại.",
                    status=400,
                )

            duration_ms = int((time.monotonic() - started) * 1000)
            summary = owner_run.demo_run.summary
            dropped_lines = len(owner_run.demo_run.result.unmapped_lines)
            run_id = owner_run.output_path.stem

            def _persist_run() -> None:
                # save_artifact (R2: upload + verify + xoá temp; local: chỉ
                # tính path tương đối, file đã nằm sẵn dưới ARTIFACT_DIR) PHẢI
                # thành công trước khi ghi run — không được để lộ một run
                # "thành công" mà artifact không thực sự tồn tại ở nơi lưu.
                artifact_ref = store.save_artifact(owner_run.output_path, run_id)
                store.create_run(
                    run_id=run_id,
                    created_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
                    status=run_registry.STATUS_COMPLETE,
                    workbook_display_name=display_name,
                    artifact_path=artifact_ref,
                    view=_build_view(summary, dropped_lines=dropped_lines),
                    tracking_evidence=tracking_evidence,
                )

            try:
                if snapshot_repo is None:
                    # Dev chưa cấu hình history store: vẫn chạy như S071B,
                    # nhưng trang kết quả nói thẳng là run này KHÔNG có lịch sử
                    # — không bao giờ để người dùng tin là đã lưu.
                    _persist_run()
                else:
                    # MỘT đơn vị công việc: lịch sử + artifact + run cùng cam
                    # kết hoặc cùng rollback (TASK-PRA-002 mục 11.2).
                    history_writer.write_run_history(
                        snapshot_repo, demo_run=owner_run.demo_run, run_id=run_id,
                        workbook_path=temp_path, display_name=display_name,
                        tracking_evidence=tracking_evidence, on_persisted=_persist_run,
                    )
            except Exception:
                # Report ĐÃ được tạo trên đĩa tạm, nhưng không lưu được vào
                # nơi lưu trữ (artifact upload lỗi, ghi metadata lỗi, hoặc ghi
                # lịch sử lỗi) — không được trả về như thể mọi thứ thành công
                # (không có run_id để Owner tra lại). Fail rõ, không giả.
                return _page(
                    error=(
                        "Báo cáo đã tạo nhưng không lưu được vào lịch sử run. "
                        "Vui lòng thử lại."
                    ),
                    status=500,
                )
        finally:
            # Workbook chỉ được xoá SAU khi history writer đã đọc xong header
            # và fingerprint của chính bytes đã upload.
            temp_path.unlink(missing_ok=True)
            if live_handle is not None:
                live_handle.cleanup()

        _record_telemetry(run_id, summary, duration_ms)
        # Post-Redirect-Get: tránh chạy lại báo cáo khi Owner bấm refresh.
        return redirect(url_for("index", run_id=run_id))

    @app.get("/artifact/<run_id>")
    def download_artifact(run_id: str):
        record = _guarded(store.get_run, run_id)
        if record is None or not record.artifact_path:
            abort(404)
        # store.artifact_response() tự resolve artifact_path CHỈ qua đúng
        # record authoritative này — browser không bao giờ cung cấp key/path
        # trực tiếp (local: chặn path traversal dưới ARTIFACT_DIR; R2: key
        # luôn tự suy từ run_id, xem app/web/storage_backend.py).
        response = _guarded(store.artifact_response, record)
        if response is None:
            abort(404)
        return response

    @app.post("/feedback")
    def submit_feedback():
        category = request.form.get("category", "")
        comment = request.form.get("comment", "")
        run_id = request.form.get("run_id") or None
        try:
            record = beta_feedback.build_feedback_record(
                category=category, comment=comment, run_id=run_id,
            )
        except beta_feedback.InvalidFeedbackError:
            return _page(error="Loại phản hồi không hợp lệ.", run_id=run_id, status=400)
        beta_feedback.save_feedback(record)
        return redirect(url_for("index", run_id=run_id, feedback="ok"))

    @app.errorhandler(RequestEntityTooLarge)
    def _too_large(_exc):
        return _page(
            error="Workbook vượt quá giới hạn 25MB cho bản Beta. Hãy chọn file nhỏ hơn.",
            status=413,
        )

    @app.errorhandler(404)
    def _not_found(_exc):
        return _page(error="Không tìm thấy tài nguyên yêu cầu.", status=404)

    @app.errorhandler(500)
    def _server_error(_exc):
        return _page(error="Có lỗi xử lý phía máy chủ. Vui lòng thử lại.", status=500)

    @app.errorhandler(503)
    def _storage_unavailable(_exc):
        return _page(
            error="Lưu trữ tạm thời không khả dụng. Vui lòng thử lại.", status=503,
        )

    return app
