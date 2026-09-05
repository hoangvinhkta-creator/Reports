"""Đọc workbook "Báo cáo Kinh doanh" cũ thành bản ghi bất biến.

Module thuần openpyxl: KHÔNG biết database tồn tại, KHÔNG gọi mạng, KHÔNG
tính lại bất kỳ con số nào. Giá trị ô đi thẳng vào bản ghi; công thức được
giữ nguyên văn ở ``formula_text`` để audit; chỗ công thức tự mâu thuẫn được
gắn mã ở ``known_defects`` (xem ``app/legacy/defects.py``) mà KHÔNG đụng tới
giá trị.

Phân loại dòng của Summary hoàn toàn theo cấu trúc công thức quan sát được
(``docs/analysis/02_FORMULA_MAPPING.md`` §3) chứ không theo vị trí dòng cố
định — số dòng người bán mỗi tháng thay đổi (tháng 03.2026 có thêm "Linh"),
nên mọi offset cứng đều sai ở kỳ nào đó.
"""

from __future__ import annotations

import hashlib
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app.legacy import defects as defect_rules
from app.legacy.models import (
    SUMMARY_COLUMN_FIELDS, DailySales, LegacyWorkbook, MonthlyReference,
    SummaryRow,
)

# Phạm vi import của PRA-001 + PHB-04 — thẩm quyền Owner (DEC-169, DEC-177).
#
# `Summary 2026` và `DataChart 2026` là REQUIRED_IMPORT: nguồn dữ liệu
# production, phải nhập đủ và đúng. Guard DEC-168 áp dụng NGUYÊN VẸN ở đây —
# một dòng có giá trị nghiệp vụ mà contract không phân loại được vẫn FAIL TO.
#
# `Summary 2025` là OPTIONAL_IMPORT (đổi bởi `DEC-177`, làm rõ `DEC-169`).
#
# Lịch sử cần đọc cho đúng: `DEC-169` ghi *"Owner KHÔNG yêu cầu"* import /
# persist / query / display `Summary 2025` — đó là một tuyên bố PHẠM VI
# ("chưa cần"), KHÔNG phải một lệnh cấm sản phẩm ("không được có"). Bối cảnh
# kỹ thuật đi kèm cũng rất cụ thể: sheet đó đã bị dán cứng thành giá trị
# tĩnh (0 ô công thức trên toàn sheet, 99 dòng value-only) nên contract phân
# loại dòng THEO CÔNG THỨC không áp dụng được cho nó.
#
# `DEC-177` (chủ dự án, 2026-09-04) mở lại phạm vi: 2025 CÓ Summary riêng và
# CÓ chi tiết theo nhân viên, và Reports nên giữ được chúng dưới dạng
# LEGACY_REFERENCE. Ngữ nghĩa OPTIONAL_IMPORT vì vậy là:
#
#   - dòng nào contract phân loại được  → NHẬP, origin = LEGACY_REFERENCE;
#   - dòng nào KHÔNG phân loại được     → KHÔNG đoán, KHÔNG bỏ im lặng:
#                                         đếm và báo lên tận giao diện;
#   - sheet vắng mặt hoặc không phân loại được dòng nào
#                                       → KHÔNG làm trượt cả workbook.
#
# Nhánh cuối là điều kiện để `DEC-169` không bị lật ngược thành một hồi quy:
# hình dạng value-only của workbook thật vẫn phải nhập được phần 2026.
SUMMARY_IMPORT_SHEETS = ("Summary 2026",)
SUMMARY_OPTIONAL_SHEETS = ("Summary 2025",)
# Tên cũ, giữ lại cho các tham chiếu hiện có; nghĩa mới = OPTIONAL_IMPORT.
SUMMARY_REFERENCE_ONLY_SHEETS = SUMMARY_OPTIONAL_SHEETS
DATACHART_SHEET = "DataChart 2026"
REQUIRED_SHEETS = SUMMARY_IMPORT_SHEETS + (DATACHART_SHEET,)

SHEET_SCOPE_REQUIRED = "REQUIRED_IMPORT"
SHEET_SCOPE_OPTIONAL = "OPTIONAL_IMPORT"

PRODUCTS_COLUMN = "D"
VS_PREV_MONTH_COLUMN = "I"
# Vùng nhãn nằm bên trái cột dữ liệu đầu tiên (C). Nhãn người bán được lấy
# NGUYÊN VĂN từ đây — không chuẩn hoá, không map sang tên nhân viên của
# pipeline (legacy và pipeline là hai origin tách biệt).
LABEL_COLUMNS = ("A", "B")

ROW_KIND_SELLER = "SELLER"
ROW_KIND_MONTH_TOTAL = "MONTH_TOTAL"
ROW_KIND_YEAR_TOTAL = "YEAR_TOTAL"
ROW_KIND_PROGRESS = "PROGRESS"

# Vùng DataChart 2026 theo docs/analysis/02_FORMULA_MAPPING.md §5.
DATACHART_FIRST_ROW = 3
DATACHART_MONTHS = 12
DATACHART_DAY_COLUMNS = tuple(get_column_letter(index) for index in range(2, 33))
DATACHART_MONTH_TOTAL = "AG"
DATACHART_PREV_YEAR = "AH"
DATACHART_VS_LAST_YEAR = "AI"
DATACHART_VS_TARGET = "AJ"
DATACHART_TARGET_YEAR_CELL = ("J", 15)
DATACHART_TARGET_PER_DAY_CELL = ("P", 15)
# Toàn bộ cột mà DataChart cần đọc — dùng để chỉ giữ đúng những cột này khi
# đọc tuần tự, thay vì dựng cả sheet trong bộ nhớ.
DATACHART_READ_COLUMNS = (
    DATACHART_TARGET_YEAR_CELL[0], DATACHART_TARGET_PER_DAY_CELL[0],
) + DATACHART_DAY_COLUMNS + (
    DATACHART_MONTH_TOTAL, DATACHART_PREV_YEAR,
    DATACHART_VS_LAST_YEAR, DATACHART_VS_TARGET,
)

YEAR_IN_SHEET_NAME = re.compile(r"(20\d{2})")
# Sheet chi tiết "MM.YYYY <Nhãn>" của workbook một-năm độc lập.
DETAIL_SHEET_NAME = re.compile(r"^(?P<month>\d{2})\.(?P<year>20\d{2})\s+(?P<label>.+)$")
STANDALONE_SUMMARY_SHEET = "Summary"

# Khối tổng kết KPI cuối năm, nằm DƯỚI 12 khối tháng của sheet Summary.
# Nhận ra bằng chính TIÊU ĐỀ mà workbook tự viết ở cột C — quan sát được
# đúng MỘT lần trên mỗi sheet Summary của năm 2025 (cả bản độc lập lẫn bản
# nhúng), và KHÔNG xuất hiện trong `Summary 2026`.
#
# Vì sao khối này KHÔNG được nhập: ở đó các cột mang ý nghĩa KHÁC HẲN bảng
# chính — `C` là "Tổng KPI" (cộng cột `N` của 12 tháng), `D` là "KPI trung
# bình", chứ không phải "Tổng đơn" và "Tổng số SP". Nhập chúng vào cùng bộ
# cột sẽ ghi ra những con số sai hẳn nghĩa, ví dụ "Tổng đơn của Ly = 10,79".
# Loại trừ là TƯỜNG MINH và có ghi lại số dòng, không phải nuốt lỗi.
SUMMARY_RECAP_HEADER = "Tổng KPI"
SUMMARY_RECAP_COLUMN = "C"
SAME_SHEET_RATIO = re.compile(r"^=\s*[A-Z]+\$?\d+\s*/\s*\$?[A-Z]+\$?\d+\s*$")


class LegacyImportError(ValueError):
    """Workbook không đúng hình dạng đã freeze cho import legacy."""


def row_has_business_values(values: dict[str, object]) -> bool:
    """Dòng có ít nhất một số ở CỘT DỮ LIỆU đã freeze của Summary.

    Chỉ xét các cột trong ``SUMMARY_COLUMN_FIELDS`` — vùng nhãn (A–B) chứa
    chữ và đôi khi cả số điều hành (số ngày trong năm), không phải giá trị
    nghiệp vụ của một dòng người bán.
    """
    return any(
        isinstance(values.get(column), (int, float))
        and not isinstance(values.get(column), bool)
        for column in SUMMARY_COLUMN_FIELDS
    )


# Kích thước khối đọc file khi băm — hằng số I/O, không phải số nghiệp vụ.
# Viết thành hằng số để bất biến "không có phép tính nào trong app/legacy"
# kiểm chứng được bằng AST (xem tests/test_legacy_importer.py).
FINGERPRINT_CHUNK_BYTES = 1_048_576


def fingerprint_file(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(FINGERPRINT_CHUNK_BYTES), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _as_decimal(value) -> Optional[Decimal]:
    """Giữ nguyên giá trị ô dưới dạng Decimal; mọi thứ không phải số → None.

    Dùng ``str(value)`` để Decimal mang đúng con số Excel hiển thị, không
    kéo theo sai số nhị phân của float.
    """
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, Decimal)):
        return Decimal(value)
    if isinstance(value, float):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    return None


def _label(value_row: dict[str, object]) -> Optional[str]:
    """Nhãn = ô chữ gần cột dữ liệu nhất trong vùng nhãn, giữ nguyên văn."""
    for column in reversed(LABEL_COLUMNS):
        value = value_row.get(column)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def _classify(formulas: dict[str, str]) -> Optional[str]:
    has_cross_sheet = any(defect_rules.sheet_reference(f) for f in formulas.values())
    if has_cross_sheet:
        return ROW_KIND_SELLER
    if any(defect_rules.is_halving_sum(f) for f in formulas.values()):
        return ROW_KIND_YEAR_TOTAL
    if any(defect_rules.sum_span(f) is not None for f in formulas.values()):
        return ROW_KIND_MONTH_TOTAL
    if any(SAME_SHEET_RATIO.match(f) for f in formulas.values()):
        return ROW_KIND_PROGRESS
    return None


def _period_from_formulas(formulas: dict[str, str]) -> Optional[tuple[int, int]]:
    """(year, month) suy từ tên sheet nguồn mà dòng người bán tham chiếu tới."""
    for formula in formulas.values():
        match = defect_rules.sheet_reference(formula)
        if match is None:
            continue
        sheet = match.group("sheet")
        parsed = defect_rules.split_sheet_label(sheet)
        if parsed is None:
            continue
        month_text = sheet.strip().split(".", 1)[0]
        try:
            return parsed[0], int(month_text)
        except ValueError:
            continue
    return None


def _sheet_year(sheet_name: str) -> Optional[int]:
    match = YEAR_IN_SHEET_NAME.search(sheet_name)
    return int(match.group(1)) if match else None


def _sheet_cells(sheet, columns: Iterable[str]) -> dict[int, dict[str, object]]:
    """Đọc TUẦN TỰ một sheet, chỉ giữ lại giá trị của các cột được yêu cầu.

    Truy cập ngẫu nhiên (``sheet["C7"]``) buộc openpyxl phải mở workbook ở
    chế độ đầy đủ, và chế độ đó dựng cây ``Cell`` cho MỌI sheet trong file
    ngay lúc ``load_workbook`` — kể cả hàng chục sheet sổ bán hàng thô mà
    import legacy không hề đọc. Trên workbook thật của Owner (~3 MB, 60+
    sheet nguồn) điều đó ngốn ~380 MB cho một lần parse và làm container
    512 MB bị OOM (S078R).

    Đọc tuần tự trên worksheet read-only thì ngược lại: sheet nào không được
    duyệt thì XML của nó không bao giờ được phân tích, và mỗi lần chỉ có một
    dòng nằm trong bộ nhớ. Kết quả trả về giống HỆT cách đọc cũ — cùng chỉ
    số dòng, cùng giá trị ô — nên không có ngữ nghĩa nghiệp vụ nào đổi.
    """
    wanted = {column_index_from_string(column): column for column in columns}
    if not wanted:
        return {}
    max_column = max(wanted)
    cells: dict[int, dict[str, object]] = {}
    for row_index, row in enumerate(
        sheet.iter_rows(min_col=1, max_col=max_column), start=1
    ):
        cells[row_index] = {
            column: row[index - 1].value if index <= len(row) else None
            for index, column in wanted.items()
        }
    return cells


def _recap_header_row(value_cells: dict[int, dict[str, object]]) -> Optional[int]:
    """Dòng đầu của khối tổng kết KPI, nếu sheet có khối đó."""
    for row_index in sorted(value_cells):
        value = value_cells[row_index].get(SUMMARY_RECAP_COLUMN)
        if isinstance(value, str) and value.strip() == SUMMARY_RECAP_HEADER:
            return row_index
    return None


def _read_rows(value_cells: dict[int, dict[str, object]],
               formula_cells: dict[int, dict[str, object]]):
    for row_index in sorted(value_cells):
        formulas = {
            column: raw
            for column, raw in formula_cells.get(row_index, {}).items()
            if isinstance(raw, str) and raw.startswith("=")
        }
        yield row_index, value_cells[row_index], formulas


def _parse_summary_sheet(
    value_sheet, formula_sheet, *, required: bool = True,
    default_year: Optional[int] = None,
) -> tuple[list[SummaryRow], list[int], list[int]]:
    """Dòng Summary của một sheet + danh sách dòng KHÔNG phân loại được.

    ``required=True`` (REQUIRED_IMPORT): còn một dòng chưa phân loại được là
    FAIL TO — guard DEC-168 nguyên vẹn, vì bỏ sót một dòng production nghĩa
    là báo cáo hiển thị thiếu số mà không ai biết.

    ``required=False`` (OPTIONAL_IMPORT, `DEC-177`): trả về đúng những dòng
    phân loại được, kèm số dòng chưa phân loại được để tầng trên BÁO LÊN
    giao diện. Không đoán, và cũng không im lặng — hai điều đó khác nhau.

    Trả về ``(rows, unaccounted, recap)``. ``recap`` là các dòng thuộc khối
    tổng kết KPI cuối năm — bị loại trừ CÓ CHỦ ĐÍCH (xem
    ``SUMMARY_RECAP_HEADER``), và số dòng được trả về để sự loại trừ đó
    kiểm chứng được thay vì phải tin.
    """
    sheet_name = value_sheet.title
    # Sheet của workbook 2025 độc lập tên đúng là "Summary" — không mang năm.
    # ``default_year`` khi đó đến từ TÊN CÁC SHEET CHI TIẾT của chính workbook
    # (`01.2025 Ly` …), tức vẫn là bằng chứng đọc được từ file, không phải
    # một năm do người viết mã chọn.
    sheet_year = _sheet_year(sheet_name) or default_year
    columns = LABEL_COLUMNS + tuple(SUMMARY_COLUMN_FIELDS)
    rows: list[SummaryRow] = []
    last_period: Optional[tuple[int, int]] = None
    unaccounted: list[int] = []

    value_cells = _sheet_cells(value_sheet, columns)
    formula_cells = _sheet_cells(formula_sheet, columns)
    recap_from = _recap_header_row(value_cells)
    recap_rows: list[int] = []
    for row_index, raw_values, formulas in _read_rows(value_cells, formula_cells):
        if recap_from is not None and row_index >= recap_from:
            # Khối tổng kết KPI — cột mang ý nghĩa khác, KHÔNG nhập. Ghi lại
            # số dòng để sự loại trừ này kiểm chứng được, không phải im lặng.
            if row_has_business_values(raw_values):
                recap_rows.append(row_index)
            continue
        row_kind = _classify(formulas)
        if row_kind is None:
            # Dòng KHÔNG khớp contract phân loại. Nếu nó vẫn mang giá trị
            # nghiệp vụ thì đây là dữ liệu của Owner mà importer không có
            # thẩm quyền diễn giải — ghi lại để báo to ở cuối sheet, TUYỆT
            # ĐỐI không đoán `row_kind` từ việc "dòng có số" (DEC-168).
            if row_has_business_values(raw_values):
                unaccounted.append(row_index)
            continue
        values = {
            field: _as_decimal(raw_values.get(column))
            for column, field in SUMMARY_COLUMN_FIELDS.items()
        }
        cell_values = {c: _as_decimal(raw_values.get(c)) for c in SUMMARY_COLUMN_FIELDS}
        seller_label = _label(raw_values)

        if row_kind == ROW_KIND_SELLER:
            period = _period_from_formulas(formulas)
            if period is not None:
                last_period = period
            year, month = period if period is not None else (sheet_year, None)
            row_defects = defect_rules.detect_seller_defects(
                seller_label=seller_label,
                values=cell_values,
                formulas=formulas,
                products_column=PRODUCTS_COLUMN,
                vs_prev_month_column=VS_PREV_MONTH_COLUMN,
            )
        elif row_kind == ROW_KIND_MONTH_TOTAL:
            # Dòng tổng đóng khối tháng ngay phía trên nó — kỳ của khối đó.
            year, month = last_period if last_period else (sheet_year, None)
            row_defects = defect_rules.detect_month_total_defects(formulas)
        else:
            # YEAR_TOTAL / PROGRESS: cấp năm, không thuộc một tháng nào.
            year, month, row_defects = sheet_year, None, {}

        if year is None:
            raise LegacyImportError(
                f"Không xác định được năm cho dòng {row_index} của sheet "
                f"'{sheet_name}' (tên sheet không chứa năm và dòng không tham "
                "chiếu sheet nguồn nào)."
            )
        rows.append(SummaryRow(
            year=year, month=month, seller_label=seller_label, row_kind=row_kind,
            sheet_name=sheet_name, sheet_row=row_index, values=values,
            formula_text=dict(formulas), known_defects=row_defects,
        ))

    if unaccounted and required:
        # FAIL LOUDLY (DEC-168). Bỏ qua im lặng một dòng có số nghĩa là báo
        # cáo hiển thị thiếu số của Owner mà không ai biết — đúng kiểu sai
        # âm thầm mà cả task này tồn tại để ngăn.
        preview = ", ".join(str(index) for index in unaccounted[:20])
        more = f" (và {len(unaccounted) - 20} dòng nữa)" if len(unaccounted) > 20 else ""
        raise LegacyImportError(
            f"Sheet '{sheet_name}': {len(unaccounted)} dòng có giá trị nghiệp vụ "
            f"nhưng không khớp contract phân loại dòng — dòng {preview}{more}. "
            "Importer KHÔNG đoán ý nghĩa dòng từ việc dòng có số. Đây là "
            "UNKNOWN / OWNER_DECISION_REQUIRED: cần Owner xác nhận ý nghĩa "
            "các dòng này trước khi mở rộng contract."
        )
    return rows, unaccounted, recap_rows


def _parse_datachart(value_sheet, formula_sheet):
    sheet_name = value_sheet.title
    year = _sheet_year(sheet_name)
    if year is None:
        raise LegacyImportError(f"Sheet '{sheet_name}' không chứa năm trong tên.")

    value_cells = _sheet_cells(value_sheet, DATACHART_READ_COLUMNS)
    formula_cells = _sheet_cells(formula_sheet, DATACHART_READ_COLUMNS)

    def _value(column: str, row_index: int):
        return value_cells.get(row_index, {}).get(column)

    def _formula(column: str, row_index: int) -> Optional[str]:
        raw = formula_cells.get(row_index, {}).get(column)
        return raw if isinstance(raw, str) and raw.startswith("=") else None

    target_year = _as_decimal(_value(*DATACHART_TARGET_YEAR_CELL))
    target_per_day = _as_decimal(_value(*DATACHART_TARGET_PER_DAY_CELL))

    daily: list[DailySales] = []
    monthly: list[MonthlyReference] = []
    for offset in range(DATACHART_MONTHS):
        row_index = DATACHART_FIRST_ROW + offset
        month = offset + 1
        for day_offset, column in enumerate(DATACHART_DAY_COLUMNS, start=1):
            amount = _as_decimal(_value(column, row_index))
            # Ô trống = ngày chưa có số, KHÔNG phải doanh số 0. Không lưu
            # dòng rỗng để trạng thái rỗng trên UI vẫn trung thực.
            if amount is None:
                continue
            daily.append(DailySales(
                year=year, month=month, day=day_offset,
                sales_vnd=amount, source_sheet=sheet_name,
            ))
        formula_text = {
            column: formula
            for column in (DATACHART_MONTH_TOTAL, DATACHART_VS_LAST_YEAR, DATACHART_VS_TARGET)
            if (formula := _formula(column, row_index)) is not None
        }
        monthly.append(MonthlyReference(
            year=year, month=month,
            sales_current_year_vnd=_as_decimal(_value(DATACHART_MONTH_TOTAL, row_index)),
            sales_prev_year_vnd=_as_decimal(_value(DATACHART_PREV_YEAR, row_index)),
            vs_last_year_ratio=_as_decimal(_value(DATACHART_VS_LAST_YEAR, row_index)),
            vs_target_ratio=_as_decimal(_value(DATACHART_VS_TARGET, row_index)),
            target_year=target_year, target_per_day=target_per_day,
            formula_text=formula_text,
        ))
    return daily, monthly


def parse_workbook(path: Path) -> LegacyWorkbook:
    """Đọc sheet REQUIRED_IMPORT (bắt buộc) + OPTIONAL_IMPORT (nếu có).

    Thiếu một sheet REQUIRED_IMPORT → lỗi rõ. Sheet OPTIONAL_IMPORT
    (``SUMMARY_OPTIONAL_SHEETS``, hiện là ``Summary 2025``) được đọc khi nó
    tồn tại: dòng nào phân loại được thì nhập, dòng nào không thì được ĐẾM
    và ghi vào ``sheets_imported`` để giao diện nói ra. Sheet vắng mặt hoặc
    không phân loại được dòng nào KHÔNG làm trượt cả workbook (`DEC-177`).
    """
    path = Path(path)
    # `read_only=True` là ràng buộc BỘ NHỚ, không phải tối ưu vặt: ở chế độ
    # đầy đủ, `load_workbook` dựng cây `Cell` cho MỌI sheet của workbook
    # ngay lúc mở — trên workbook thật của Owner đó là 60+ sheet sổ bán
    # hàng thô mà import legacy không đọc dòng nào. Hai workbook (giá trị +
    # công thức) phải sống đồng thời để ghép value/formula theo từng dòng,
    # nên chi phí đó bị nhân đôi và làm container 512 MB bị Render kill
    # (S078R). Ở chế độ read-only, sheet nào không được duyệt thì XML của
    # nó không bao giờ được phân tích.
    values_wb = load_workbook(path, data_only=True, read_only=True)
    formulas_wb = load_workbook(path, data_only=False, read_only=True)
    try:
        missing = [name for name in REQUIRED_SHEETS if name not in values_wb.sheetnames]
        if missing:
            raise LegacyImportError(
                "Workbook thiếu sheet bắt buộc cho import legacy: "
                + ", ".join(missing)
            )
        summary_rows: list[SummaryRow] = []
        sheets_imported: list[dict] = []

        def _record(name: str, scope: str, imported: int, unclassified: list[int],
                    recap: list[int] = ()) -> None:
            entry = {
                "sheet_name": name, "state": values_wb[name].sheet_state,
                "scope": scope, "imported_rows": str(imported),
            }
            if recap:
                entry["recap_rows_excluded"] = str(len(recap))
            if unclassified:
                # Số dòng Owner có số mà contract chưa đọc được. Ghi cả vài
                # số dòng đầu để Owner mở đúng chỗ trong workbook mà đối
                # chiếu, thay vì phải dò cả sheet.
                entry["unclassified_rows"] = str(len(unclassified))
                entry["unclassified_preview"] = ", ".join(
                    str(index) for index in unclassified[:20])
            sheets_imported.append(entry)

        for name in SUMMARY_IMPORT_SHEETS:
            rows, unaccounted, recap = _parse_summary_sheet(
                values_wb[name], formulas_wb[name], required=True)
            summary_rows.extend(rows)
            _record(name, SHEET_SCOPE_REQUIRED, len(rows), unaccounted, recap)

        daily, monthly = _parse_datachart(
            values_wb[DATACHART_SHEET], formulas_wb[DATACHART_SHEET],
        )
        sheets_imported.append({
            "sheet_name": DATACHART_SHEET,
            "state": values_wb[DATACHART_SHEET].sheet_state,
            "scope": SHEET_SCOPE_REQUIRED,
            "imported_rows": str(len(monthly)),
        })

        # OPTIONAL_IMPORT sau REQUIRED: một sự cố ở đây không được phép làm
        # mất phần production đã đọc xong ở trên.
        for name in SUMMARY_OPTIONAL_SHEETS:
            if name not in values_wb.sheetnames:
                continue
            rows, unaccounted, recap = _parse_summary_sheet(
                values_wb[name], formulas_wb[name], required=False)
            summary_rows.extend(rows)
            _record(name, SHEET_SCOPE_OPTIONAL, len(rows), unaccounted, recap)
    finally:
        values_wb.close()
        formulas_wb.close()

    return LegacyWorkbook(
        source_file_name=path.name,
        file_fingerprint=fingerprint_file(path),
        file_size=path.stat().st_size,
        sheets_imported=sheets_imported,
        summary_rows=summary_rows,
        daily_sales=daily,
        monthly_reference=monthly,
    )


# ---------------------------------------------------------------------------
# WORKBOOK MỘT-NĂM ĐỘC LẬP (`DEC-178`) — nguồn chuẩn lịch sử của một năm.
#
# Hình dạng khác hẳn workbook 2026: sheet tổng hợp tên đúng là ``Summary``
# (không mang năm), không có ``DataChart``, và mỗi kỳ × người bán có một
# sheet chi tiết riêng ``MM.YYYY <Nhãn>``.
#
# PHB-04 V1 nhập MỘT sheet: ``Summary``. 74 sheet chi tiết của workbook 2025
# **cố ý KHÔNG được nhập** — xem `LEGACY_LINE_DETAIL_2025 = DEFERRED` ở
# ``docs/tasks/PHB-04-legacy-reference-v1.md``: chúng mang tên, số điện
# thoại và địa chỉ khách hàng, nên đưa vào history store là một quyết định
# quản trị dữ liệu cá nhân, không phải một chi tiết triển khai của PHB-04.
# Tên sheet vẫn được ghi lại làm dấu vết nguồn, không kèm một ô dữ liệu nào.
# ---------------------------------------------------------------------------

SOURCE_AUTHORITY_YEAR = "AUTHORITATIVE_YEAR"
SOURCE_AUTHORITY_SNAPSHOT = "WORKBOOK_SNAPSHOT"

SHEET_SCOPE_DETAIL_NOT_INGESTED = "DETAIL_NOT_INGESTED"


def detail_sheets(sheet_names: Iterable[str]) -> list[tuple[int, int, str]]:
    """(năm, tháng, nhãn) của các sheet chi tiết, theo ĐÚNG tên sheet."""
    found = []
    for name in sheet_names:
        match = DETAIL_SHEET_NAME.match(name)
        if match is None:
            continue
        found.append((
            int(match.group("year")), int(match.group("month")),
            match.group("label").strip(),
        ))
    return found


def is_standalone_year_workbook(sheet_names: Iterable[str]) -> bool:
    names = list(sheet_names)
    return STANDALONE_SUMMARY_SHEET in names and bool(detail_sheets(names))


def parse_year_workbook(path: Path) -> LegacyWorkbook:
    """Đọc workbook lịch sử MỘT NĂM độc lập (ví dụ `Báo cáo Kinh doanh 2025`).

    Năm của workbook suy ra từ TÊN các sheet chi tiết, không từ tên file và
    không từ một hằng số: một workbook 2024 cùng hình dạng sẽ tự cho ra 2024.
    Nhiều năm lẫn trong một file là hình dạng chưa từng được xác nhận, nên
    nó FAIL TO thay vì bị đoán.

    Guard `DEC-168` áp dụng ĐẦY ĐỦ ở đây: ``Summary`` là sheet REQUIRED, nên
    một dòng có giá trị nghiệp vụ mà contract không phân loại được vẫn làm
    trượt cả lần nhập. Đây là nguồn CHUẨN của năm đó — bỏ sót một dòng của
    nó nghĩa là báo cáo lịch sử thiếu số mà không ai biết.
    """
    path = Path(path)
    values_wb = load_workbook(path, data_only=True, read_only=True)
    formulas_wb = load_workbook(path, data_only=False, read_only=True)
    try:
        names = values_wb.sheetnames
        if STANDALONE_SUMMARY_SHEET not in names:
            raise LegacyImportError(
                "Workbook lịch sử một năm phải có sheet "
                f"'{STANDALONE_SUMMARY_SHEET}'. Sheet hiện có: "
                + ", ".join(names[:10])
            )
        details = detail_sheets(names)
        years = {year for year, _month, _label in details}
        if not years:
            raise LegacyImportError(
                "Không tìm thấy sheet chi tiết nào dạng 'MM.YYYY <Nhãn>', nên "
                "không xác định được năm của workbook. KHÔNG đoán năm từ tên file."
            )
        if len(years) > 1:
            raise LegacyImportError(
                "Workbook chứa sheet chi tiết của NHIỀU năm ("
                + ", ".join(str(year) for year in sorted(years))
                + "). Hình dạng này chưa được xác nhận — cần chủ dự án xác nhận "
                "ý nghĩa trước khi nhập."
            )
        year = years.pop()

        summary_rows, _unaccounted, recap = _parse_summary_sheet(
            values_wb[STANDALONE_SUMMARY_SHEET],
            formulas_wb[STANDALONE_SUMMARY_SHEET],
            required=True, default_year=year,
        )
        summary_entry = {
            "sheet_name": STANDALONE_SUMMARY_SHEET,
            "state": values_wb[STANDALONE_SUMMARY_SHEET].sheet_state,
            "scope": SHEET_SCOPE_REQUIRED,
            "imported_rows": str(len(summary_rows)),
            "year": str(year),
        }
        if recap:
            summary_entry["recap_rows_excluded"] = str(len(recap))
        sheets_imported = [summary_entry]
        # Sheet chi tiết: GHI TÊN, KHÔNG ĐỌC Ô NÀO. Dấu vết để chủ dự án
        # thấy chúng tồn tại và đang bị hoãn có chủ đích, không phải bị bỏ sót.
        for detail_year, month, label in sorted(details):
            sheets_imported.append({
                "sheet_name": f"{month:02d}.{detail_year} {label}",
                "state": "visible",
                "scope": SHEET_SCOPE_DETAIL_NOT_INGESTED,
                "imported_rows": "0",
            })
    finally:
        values_wb.close()
        formulas_wb.close()

    return LegacyWorkbook(
        source_file_name=path.name,
        file_fingerprint=fingerprint_file(path),
        file_size=path.stat().st_size,
        sheets_imported=sheets_imported,
        summary_rows=summary_rows,
        daily_sales=[],
        monthly_reference=[],
        source_authority=SOURCE_AUTHORITY_YEAR,
    )
