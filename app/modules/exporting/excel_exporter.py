"""Trình bày kết quả đã tính; không tra giá hay tính lại nghiệp vụ."""

from __future__ import annotations

from collections import Counter, defaultdict, deque
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.domain.models import (
    PRICE_SOURCE_HISTORICAL_CONFIRMED_REPORT,
    PRICE_SOURCE_OWNER_MANUAL_LEGACY_CONFIRMATION,
    RawRow,
    WorkingLine,
)
from app.modules.pricing.resolution.composition import PriceResolutionRecord
from app.modules.pricing.resolution.unresolved_descriptions import (
    UnresolvedDescriptionGroup,
    aggregate_unresolved_descriptions,
)
from app.modules.validation.models import (
    CATEGORY_MISSING_PURCHASE_PRICE, SEVERITY_ERROR, SEVERITY_INFO, SCOPE_ORDER,
)
from app.pipeline import ImportResult


LINE_FIELDS = (
    "Ngày bán", "OrderID / Số BH", "Nhân viên", "Sản phẩm", "Số lượng",
    "Doanh thu", "Giá nhập kế toán / công khai", "Lợi nhuận kế toán",
    "Lợi nhuận KPI", "Trạng thái", "Pending reason", "Nguồn giá",
    "Dòng nguồn", "Trạng thái đơn",
)
REVIEW_FIELDS = (
    "Ngày bán", "OrderID", "Nhân viên", "Sản phẩm", "Trạng thái", "Lý do",
    "Chi tiết", "Tệp nguồn", "Sheet nguồn", "Dòng nguồn", "Namespace",
    "Mã sản phẩm", "Raw identity key", "Nguồn giá", "Quy tắc giá",
    "Capture giá", "Capture danh mục", "Identity revision", "Tracking reason",
    "Fallback blocked by", "Fallback detail", "KPI provenance",
)
UNRESOLVED_FIELDS = (
    "Tên hàng trên chứng từ", "Khoá inv.map (Tracking)", "Số dòng", "Số đơn",
    "Doanh thu (VND)", "Số cách viết", "Lý do chưa định danh",
)
"""PHB-01 §4B — CỘT ĐẦU TIÊN là câu tên hàng nguyên văn, cố ý.

Đường dùng thật là: Owner bôi đen cột này trong Excel, Ctrl+C, rồi dán thẳng
vào ô nhập của màn "Phân loại theo tên hàng" bên Tracking. Excel chép nhiều
cột ra clipboard bằng ký tự TAB, nên bên Tracking chỉ cần lấy phần trước dấu
TAB đầu tiên là ra đúng câu tên hàng — KHÔNG cắt theo dấu phẩy, vì tên hàng
thật có dấu phẩy ("Tivi Samsung 55, model 2026"). Đổi thứ tự cột ở đây là làm
hỏng thao tác dán ấy.
"""

MONEY_FORMAT = '#,##0;[Red](#,##0);"0"'


class ReportIntegrityError(ValueError):
    """Không thể xuất báo cáo đầy đủ, đối chiếu được với nguồn."""


@dataclass(frozen=True)
class ReportSummary:
    input_orders: int
    accounted_orders: int
    total_lines: int
    auto_orders: int
    review_orders: int
    review_lines: int
    error_count: int
    review_reason_counts: dict[str, int]
    unresolved_description_count: int = 0
    """PHB-01 — số CÂU TÊN HÀNG (khoá `inv.map`) đang chờ Owner phân loại.

    Mặc định `0` để mọi lời gọi `ReportSummary(...)` đã có (test, telemetry
    cũ) giữ nguyên hành vi; đường production luôn truyền giá trị thật.
    Đếm theo khoá chứ không theo dòng — đó mới là số lần Owner phải bấm.
    """

    @property
    def order_accounting_rate(self) -> float:
        return self.accounted_orders / self.input_orders


@dataclass(frozen=True)
class _PresentedLine:
    line: WorkingLine
    record: PriceResolutionRecord | None
    reasons: tuple[str, ...]
    details: tuple[str, ...]

    @property
    def status(self) -> str:
        return "PENDING" if self.reasons else "AUTO"


def _record_key(order_id, product, sale_date):
    return order_id, product, sale_date


def _present_lines(result, records, raw_rows):
    lines = sorted(
        (line for order in result.orders for line in order.lines),
        key=lambda line: line.raw.source_row,
    )
    # Đối chiếu cả dòng anh em và số lần xuất hiện, không chỉ đếm OrderID.
    def source_key(raw):
        return (raw.source_file, raw.source_sheet, raw.source_row,
                raw.order_id, raw.row_hash)

    if not raw_rows:
        raise ReportIntegrityError("Không có đơn hàng trong workbook nguồn.")
    if Counter(map(source_key, raw_rows)) != Counter(source_key(l.raw) for l in lines):
        raise ReportIntegrityError("Dòng nguồn và kết quả production không khớp.")
    if Counter(order.order_id for order in result.orders) != Counter(
        {raw.order_id: 1 for raw in raw_rows}
    ) or any(line.order_id != order.order_id
             for order in result.orders for line in order.lines):
        raise ReportIntegrityError("OrderID nguồn và kết quả production không khớp.")

    by_key = defaultdict(deque)
    for record in records:
        by_key[_record_key(record.order_id, record.raw_product_identity,
                           record.sale_date)].append(record)

    presented = []
    for line in lines:
        record = None
        if line.price_source not in (
            PRICE_SOURCE_HISTORICAL_CONFIRMED_REPORT,
            PRICE_SOURCE_OWNER_MANUAL_LEGACY_CONFIRMATION,
        ):
            matches = by_key[_record_key(line.order_id, line.product_raw, line.date)]
            if not matches:
                raise ReportIntegrityError("Thiếu PriceResolutionRecord của dòng.")
            record = matches.popleft()
            if (record.price_vnd != line.accounting_purchase_price
                    or record.price_source != line.price_source):
                raise ReportIntegrityError("Bản ghi giá không khớp kết quả dòng.")

        reasons, details = [], []
        if record and record.reason:
            reasons.append(record.reason.value)
            details.append(record.detail)
        # Chỉ chiếu finding lên đúng dòng/phạm vi đơn mà provenance chỉ tới.
        for item in result.review_queue.items:
            if (item.severity == SEVERITY_INFO
                    and item.category != CATEGORY_MISSING_PURCHASE_PRICE):
                continue
            applies = any(
                row.source_file == line.raw.source_file
                and row.source_row == line.raw.source_row
                for row in item.provenance.rows
            ) or (item.scope == SCOPE_ORDER and item.order_id == line.order_id
                  and item.source_file == line.raw.source_file)
            if applies:
                reasons.append(item.category)
                # Record có lý do giá chính xác hơn thông báo aggregate cũ.
                if not (item.category == CATEGORY_MISSING_PURCHASE_PRICE and record):
                    details.append(item.message)

        # Chỉ phơi bày các kết quả còn trống, không tự suy đoán nguyên nhân.
        #
        # DEC-PAN-001 (PRICE_AUTHORITY_NORMALIZATION) — vòng lặp này CHỈ còn
        # `eligible_kpi_profit`. Hai trường `accounting_*` đã bị gỡ khỏi đường
        # sinh reason vì Owner đã xác nhận Reports KHÔNG có nguồn giá nhập kế
        # toán độc lập: `accounting_purchase_price` là LEGACY_INTERNAL_PP_CARRIER
        # chở đúng Tracking PP tại ngày bán, và `accounting_profit` là
        # LEGACY_DERIVED_FIELD suy ra từ nó. Khi PP không resolve được thì cả
        # hai cùng None — phát ba mã cho MỘT nguyên nhân gốc khiến người đọc
        # tưởng có ba việc phải làm, trong khi việc thật đã nằm ở
        # `Missing.PurchasePrice`/`IDENTITY_*`. Tên trường legacy không tự nó
        # tạo ra một business authority (bài học ADR/DEC-PAN-001).
        #
        # `eligible_kpi_profit` GIỮ LẠI vì nó KHÔNG phải hệ quả dẫn xuất thuần
        # tuý: khi authority KPI hỏng (`config/eligible_costs.yaml` thiếu/sai,
        # hoặc confirmed-adjustment source UNAVAILABLE) nó là None NGAY CẢ KHI
        # identity đã nhận diện, PP đã resolve và `kpi_purchase_price` đã có —
        # lúc đó đây là mã DUY NHẤT báo cho người vận hành rằng một authority
        # cần được sửa (xem `tests/test_demo.py::
        # test_kpi_unavailable_is_queued_even_with_resolved_price`). Gỡ nó đi
        # sẽ giấu mất một lỗi authority thật.
        for field, label in (
            ("eligible_kpi_profit", "Lợi nhuận KPI"),
        ):
            if getattr(line, field) is None:
                reasons.append(f"Pending.{field}")
                details.append(f"{label}: production chưa trả kết quả.")
        presented.append(_PresentedLine(
            line, record, tuple(dict.fromkeys(reasons)), tuple(dict.fromkeys(details))
        ))
    if any(by_key.values()):
        raise ReportIntegrityError("Dư PriceResolutionRecord không gắn được vào dòng.")
    return presented


# Alias public: cùng object đã in ra XLSX — không có nguồn sự thật thứ hai.
PresentedLine = _PresentedLine
present_lines = _present_lines


def unresolved_description_groups(
    views: "list[_PresentedLine]",
) -> tuple[UnresolvedDescriptionGroup, ...]:
    """PHB-01 §4A — các câu tên hàng chờ phân loại, lấy từ CHÍNH các dòng đã
    trình bày.

    Đi qua `views` chứ không qua `records` thô vì hai lý do, cả hai đều là
    tính đúng đắn chứ không phải tiện tay:

    1. `_present_lines()` đã đối chiếu 1-1 số dòng/OrderID với workbook nguồn
       và ném `ReportIntegrityError` nếu lệch. Đếm trên tập đã qua cửa ấy thì
       "số dòng ảnh hưởng" không thể nhiều hơn hay ít hơn dữ liệu thật.
    2. Doanh thu nằm ở `WorkingLine`, không nằm ở bản ghi giá. Ghép ở đây là
       ghép đúng cặp dòng-bản ghi mà báo cáo đang in ra, không phải một phép
       nối lại tự làm.
    """
    return aggregate_unresolved_descriptions(
        (view.record, view.line.total_sales)
        for view in views
        if view.record is not None
    )


def _append(sheet, values):
    sheet.append(values)
    # Chuỗi nguồn là dữ liệu hiển thị, không phải công thức Excel.
    for cell in sheet[sheet.max_row]:
        if isinstance(cell.value, str):
            cell.data_type = "s"


def _style_table(sheet, widths, *, status_column=None):
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="17365D")
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 38
    for row in sheet.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            cell.font = Font(name="Calibri", size=11, color="203047")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F1F5F9")
            if isinstance(cell.value, str):
                width = max(1, widths[cell.column - 1] - 2)
                max_lines = max(max_lines, sum(
                    max(1, (len(part) + width - 1) // width)
                    for part in cell.value.split("\n")
                ))
        sheet.row_dimensions[row[0].row].height = min(409, max(30, max_lines * 16))
        if status_column:
            cell = row[status_column - 1]
            cell.fill = PatternFill("solid", fgColor=(
                "DCFCE7" if cell.value == "AUTO" else "FEF3C7"
            ))
        row[0].number_format = "dd/mm/yyyy"


def export_report(
    result: ImportResult,
    records: tuple[PriceResolutionRecord, ...],
    raw_rows: list[RawRow],
    *,
    sales_path: Path,
    tracking_capture: Path,
    tracking_catalog: Path,
    output_path: Path,
    processed_at: datetime,
) -> ReportSummary:
    """Xuất một snapshot kết quả; ô trống giữ nguyên nghĩa chưa xác định."""
    views = _present_lines(result, records, raw_rows)
    unresolved = unresolved_description_groups(views)
    pending_orders = {v.line.order_id for v in views if v.reasons}
    review_reason_counts: dict[str, int] = defaultdict(int)
    for view in views:
        for reason in view.reasons:
            review_reason_counts[reason] += 1
    summary = ReportSummary(
        input_orders=len({r.order_id for r in raw_rows}),
        accounted_orders=len(result.orders), total_lines=len(views),
        auto_orders=len(result.orders) - len(pending_orders),
        review_orders=len(pending_orders),
        review_lines=sum(bool(v.reasons) for v in views),
        error_count=len(result.review_queue.by_severity(SEVERITY_ERROR)),
        review_reason_counts=dict(review_reason_counts),
        unresolved_description_count=len(unresolved),
    )
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    line_sheet = workbook.create_sheet("Order Lines")
    review_sheet = workbook.create_sheet("Review Queue")
    unresolved_sheet = workbook.create_sheet("Chưa định danh")
    _append(line_sheet, LINE_FIELDS)
    _append(review_sheet, REVIEW_FIELDS)
    _append(unresolved_sheet, UNRESOLVED_FIELDS)
    # PHB-01 §4B — MỘT dòng cho MỖI khoá inv.map, thứ tự xác định do
    # `aggregate_unresolved_descriptions()` chốt. Sheet luôn có mặt kể cả khi
    # rỗng: một sheet biến mất khiến người dùng phải đoán xem "không có mục
    # nào" hay "bản báo cáo này chưa có tính năng đó".
    for group in unresolved:
        _append(unresolved_sheet, (
            group.raw_description, group.inv_map_key, group.line_count,
            group.order_count, group.revenue_vnd, group.description_variants,
            "\n".join(group.pending_reasons),
        ))
    for view in views:
        line, record = view.line, view.record
        employee = line.employee_normalized or line.employee_raw
        _append(line_sheet, (
            line.date, line.order_id, employee, line.product_raw, line.quantity,
            line.total_sales, line.accounting_purchase_price, line.accounting_profit,
            line.eligible_kpi_profit, view.status, "\n".join(view.reasons),
            line.price_source, line.raw.source_row,
            "REVIEW_QUEUE" if line.order_id in pending_orders else "AUTO",
        ))
        if not view.reasons:
            continue
        identity = record.identity if record else None
        evidence = record.evidence if record else None
        reconstruction = record.tracking_reconstruction if record else None
        _append(review_sheet, (
            line.date, line.order_id, employee, line.product_raw, view.status,
            "\n".join(view.reasons), "\n".join(view.details), line.raw.source_file,
            line.raw.source_sheet, line.raw.source_row,
            identity.namespace.value if identity else None,
            identity.source_product_code if identity else None,
            record.raw_identity_key if record else None, line.price_source,
            record.rule.value if record else None,
            evidence.tracking_price_history_capture_id if evidence else None,
            evidence.tracking_catalog_capture_id if evidence else None,
            evidence.identity_store_revision if evidence else None,
            reconstruction.reason.value if reconstruction and reconstruction.reason else None,
            record.fallback_blocked_by.value if record and record.fallback_blocked_by else None,
            record.fallback_blocked_detail if record else None,
            line.kpi_purchase_price_provenance,
        ))
    # Finding cấp lô không có dòng (ví dụ master nhân viên thiếu doanh số)
    # không được gán nhầm cho mọi đơn hoặc làm tăng số đơn Review Queue.
    batch_items = [item for item in result.review_queue.items
                   if item.severity != SEVERITY_INFO
                   and not item.provenance.rows and not item.order_id]
    for item in batch_items:
        _append(review_sheet, (None, None, None, None, "REVIEW_BATCH",
                              item.category, item.message, item.source_file))

    _append(summary_sheet, ("REPORTS — DEMO V1", "Giá trị"))
    kpi_values = [v.line.eligible_kpi_profit for v in views
                  if v.line.eligible_kpi_profit is not None]
    revenues = [v.line.total_sales for v in views if v.line.total_sales is not None]
    fields = (
        ("Tệp đầu vào", sales_path.name),
        ("Thời điểm xử lý", processed_at.isoformat(timespec="seconds")),
        ("Tổng đơn đầu vào", summary.input_orders),
        ("Đơn đã đối chiếu", summary.accounted_orders),
        ("Tổng dòng", summary.total_lines),
        ("AUTO — số đơn", summary.auto_orders),
        ("PENDING / Review Queue — số đơn", summary.review_orders),
        ("Tỷ lệ đối chiếu đơn", summary.order_accounting_rate),
        ("Tỷ lệ tự động — theo đơn", summary.auto_orders / summary.input_orders),
        ("Tổng doanh thu đã xác định (VND)", sum(revenues, Decimal(0)) if revenues else None),
        ("Tổng lợi nhuận KPI đủ điều kiện (VND)", sum(kpi_values, Decimal(0)) if kpi_values else None),
        ("Dòng có lợi nhuận KPI", len(kpi_values)),
        ("Dòng chưa xác định doanh thu", len(views) - len(revenues)),
        ("Dòng cần Review Queue", summary.review_lines),
        ("Tên hàng chưa định danh — chờ phân loại", summary.unresolved_description_count),
        ("Finding cấp lô cần xem", len(batch_items)),
        ("Capture giá đầu vào", tracking_capture.name),
        ("Capture danh mục đầu vào", tracking_catalog.name),
        ("Đọc trạng thái", "AUTO: mọi dòng trong đơn có kết quả và không có finding cần xem. "
         "Review Queue: ít nhất một dòng cần kiểm tra. REVIEW_BATCH không tính vào số đơn."),
        ("Đọc số tiền", "Đơn vị VND. Ô trống là chưa xác định, không phải 0. Tổng chỉ cộng "
         "giá trị đã có; đây là snapshot, không tính lại khi sửa ô."),
        ("Nguồn giá", "Giá nhập theo kết quả production từng dòng. Giá lịch sử có nguồn "
         "riêng, không gắn nhãn thành Tracking PP. Không dùng PP YAML cũ."),
    )
    for field in fields:
        _append(summary_sheet, field)
    _style_table(summary_sheet, [45, 92])
    summary_sheet.freeze_panes = "B2"
    summary_sheet.auto_filter.ref = None
    for row in (9, 10):
        summary_sheet.cell(row, 2).number_format = "0.0%"
    for row in (11, 12):
        summary_sheet.cell(row, 2).number_format = MONEY_FORMAT
    _style_table(line_sheet, [14, 20, 22, 43, 12, 20, 23, 22, 22, 14, 48, 34, 12, 20],
                 status_column=10)
    for row in line_sheet.iter_rows(min_row=2, min_col=5, max_col=9):
        for cell in row:
            if cell.column == 5:
                cell.number_format = (
                    "#,##0" if cell.value is None or cell.value == int(cell.value)
                    else "#,##0.##"
                )
            else:
                cell.number_format = MONEY_FORMAT
    _style_table(review_sheet, [14, 20, 22, 43, 18, 48, 95, 28, 28, 12, 20, 24,
                               40, 34, 30, 35, 35, 16, 36, 36, 65, 42], status_column=5)
    _style_table(unresolved_sheet, [60, 40, 10, 10, 20, 14, 40])
    # `_style_table` đóng băng ở "E2" cho các bảng có bốn cột định danh đầu;
    # bảng này chỉ có MỘT cột chữ dài nên đóng băng ngay sau nó.
    unresolved_sheet.freeze_panes = "B2"
    for row in unresolved_sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Không ghi đè workbook/capture/artifact đã có của Owner.
    with output_path.open("xb") as handle:
        workbook.save(handle)
    workbook.close()
    return summary
