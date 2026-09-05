"""So hai nguồn lịch sử 2025 — bằng chứng cho `DEC-178`, KHÔNG phải sản phẩm.

Hai nguồn cùng nói về năm 2025:

* SOURCE_A — workbook lịch sử một năm độc lập (`Báo cáo Kinh doanh 2025`).
  Sheet ``Summary`` còn NGUYÊN công thức, liên kết tới 74 sheet chi tiết.
  `DEC-178`: đây là nguồn CHUẨN.
* SOURCE_B — bản sao ``Summary 2025`` nhúng trong workbook 2026. Value-only,
  không còn công thức. Bằng chứng THỨ CẤP.

Script này KHÔNG hoà giải hai nguồn và KHÔNG sinh ra một con số thứ ba. Nó
chỉ ĐO độ lệch để báo cáo, theo đúng ô (dòng × cột) — hai sheet có cùng bố
cục 755 dòng nên so theo toạ độ ô là so đúng cặp giá trị, không cần đoán
dòng nào ứng với dòng nào.

Cách chạy:

    python3 -m tools.analysis.compare_legacy_2025_sources SOURCE_A.xlsx SOURCE_B.xlsx

Không ghi gì vào database. Không sửa file nào.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.legacy.models import SUMMARY_COLUMN_FIELDS
from app.legacy.parser import STANDALONE_SUMMARY_SHEET

SOURCE_B_SHEET = "Summary 2025"

# "Chỉ là làm tròn" được xác định bằng CƠ CHẾ, không bằng một ngưỡng tự chọn:
# B chỉ là bản làm tròn của A khi B == round(A, n) với một n nhỏ. Cách này
# giải thích được ĐIỀU GÌ đã xảy ra (bản sao nhúng lưu giá trị đã làm tròn để
# hiển thị) thay vì chỉ nói "lệch ít". Nó cũng không thể bị vặn cho ra kết
# quả mong muốn: một đơn hàng lệch (105 so với 104) không bao giờ là
# round(105, n), nên nó luôn lộ ra là SUBSTANTIVE.
ROUNDING_DECIMALS = range(0, 7)

EXACT_MATCH = "EXACT_MATCH"
ROUNDING_ONLY = "ROUNDING_ONLY"
SUBSTANTIVE_DIFFERENCE = "SUBSTANTIVE_DIFFERENCE"
MISSING_IN_A = "MISSING_IN_A"
MISSING_IN_B = "MISSING_IN_B"

CLASSES = (EXACT_MATCH, ROUNDING_ONLY, SUBSTANTIVE_DIFFERENCE,
           MISSING_IN_A, MISSING_IN_B)


@dataclass
class Comparison:
    counts: dict = field(default_factory=lambda: {name: 0 for name in CLASSES})
    substantive: list[tuple] = field(default_factory=list)

    @property
    def total(self) -> int:
        return sum(self.counts.values())


def _decimal(value):
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return None
    return Decimal(str(value))


def classify(a, b) -> str:
    if a is None and b is None:
        return EXACT_MATCH
    if a is None:
        return MISSING_IN_A
    if b is None:
        return MISSING_IN_B
    if a == b:
        return EXACT_MATCH
    for places in ROUNDING_DECIMALS:
        if b == round(a, places):
            return ROUNDING_ONLY
    return SUBSTANTIVE_DIFFERENCE


def compare(source_a: Path, source_b: Path) -> Comparison:
    book_a = load_workbook(source_a, data_only=True)
    book_b = load_workbook(source_b, data_only=True)
    try:
        sheet_a = book_a[STANDALONE_SUMMARY_SHEET]
        sheet_b = book_b[SOURCE_B_SHEET]
        result = Comparison()
        last_row = max(sheet_a.max_row or 0, sheet_b.max_row or 0)
        for row in range(1, last_row + 1):
            label = sheet_a.cell(row, 2).value or sheet_b.cell(row, 2).value
            for column in SUMMARY_COLUMN_FIELDS:
                a = _decimal(sheet_a[f"{column}{row}"].value)
                b = _decimal(sheet_b[f"{column}{row}"].value)
                if a is None and b is None:
                    continue
                verdict = classify(a, b)
                result.counts[verdict] += 1
                if verdict == SUBSTANTIVE_DIFFERENCE:
                    result.substantive.append((row, column, label, a, b))
        return result
    finally:
        book_a.close()
        book_b.close()


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 2
    result = compare(Path(argv[0]), Path(argv[1]))
    for name in CLASSES:
        print(f"{name:24} = {result.counts[name]}")
    print(f"{'TOTAL_CELLS_COMPARED':24} = {result.total}")
    print(f"{'SUBSTANTIVE_ROWS':24} = "
          f"{len({item[0] for item in result.substantive})}")
    for row, column, label, a, b in result.substantive[:40]:
        print(f"  DIFF {column}{row} [{label}] A={a} B={b}")
    if len(result.substantive) > 40:
        print(f"  ... và {len(result.substantive) - 40} ô nữa")
    # Lệch KHÔNG phải lỗi: `DEC-178` đã chốt nguồn nào thắng. Script này là
    # công cụ ĐO, nên nó luôn thoát 0 khi đọc được cả hai file.
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
