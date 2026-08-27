"""Reconcile ConversionScheme resolution against the real reference workbook.

CHECK-108A1-14 and CHECK-108A1-15.

**Every dimension used to build an expected result is derived from production
config or the production mapper.** Nothing in this file says "row X means
group Y" — an earlier version did, and a hand-written mapping that exists only
to make the engine agree with the workbook produces a PASS that proves nothing.

That honesty has a cost, and it is stated rather than hidden: only Summary rows
whose label **is** a production employee (`normalized` in
`config/employees.yaml`) can be reconciled independently. The rows labelled
`Nội thành` and `Gia dụng` are report-layer aggregates — the workbook's own
channel sheets carry no employee column at all — and `Linh` / `Fanpage` are
legacy names deliberately absent from master data (DEC-127 §8). For those there
is no production artifact linking the label to an Employee, EmployeeGroup or
ProductGroup, so they are reported as **UNVERIFIABLE**, not as passes.

What is verified for the rows that can be:

    label -> employees.yaml -> (normalized, group, default_lead_source)
                            -> lead_source.yaml (classifier) -> reachable sources
                            -> conversion_rates.yaml (resolver) -> reachable rates
    and every rate the workbook formula uses must be one of those.

This works TODAY even though TASK-108B is blocked: the profit figures stay in
the workbook, so nothing here needs `EligibleKpiProfit`.

Real data files are never committed and are deleted after use (DEC-108).

Usage:
    python tools/analysis/reconcile_conversion.py --workbook <path> [--raw <path>]
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from app.modules.conversion.scheme_resolver import ConversionSchemeResolver  # noqa: E402
from app.modules.domain.models import ADS  # noqa: E402
from app.modules.lead_source.classifier import LeadSourceClassifier  # noqa: E402
from app.modules.mapping.employee_mapper import (  # noqa: E402
    EmployeeMapper,
    load_employee_master,
)
# F1–F5 now live in production (TD-001, TASK-110). Re-exported here unchanged
# so this script's behaviour and output stay exactly as signed off in
# CHECK-108A1-15 — that output is shipped evidence, not a draft.
from app.modules.validation.employee_mapping import (  # noqa: E402
    MappingInput,
    collect_mapping_stats,
    evaluate_raw_mapping,
    norm,
)

CONFIG = Path(__file__).resolve().parents[2] / "config"


def production_employees() -> dict[str, dict]:
    """`normalized name -> its own row` straight out of production config.

    This is the only bridge between a Summary row label and the engine's
    dimensions, and it is a lookup into the real file — not a table written
    here to make the numbers agree.
    """
    # HD-110-10: đi qua biên canonical thay vì `load_yaml` thô, nên chế độ
    # chỉ-`--workbook` cũng fail-fast trên master hỏng. Không đổi logic đối
    # chiếu, không đổi vòng khớp prefix riêng đã freeze.
    master = load_employee_master(CONFIG / "employees.yaml")
    return {norm(row["normalized"]): row for row in master.records}


def reachable_rates(
    employee_row: dict,
    classifier: LeadSourceClassifier,
    resolver: ConversionSchemeResolver,
    as_of: date,
) -> dict[str, Decimal]:
    """Every rate production can produce for this employee on this date.

    Reachable lead sources come from the production classifier: the automatic
    verdict with no ADS keyword present, plus ADS (the keyword rule can raise
    any order). Tín Phát's `default_lead_source: ADS` therefore narrows its
    reachable set to ADS alone — which is a stricter check, not a looser one.
    """
    normalized = norm(employee_row["normalized"])
    without_keyword = classifier.classify_auto(
        [], employee_row.get("default_lead_source"), normalized
    ).lead_source

    rates: dict[str, Decimal] = {}
    for lead_source in {without_keyword, ADS}:
        got = resolver.resolve_auto(
            employee=normalized,
            employee_group=employee_row.get("group"),
            lead_source=lead_source,
            product_group=resolver.default_product_group,
            as_of=as_of,
        )
        if got.rate is not None:
            rates[lead_source] = got.rate
    return rates


def reconcile_summary(
    workbook: Path,
    classifier: LeadSourceClassifier,
    resolver: ConversionSchemeResolver,
) -> int:
    employees = production_employees()
    wb = openpyxl.load_workbook(workbook, data_only=False)
    ws = wb["Summary 2026"]

    label_at_row: dict[int, str] = {}
    period_at_row: dict[int, date] = {}
    current: date | None = None
    for row in ws.iter_rows(min_col=1, max_col=2):
        a, b = row[0].value, row[1].value
        if hasattr(a, "year"):
            current = a.date() if hasattr(a, "date") else a
        label = norm(b)
        if label and label != "Tiến độ":
            label_at_row[row[1].row] = label
            if current:
                period_at_row[row[1].row] = current

    verified = mismatched = 0
    unverifiable: Counter = Counter()
    details: list[str] = []

    for cell in (c for r in ws.iter_rows(min_col=6, max_col=6) for c in r):
        formula = cell.value
        if not isinstance(formula, str) or not formula.startswith("="):
            continue
        if "SUM" in formula or "%" not in formula:
            continue
        label = label_at_row.get(cell.row)
        as_of = period_at_row.get(cell.row)
        if not label or not as_of:
            continue

        employee_row = employees.get(label)
        if employee_row is None:
            unverifiable[label] += 1
            continue

        wanted = {Decimal(r) / 100 for r in re.findall(r"(\d+(?:\.\d+)?)%", formula)}
        rates = reachable_rates(employee_row, classifier, resolver, as_of)
        if wanted <= set(rates.values()):
            verified += 1
        else:
            mismatched += 1
            details.append(
                f"  [LỆCH] {as_of:%m.%Y} {label:10} {cell.coordinate:5} "
                f"workbook dùng {sorted(wanted)} nhưng engine phân giải {rates}"
            )

    wb.close()

    total_unverifiable = sum(unverifiable.values())
    print("=" * 78)
    print("CHECK-108A1-14 — Đối chiếu cột F, Summary 2026")
    print("=" * 78)
    print(f"  VERIFIED — đối chiếu độc lập được : {verified + mismatched}")
    print(f"      khớp                          : {verified}")
    print(f"      MISMATCH                      : {mismatched}")
    print(f"  LIMITED  — không đối chiếu được   : {total_unverifiable}")
    for label, count in unverifiable.most_common():
        print(f"      {label:12} {count:3}  — nhãn không phải Employee trong "
              "config/employees.yaml")
    if details:
        print()
        for line in details:
            print(line)
    print()
    print("  GIỚI HẠN (không tạo PASS giả): các nhãn trên là bút toán gộp ở tầng")
    print("  báo cáo (`Nội thành`, `Gia dụng`) hoặc tên legacy ngoài master data")
    print("  (`Linh`, `Fanpage`). Không có artifact production nào nối nhãn đó với")
    print("  Employee/EmployeeGroup/ProductGroup, nên chúng KHÔNG được tính là")
    print("  đối chiếu thành công. Sheet kênh trong chính workbook cũng không có")
    print("  cột nhân viên để suy ra.")
    print()
    return mismatched


def reconcile_raw(raw: Path, mapper: EmployeeMapper) -> int:
    # HD-110-12: script dựng ĐÚNG một `MappingStats` canonical rồi trao cho
    # tiêu chí, thay vì tự đếm và tự khớp prefix song song với production. Nhờ
    # `enumerate` từ `min_row=6` mà nó có số dòng thật, nên provenance của mọi
    # finding dẫn từ dòng thô có thật — và con số trong message đọc từ chính
    # provenance đó, không từ một counter thứ hai (RC-3).
    #
    # Vòng khớp prefix riêng của script đã biến mất cùng bộ đếm song song:
    # giữ nó lại là giữ một bản cài đặt thứ hai của cùng một quy tắc, đúng
    # thứ INVARIANT A tồn tại để loại bỏ. Output trên config hợp lệ không đổi,
    # và điều đó được kiểm bằng L3/CHECK-108A1-15.
    wb = openpyxl.load_workbook(raw, read_only=True, data_only=True)
    inputs = []
    for offset, row in enumerate(wb.active.iter_rows(min_row=6, values_only=True)):
        if not row[1] or len(row) < 13 or not row[12]:
            continue
        raw_value = norm(row[12])
        when = row[0].date() if hasattr(row[0], "date") else row[0]
        when = when if isinstance(when, date) else None
        result = mapper.resolve(raw_value, when)
        inputs.append(
            MappingInput(
                source_file=raw.name,
                source_row=6 + offset,
                employee_raw=raw_value,
                when=when,
                normalized=result.normalized,
                group=result.group,
            )
        )
    wb.close()

    stats = collect_mapping_stats(inputs, mapper)
    mapped = stats.mapped
    unmapped = stats.unmapped
    groups = stats.groups
    dataset_start = stats.dataset_start
    dataset_end = stats.dataset_end

    verdict = evaluate_raw_mapping(stats)

    span = (f"{dataset_start} .. {dataset_end}"
            if dataset_start and dataset_end else "không xác định")
    print("=" * 78)
    print("CHECK-108A1-15 — Employee mapping trên file thô toàn công ty")
    print("=" * 78)
    print(f"  Phạm vi ngày của dữ liệu  : {span}")
    print(f"  VERIFIED  — dòng map được : {sum(mapped.values())}")
    for name, count in mapped.most_common():
        print(f"      {name:10} {groups[name]:16} {count:6}")
    print(f"  UNMAPPED  — dòng chưa map : {sum(unmapped.values())}  -> Review "
          "Queue, KHÔNG nhận tỉ lệ nào (DEC-127 §8)")
    for name, count in unmapped.most_common():
        print(f"      {name:34} {count}")
    print()

    print("  HARD FAILURE (invariant — quyết định exit code):")
    if verdict.hard_failures:
        for line in verdict.hard_failures:
            print(f"      {line}")
    else:
        print(f"      F1 group referential integrity : OK "
              f"({len(mapper.master.group_codes)} group khai báo)")
        print("      F3 không đụng độ prefix cùng kỳ : OK")
        print("      F5 mapping không rỗng           : OK")

    print()
    print("  WARNING / REVIEW SIGNAL (chẩn đoán — KHÔNG làm exit non-zero):")
    if verdict.warnings:
        for line in verdict.warnings:
            print(f"      {line}")
    elif not mapped:
        print("      (bỏ qua — không có nhân viên nào map được, xem HARD FAILURE)")
    else:
        largest_unmapped = max(unmapped.values(), default=0)
        smallest_name, smallest = min(mapped.items(), key=lambda kv: kv[1])
        print("      F2 nhân viên hiệu lực không có dòng : không có")
        print(f"      F4 biên khối lượng                  : unmapped lớn nhất "
              f"{largest_unmapped} < {smallest_name} {smallest}")

    if verdict.info:
        print()
        print("  INFO (vắng mặt hợp lệ theo effective window):")
        for line in verdict.info:
            print(f"      {line}")
    print()
    return len(verdict.hard_failures)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--raw", type=Path, default=None)
    args = parser.parse_args()

    classifier = LeadSourceClassifier.from_yaml(CONFIG / "lead_source.yaml")
    resolver = ConversionSchemeResolver.from_yaml(CONFIG / "conversion_rates.yaml")
    failures = reconcile_summary(args.workbook, classifier, resolver)

    if args.raw and args.raw.exists():
        mapper = EmployeeMapper.from_yaml(CONFIG / "employees.yaml")
        failures += reconcile_raw(args.raw, mapper)

    print("KẾT QUẢ:", "PASS" if failures == 0 else f"FAIL ({failures} ô lệch)")
    sys.exit(1 if failures else 0)


if __name__ == "__main__":
    main()
