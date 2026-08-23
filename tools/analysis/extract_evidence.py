"""Extract verifiable facts from the two sample workbooks.

Every number quoted in docs/analysis/ comes from this script, so a reviewer can
re-run it and check the claim rather than trusting the prose.

Usage:
    python tools/analysis/extract_evidence.py \
        --raw data/samples/So_chi_tiet_ban_hang.xlsx \
        --report data/samples/Bao_cao_Kinh_doanh_2026.xlsx \
        --out docs/analysis/_evidence
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

RAW_HEADER_ROW = 4
RAW_FIRST_DATA_ROW = 6

# Raw column positions, 0-based, as they appear in "SỔ CHI TIẾT BÁN HÀNG".
RAW = {
    "date": 0, "order_id": 1, "note": 2, "product": 3, "customer_code": 4,
    "customer": 5, "address": 6, "phone": 7, "qty": 8, "unit_price": 9,
    "sales": 10, "discount": 11, "employee": 12, "shipper": 13,
    "trip_pay": 14, "imei": 15, "profit": 16,
}

ADS_KEYWORDS = ("ADS",)


def normalize_text(value) -> str:
    """NFC-normalize and collapse whitespace, per spec section 13."""
    if value is None:
        return ""
    text = unicodedata.normalize("NFC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def contains_ads(note: str) -> bool:
    upper = normalize_text(note).upper()
    return any(keyword in upper for keyword in ADS_KEYWORDS)


def read_raw(path: Path) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [
        row
        for row in sheet.iter_rows(min_row=RAW_FIRST_DATA_ROW, values_only=True)
        if row[RAW["order_id"]]
    ]
    workbook.close()
    return rows


def analyse_raw(rows: list[tuple]) -> dict:
    orders: dict[str, list[tuple]] = defaultdict(list)
    for row in rows:
        orders[str(row[RAW["order_id"]]).strip()].append(row)

    dates = [row[RAW["date"]] for row in rows if row[RAW["date"]]]
    ads_orders = [
        order_id
        for order_id, lines in orders.items()
        if any(contains_ads(line[RAW["note"]]) for line in lines)
    ]

    lines_per_order = Counter(len(lines) for lines in orders.values())
    products = Counter(normalize_text(row[RAW["product"]]) for row in rows)
    non_product = {
        name: count
        for name, count in products.items()
        if re.search(
            r"chi phí vận chuyển|công lắp đặt|chênh vat|chiết khấu|voucher|phí ",
            name,
            re.IGNORECASE,
        )
    }

    return {
        "line_count": len(rows),
        "distinct_order_count": len(orders),
        "date_min": str(min(dates)),
        "date_max": str(max(dates)),
        "employees": Counter(
            normalize_text(row[RAW["employee"]]) for row in rows
        ).most_common(),
        "orders_matching_ads_rule": len(ads_orders),
        "ads_sample_order_ids": ads_orders[:20],
        "lines_per_order_distribution": dict(sorted(lines_per_order.items())),
        "multi_line_order_count": sum(
            count for size, count in lines_per_order.items() if size > 1
        ),
        "rows_missing_employee": sum(
            1 for row in rows if not normalize_text(row[RAW["employee"]])
        ),
        "rows_missing_qty": sum(1 for row in rows if not row[RAW["qty"]]),
        "rows_negative_profit": sum(
            1 for row in rows if isinstance(row[RAW["profit"]], (int, float))
            and row[RAW["profit"]] < 0
        ),
        "rows_with_discount": sum(
            1 for row in rows if isinstance(row[RAW["discount"]], (int, float))
            and row[RAW["discount"]] != 0
        ),
        "non_product_line_types": dict(
            sorted(non_product.items(), key=lambda kv: -kv[1])[:30]
        ),
    }


def aggregate_raw_by_month_employee(rows: list[tuple]) -> dict:
    buckets: dict[str, dict] = defaultdict(
        lambda: {"orders": set(), "qty": 0.0, "sales": 0.0, "profit": 0.0}
    )
    for row in rows:
        date = row[RAW["date"]]
        if not date:
            continue
        key = f"{date.month:02d}.{date.year}|{normalize_text(row[RAW['employee']])}"
        bucket = buckets[key]
        bucket["orders"].add(str(row[RAW["order_id"]]).strip())
        bucket["qty"] += row[RAW["qty"]] or 0
        bucket["sales"] += row[RAW["sales"]] or 0
        bucket["profit"] += row[RAW["profit"]] or 0
    return {
        key: {
            "orders": len(value["orders"]),
            "qty": value["qty"],
            "sales_thousands": round(value["sales"] / 1000, 0),
            "profit_thousands": round(value["profit"] / 1000, 0),
        }
        for key, value in sorted(buckets.items())
    }


def analyse_report(path: Path) -> dict:
    formulas = openpyxl.load_workbook(path, data_only=False)
    values = openpyxl.load_workbook(path, data_only=True)

    layouts: dict[str, list[str]] = defaultdict(list)
    header_row1: dict[str, dict[str, str]] = {}
    purchase_source = Counter()
    adjustment_notes = Counter()
    manual_price_overrides = 0
    price_cells_total = 0

    for name in formulas.sheetnames:
        if name.startswith(("Summary", "DataChart")):
            continue
        sheet = formulas[name]
        header = tuple(
            normalize_text(cell.value) for cell in sheet[2]
        )
        layouts[header].append(name)
        header_row1.setdefault(
            header,
            {
                cell.coordinate: str(cell.value)
                for cell in sheet[1]
                if cell.value is not None
            },
        )

        # Channel sheets (Nội thành / Gia dụng) have no "Nơi nhập" column and
        # therefore sit one column to the left of the individual layout.
        offset = 0 if header[2] == "Nơi nhập" else -1
        for index in range(3, sheet.max_row + 1):
            if offset == 0:
                purchase_source[normalize_text(sheet.cell(index, 3).value)] += 1
            note = sheet.cell(index, 10 + offset).value
            if note:
                adjustment_notes[normalize_text(note)] += 1
            actual_price = sheet.cell(index, 12 + offset).value
            if actual_price is not None:
                price_cells_total += 1
                if not (isinstance(actual_price, str) and actual_price.startswith("=")):
                    manual_price_overrides += 1

    conversion = []
    for summary_name in ("Summary 2026", "Summary 2025"):
        if summary_name not in formulas.sheetnames:
            continue
        sheet = formulas[summary_name]
        period = None
        for index in range(1, sheet.max_row + 1):
            first = sheet.cell(index, 1).value
            if hasattr(first, "year"):
                period = f"{first.month:02d}.{first.year}"
            employee = sheet.cell(index, 2).value
            formula = sheet.cell(index, 6).value
            if isinstance(formula, str) and formula.startswith("=") and employee:
                conversion.append(
                    {
                        "summary": summary_name,
                        "period": period,
                        "employee": normalize_text(employee),
                        "converted_revenue_formula": formula,
                        "target": sheet.cell(index, 13).value,
                        "bonus_formula": sheet.cell(index, 15).value,
                        "workdays": sheet.cell(index, 16).value,
                    }
                )

    sheet_totals = {}
    for name in values.sheetnames:
        if name.startswith(("Summary", "DataChart")):
            continue
        sheet = values[name]
        header = normalize_text(formulas[name].cell(2, 3).value)
        offset = 0 if header == "Nơi nhập" else -1
        sheet_totals[name] = {
            "orders": sheet.cell(1, 2).value,
            "products": sheet.cell(1, 5 + offset).value,
            "sales": sheet.cell(1, 8 + offset).value,
            "profit": sheet.cell(1, 9 + offset).value,
            "gross_profit": sheet.cell(1, 13 + offset).value,
        }

    result = {
        "sheet_count": len(formulas.sheetnames),
        "layouts": [
            {
                "sheets": names,
                "sheet_count": len(names),
                "header": [column for column in header if column],
                "row1_formulas": header_row1[header],
            }
            for header, names in layouts.items()
        ],
        "conversion_rows": conversion,
        "purchase_source_top": purchase_source.most_common(25),
        "adjustment_notes_top": adjustment_notes.most_common(25),
        "manual_price_overrides": manual_price_overrides,
        "price_cells_total": price_cells_total,
        "sheet_totals": sheet_totals,
    }
    formulas.close()
    values.close()
    return result


def count_ads_in_workbook(path: Path) -> int:
    """Count cells whose text contains the ADS keyword, values and formulas."""
    workbook = openpyxl.load_workbook(path, data_only=False)
    hits = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and contains_ads(cell):
                    hits += 1
    workbook.close()
    return hits


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)

    raw_rows = read_raw(args.raw)
    evidence = {
        "raw": analyse_raw(raw_rows),
        "raw_by_month_employee": aggregate_raw_by_month_employee(raw_rows),
        "report": analyse_report(args.report),
        "ads_keyword_cell_hits": {
            "raw_workbook": count_ads_in_workbook(args.raw),
            "report_workbook": count_ads_in_workbook(args.report),
        },
    }

    target = args.out / "evidence.json"
    target.write_text(
        json.dumps(evidence, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(f"wrote {target}")
    print(f"raw lines                 : {evidence['raw']['line_count']}")
    print(f"distinct orders           : {evidence['raw']['distinct_order_count']}")
    print(f"orders matching ADS rule  : {evidence['raw']['orders_matching_ads_rule']}")
    print(f"ADS cells in raw workbook : {evidence['ads_keyword_cell_hits']['raw_workbook']}")
    print(f"ADS cells in report       : {evidence['ads_keyword_cell_hits']['report_workbook']}")
    print(f"report layouts            : {len(evidence['report']['layouts'])}")


if __name__ == "__main__":
    main()
