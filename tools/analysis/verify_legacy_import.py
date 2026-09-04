"""Đối chiếu từng ô Excel legacy với bản ghi trong history DB (CHECK-PRA001-01).

Đây là bằng chứng E1 cho fidelity trên FILE THẬT: script đọc lại workbook
bằng openpyxl ``data_only=True`` và so từng ô số của các sheet
REQUIRED_IMPORT — ``Summary 2026`` và ``DataChart 2026`` — với giá trị đã
lưu. Script KHÔNG sửa gì — nó chỉ đọc.

``Summary 2025`` là OPTIONAL_IMPORT (`DEC-177` làm rõ `DEC-169`): những
dòng của nó ĐÃ nhập được vẫn phải khớp giá trị nguồn như mọi dòng khác, còn
dòng chưa phân loại được thì được ĐẾM và in ra
(``SUMMARY_OPTIONAL_UNIMPORTED``) chứ không làm script trượt. Lý do bất đối
xứng này: thiếu một dòng REQUIRED nghĩa là số production sai; thiếu một dòng
OPTIONAL nghĩa là còn một phần lịch sử chưa đọc được — cần nói ra, không cần
chặn.

Fidelity ở đây gồm HAI phần, không phải một:

1. **VALUE MATCH** — ô đã nhập có đúng giá trị nguồn không.
2. **SOURCE COVERAGE** — có dòng nguồn nào mang giá trị nghiệp vụ mà KHÔNG
   được nhập không.

Chỉ có (1) là không đủ, và đó là lỗi thật của bản trước repair: duyệt từ
DB → Excel nên một dòng (hoặc cả một sheet) bị bỏ qua khi import vẫn cho ra
``matched>0 mismatched=0``. Một báo cáo thiếu hẳn một kỳ mà vẫn "khớp
100%" là loại bằng chứng tệ hơn không có bằng chứng. Vì vậy vòng lặp
Summary duyệt từ EXCEL → DB (FIND-PRA001-R01). Việc thu hẹp phạm vi về
sheet REQUIRED_IMPORT (DEC-169) KHÔNG làm yếu guard này: trong phạm vi đó,
mọi dòng nguồn có giá trị nghiệp vụ vẫn bị soi từ phía Excel.

Cách chạy (workbook thật KHÔNG commit vào repo):

    HISTORY_DATABASE_URL=... python3 -m tools.analysis.verify_legacy_import \\
        "/duong/dan/Bao cao Kinh doanh 2026.xlsx" [LEG-...]

Thoát 0 khi không lệch giá trị, không thiếu dòng nguồn, và không có
dòng REFERENCE_ONLY nào bị persist; ngược lại thoát 1.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

import tools.db as history_db
from app.legacy.models import SUMMARY_COLUMN_FIELDS
from app.legacy.parser import (
    DATACHART_DAY_COLUMNS, DATACHART_FIRST_ROW, DATACHART_SHEET,
    SUMMARY_IMPORT_SHEETS, SUMMARY_OPTIONAL_SHEETS,
    row_has_business_values,
)
from app.web import history_store


@dataclass
class VerificationResult:
    """Kết quả đối chiếu: khớp giá trị VÀ phủ hết dòng nguồn."""

    matched: int = 0
    mismatches: list[str] = field(default_factory=list)
    summary_source_rows_with_values: int = 0
    summary_imported_rows: int = 0
    summary_unaccounted_rows: list[str] = field(default_factory=list)
    # Dòng nguồn của sheet OPTIONAL_IMPORT có số nhưng chưa nhập được
    # (`DEC-177`). ĐO LƯỜNG, không phải lỗi: nó là thước đo phần lịch sử
    # còn thiếu, và là con số Owner cần thấy để biết còn phải cấp gì.
    optional_unimported_rows: list[str] = field(default_factory=list)
    optional_imported_rows: int = 0

    @property
    def ok(self) -> bool:
        return (
            not self.mismatches
            and not self.summary_unaccounted_rows
        )


def _decimal(value):
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return None
    return Decimal(str(value))


def verify(workbook_path: Path, repository, import_id=None) -> VerificationResult:
    sheets = load_workbook(workbook_path, data_only=True)
    result = VerificationResult()

    stored_summary = {}
    all_summary_sheets = SUMMARY_IMPORT_SHEETS + SUMMARY_OPTIONAL_SHEETS
    for year in {int(name.split()[-1]) for name in all_summary_sheets}:
        for row in repository.query_summary(year, import_id=import_id):
            stored_summary[(row["sheet_name"], row["sheet_row"])] = row
    result.summary_imported_rows = sum(
        1 for name, _ in stored_summary if name in SUMMARY_IMPORT_SHEETS)
    result.optional_imported_rows = sum(
        1 for name, _ in stored_summary if name in SUMMARY_OPTIONAL_SHEETS)

    # Duyệt từ EXCEL → DB: đây là chiều DUY NHẤT phát hiện được dòng nguồn
    # bị bỏ qua khi import. Chiều ngược lại (DB → Excel) không bao giờ thấy
    # một dòng chưa từng được nhập.
    for sheet_name in all_summary_sheets:
        optional = sheet_name in SUMMARY_OPTIONAL_SHEETS
        if sheet_name not in sheets.sheetnames:
            if not optional:
                result.mismatches.append(f"{sheet_name}: thiếu sheet trong workbook nguồn")
            continue
        sheet = sheets[sheet_name]
        for sheet_row in range(1, (sheet.max_row or 0) + 1):
            source_values = {
                column: sheet[f"{column}{sheet_row}"].value
                for column in SUMMARY_COLUMN_FIELDS
            }
            if not row_has_business_values(source_values):
                continue
            if not optional:
                result.summary_source_rows_with_values += 1
            row = stored_summary.get((sheet_name, sheet_row))
            if row is None:
                # Dòng nguồn có số nhưng KHÔNG có bản ghi tương ứng. Với
                # sheet REQUIRED đây là lỗi; với OPTIONAL đây là số đo.
                target = (result.optional_unimported_rows if optional
                          else result.summary_unaccounted_rows)
                target.append(f"{sheet_name}!{sheet_row}")
                continue
            for column, field_name in SUMMARY_COLUMN_FIELDS.items():
                expected = _decimal(source_values[column])
                actual = row[field_name]
                if actual == expected:
                    result.matched += 1
                else:
                    result.mismatches.append(
                        f"{sheet_name}!{column}{sheet_row}: excel={expected} db={actual}"
                    )

    # Bản ghi tồn tại trong DB nhưng dòng nguồn không còn giá trị nghiệp vụ
    # nào — cũng là lệch, theo chiều ngược lại.
    for (sheet_name, sheet_row) in stored_summary:
        if sheet_name not in sheets.sheetnames:
            result.mismatches.append(f"{sheet_name}!{sheet_row}: sheet nguồn không tồn tại")
            continue
        sheet = sheets[sheet_name]
        source_values = {
            column: sheet[f"{column}{sheet_row}"].value for column in SUMMARY_COLUMN_FIELDS
        }
        if not row_has_business_values(source_values):
            result.mismatches.append(
                f"{sheet_name}!{sheet_row}: có bản ghi trong DB nhưng dòng nguồn không có giá trị"
            )

    chart = sheets[DATACHART_SHEET]
    year = int(DATACHART_SHEET.split()[-1])
    for month in range(1, 13):
        sheet_row = DATACHART_FIRST_ROW + month - 1
        stored = {row["day"]: row["sales_vnd"]
                  for row in repository.query_daily(year, month, import_id=import_id)}
        # DataChart đã duyệt từ phía Excel sẵn: ô có số mà DB không có sẽ ra
        # expected != None, actual = None → tính là lệch, không im lặng.
        for day, column in enumerate(DATACHART_DAY_COLUMNS, start=1):
            expected = _decimal(chart[f"{column}{sheet_row}"].value)
            actual = stored.get(day)
            if actual == expected:
                result.matched += 1
            else:
                result.mismatches.append(
                    f"{DATACHART_SHEET}!{column}{sheet_row}: excel={expected} db={actual}"
                )
    return result


def main(argv: list[str]) -> int:
    if not argv:
        print(__doc__)
        return 2
    workbook_path = Path(argv[0])
    import_id = argv[1] if len(argv) > 1 else None
    repository = history_store.build(engine=history_db.build_engine())
    result = verify(workbook_path, repository, import_id)
    for line in result.mismatches:
        print(f"MISMATCH {line}")
    for line in result.summary_unaccounted_rows:
        print(f"UNACCOUNTED {line}")
    for line in result.optional_unimported_rows:
        print(f"OPTIONAL_UNIMPORTED {line}")
    print(f"SUMMARY_SOURCE_ROWS_WITH_VALUES = {result.summary_source_rows_with_values}")
    print(f"SUMMARY_IMPORTED_ROWS           = {result.summary_imported_rows}")
    print(f"SUMMARY_UNACCOUNTED_ROWS        = {len(result.summary_unaccounted_rows)}")
    print(f"SUMMARY_OPTIONAL_IMPORTED       = {result.optional_imported_rows}")
    print(
        "SUMMARY_OPTIONAL_UNIMPORTED     = "
        f"{len(result.optional_unimported_rows)}"
    )
    print(f"matched={result.matched} mismatched={len(result.mismatches)}")
    # Thiếu dòng nguồn là FAIL, ngang hàng với lệch giá trị: một bản nhập
    # thiếu kỳ mà báo "khớp 100%" còn tệ hơn không đối chiếu.
    return 0 if result.ok else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
